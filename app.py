import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import re
import base64
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

from dotenv import load_dotenv, find_dotenv

# Try to load .env from same directory as app.py first, then search upward
_app_dir  = os.path.dirname(os.path.abspath(__file__))
_env_path = os.path.join(_app_dir, ".env")
if os.path.exists(_env_path):
    load_dotenv(_env_path, override=True)
else:
    # Fallback: search parent directories
    _found = find_dotenv(usecwd=True)
    if _found:
        load_dotenv(_found, override=True)

# ==============================
# FEATURE STORE PATHS
# ==============================
FEATURE_STORE_PATH    = "feature_store/claims_json"
AUDIT_LOG_PATH        = "feature_store/audit_log.json"
HASH_STORE_PATH       = "feature_store/hash_store.json"
JSON_EXPORT_TABLE_PATH = "feature_store/json_export_table.json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)
os.makedirs("feature_store", exist_ok=True)

# ==============================
# AUDIT LOG HELPERS
# ==============================
def _load_audit_log() -> list:
    if os.path.exists(AUDIT_LOG_PATH):
        try:
            with open(AUDIT_LOG_PATH) as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_audit_log(log: list):
    with open(AUDIT_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2)

def _append_audit(entry: dict):
    log = _load_audit_log()
    log.append(entry)
    _save_audit_log(log)

# ==============================
# HASH STORE HELPERS
# ==============================
def _load_hash_store() -> dict:
    if os.path.exists(HASH_STORE_PATH):
        try:
            with open(HASH_STORE_PATH) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def _save_hash_store(store: dict):
    with open(HASH_STORE_PATH, "w") as f:
        json.dump(store, f, indent=2)

def _compute_file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

def _compute_sheet_sha256(file_path: str, sheet_name: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "rb") as f:
            return hashlib.sha256(f.read()).hexdigest()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    h  = hashlib.sha256()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            h.update(str(cell).encode("utf-8"))
    wb.close()
    return h.hexdigest()

# ==============================
# JSON EXPORT TABLE HELPERS
# ==============================
def _load_json_export_table() -> list:
    if os.path.exists(JSON_EXPORT_TABLE_PATH):
        try:
            with open(JSON_EXPORT_TABLE_PATH) as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_json_export_table(table: list):
    with open(JSON_EXPORT_TABLE_PATH, "w") as f:
        json.dump(table, f, indent=2)

def _append_json_export(entry: dict):
    table = _load_json_export_table()
    for existing in table:
        if (existing.get("filename") == entry.get("filename") and
                existing.get("sheet")    == entry.get("sheet") and
                existing.get("type")     == entry.get("type")):
            existing.update(entry)
            _save_json_export_table(table)
            return
    table.append(entry)
    _save_json_export_table(table)

# ==============================
# LOGO HELPER
# ==============================
def _load_logo_b64() -> str:
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "valuemomentum_logo.png"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "valuemomentum_logo.png"),
        "valuemomentum_logo.png",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, "rb") as f:
                return base64.b64encode(f.read()).decode()
    return ""

_LOGO_B64 = _load_logo_b64()

def _logo_img_tag(height: int = 38) -> str:
    if not _LOGO_B64:
        return ""
    return (
        f'<img src="data:image/png;base64,{_LOGO_B64}" '
        f'style="height:{height}px;margin-right:14px;vertical-align:middle;'
        f'border-radius:4px;background:#1e1e2e;padding:3px 6px;" />'
    )

# ==============================
# YAML CONFIG LOADER
# ==============================
CONFIG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")

def _parse_yaml_simple(text: str) -> dict:
    def _cast(v: str):
        v = v.strip()
        if not v or v.lower() in ("null", "~", ""): return None
        if v.lower() == "true":  return True
        if v.lower() == "false": return False
        try:    return int(v)
        except: pass
        try:    return float(v)
        except: pass
        return v.strip('"').strip("'")

    lines   = text.splitlines()
    root    = {}
    stack   = [(0, root)]
    cur_key = None
    for raw in lines:
        if not raw.strip() or raw.strip().startswith("#"): continue
        indent = len(raw) - len(raw.lstrip())
        line   = raw.strip()
        while len(stack) > 1 and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1]
        if line.startswith("- "):
            val = line[2:].strip()
            if cur_key and isinstance(parent, dict):
                if not isinstance(parent.get(cur_key), list):
                    parent[cur_key] = []
                parent[cur_key].append(_cast(val))
        elif ":" in line:
            parts = line.split(":", 1)
            key   = parts[0].strip().strip('"').strip("'")
            val   = parts[1].strip() if len(parts) > 1 else ""
            if " #" in val: val = val[:val.index(" #")].strip()
            cur_key = key
            if val:
                parent[key] = _cast(val)
            else:
                parent[key] = {}
                stack.append((indent + 2, parent[key]))
    return root

def load_schema_config(schema_filename: str) -> dict | None:
    path = os.path.join(CONFIG_DIR, schema_filename)
    if not os.path.exists(path): return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _parse_yaml_simple(f.read())
    except Exception:
        return None

def _merge_schema_from_config(hardcoded: dict, cfg: dict | None) -> dict:
    if not cfg: return hardcoded
    merged = dict(hardcoded)
    schema_block = cfg.get("schema", {})
    for k in ("version", "description"):
        if schema_block.get(k): merged[k] = schema_block[k]
    if cfg.get("required_fields"):
        rf = cfg["required_fields"]
        if isinstance(rf, dict): rf = list(rf.keys())
        if isinstance(rf, list): merged["required_fields"] = [str(f) for f in rf if f]
    if cfg.get("accepted_fields"):
        af = cfg["accepted_fields"]
        if isinstance(af, dict): af = list(af.keys())
        if isinstance(af, list): merged["accepted_fields"] = [str(f) for f in af if f]
    if cfg.get("field_aliases") and isinstance(cfg["field_aliases"], dict):
        aliases = {}
        for field, vals in cfg["field_aliases"].items():
            if isinstance(vals, list):  aliases[field] = [str(v) for v in vals if v]
            elif isinstance(vals, str): aliases[field] = [vals]
        if aliases: merged["field_aliases"] = aliases
    conf_block = cfg.get("confidence", {})
    if isinstance(conf_block, dict):
        if conf_block.get("field_thresholds") and isinstance(conf_block["field_thresholds"], dict):
            merged["field_thresholds"] = {k: int(v) for k, v in conf_block["field_thresholds"].items() if v is not None}
        if conf_block.get("global_threshold") is not None:
            merged["config_threshold"] = int(conf_block["global_threshold"])
    if cfg.get("export") and isinstance(cfg["export"], dict):
        merged["export_config"] = cfg["export"]
    return merged

_CONFIG_LOAD_STATUS = {}

def _load_all_configs(hardcoded_schemas: dict) -> dict:
    filemap = {"Guidewire": "guidewire.yaml", "Duck Creek": "duck_creek.yaml"}
    result  = {}
    for name, schema in hardcoded_schemas.items():
        fname = filemap.get(name)
        cfg   = load_schema_config(fname) if fname else None
        result[name] = _merge_schema_from_config(schema, cfg)
        _CONFIG_LOAD_STATUS[name] = {
            "file": fname, "loaded": cfg is not None,
            "path": os.path.join(CONFIG_DIR, fname) if fname else "",
        }
    return result

# ==============================
# UNICODE NORMALIZER
# ==============================
_DASH_TABLE = str.maketrans({
    '\u2013': '-', '\u2014': '-', '\u2012': '-', '\u2015': '-',
    '\u2212': '-', '\ufe58': '-', '\ufe63': '-', '\uff0d': '-',
    '\u2018': "'", '\u2019': "'", '\u201c': '"', '\u201d': '"',
    '\u00a0': ' ', '\u202f': ' ',
})

def normalize_str(s: str) -> str:
    if not s: return s
    return s.translate(_DASH_TABLE)

# ==============================
# SCHEMA DEFINITIONS
# ==============================
_HARDCODED_SCHEMAS = {
    "Guidewire": {
        "color": "#4f9cf9", "icon": "🔵", "css_cls": "guide",
        "version": "ClaimCenter 10.x",
        "description": "Guidewire ClaimCenter 10.x compatible format",
        "date_format": "YYYY-MM-DD",
        "amount_format": "decimal",
        "status_values": ["open", "closed", "pending", "reopened", "denied", "submitted", "draft"],
        "required_fields": [
            "Claim Number", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Status", "Line of Business", "Policy Number",
        ],
        "accepted_fields": [
            "Claim Number", "Claimant Name", "Loss Date", "Date Reported",
            "Total Incurred", "Total Paid", "Reserve", "Indemnity Paid",
            "Medical Paid", "Expense Paid", "Status", "Line of Business",
            "Policy Number", "Policy Period Start", "Policy Period End",
            "Carrier", "Insured Name", "Description of Loss",
            "Cause of Loss", "Litigation Flag", "Adjuster Name",
            "Adjuster Phone", "Branch Code", "Department Code",
            "Coverage Type", "Deductible", "Subrogation Amount",
            "Recovery Amount", "Open/Closed", "Reopen Date",
            "Last Activity Date", "Days Lost", "State", "Notes",
            "Job Title", "Body Part", "Vehicle ID", "At Fault",
            "Building Damage", "Contents Damage", "Business Interruption Loss",
            "Net Paid", "Services Involved", "Location",
        ],
        "field_aliases": {
            "Claim Number":      ["claim_id","claim number","claim no","claim#","claimid","claim ref","claim #"],
            "Claimant Name":     ["claimant name","claimant","insured name","name","injured party","employee name","driver name"],
            "Loss Date":         ["date of loss","loss date","loss dt","date of accident","incident date","date of injury","injury date","date of incident"],
            "Date Reported":     ["date reported","reported date","report date"],
            "Total Incurred":    ["total incurred","incurred","total incurred amount"],
            "Total Paid":        ["total paid","amount paid","paid amount","net paid"],
            "Reserve":           ["reserve","outstanding reserve","case reserve"],
            "Indemnity Paid":    ["indemnity paid","indemnity","wage loss paid","ttd paid","bi paid"],
            "Medical Paid":      ["medical paid","medical","med paid"],
            "Expense Paid":      ["expense paid","expense","legal expense","defense costs"],
            "Status":            ["status","claim status","open/closed"],
            "Line of Business":  ["line of business","lob","coverage line"],
            "Policy Number":     ["policy number","policy no","policy#","policy id","policy :","policy #"],
            "Insured Name":      ["insured name","insured","employer name"],
            "Description of Loss": ["description of loss","loss description","description","narrative","nature of injury","nature of claim","type of loss","cause of loss"],
            "Cause of Loss":     ["cause of loss","cause","type of loss","peril","nature of injury","nature of claim"],
            "Adjuster Name":     ["adjuster name","adjuster","examiner"],
            "Coverage Type":     ["coverage","coverage type"],
            "Deductible":        ["deductible","deductible amount"],
            "Days Lost":         ["days lost","days of disability","lost days","disability days","days missed"],
            "Job Title":         ["job title","occupation","position","employee title"],
            "Body Part":         ["body part","body part injured","part of body"],
            "Vehicle ID":        ["vehicle id","vehicle","unit number","vin"],
            "At Fault":          ["at fault","fault","liable","at-fault"],
            "Building Damage":   ["building damage","structure damage","building loss"],
            "Contents Damage":   ["contents damage","contents loss","stock loss"],
            "Business Interruption Loss": ["bi loss","business interruption","business income loss"],
            "Net Paid":          ["net paid","pd paid","property damage paid","net claim payment"],
            "Services Involved": ["services involved","professional services","service type"],
            "Location":          ["location","property location","site","premises"],
        },
    },
    "Duck Creek": {
        "color": "#f5c842", "icon": "🟡", "css_cls": "duck",
        "version": "Claims 6.x",
        "description": "Duck Creek Claims 6.x transaction format",
        "date_format": "MM/DD/YYYY",
        "amount_format": "decimal",
        "status_values": ["Open", "Closed", "Pending", "Reopen", "Denied", "Settled"],
        "required_fields": [
            "Claim Id", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Policy Number", "Claim Status",
        ],
        "accepted_fields": [
            "Claim Id", "Transaction Id", "Claimant Name", "Loss Date",
            "Date Reported", "Total Incurred", "Total Paid", "Reserve",
            "Indemnity Paid", "Medical Paid", "Expense Paid",
            "Policy Number", "Policy Effective Date", "Policy Expiry Date",
            "Claim Status", "Cause of Loss", "Description of Loss",
            "Insured Name", "Carrier Name", "Line of Business",
            "Adjuster Id", "Adjuster Name", "Office Code",
            "Jurisdiction", "State Code", "Deductible Amount",
            "Subrogation Flag", "Recovery Amount", "Litigation Flag",
            "Date Closed", "Date Reopened", "Last Updated Date", "Days Lost",
            "Notes", "Job Title", "Body Part", "Vehicle ID", "At Fault",
            "Building Damage", "Contents Damage", "Business Interruption Loss",
            "Net Paid", "Services Involved", "Property Location", "Coverage",
        ],
        "field_aliases": {
            "Claim Id":          ["claim_id","claim number","claim no","claim#","claimid","claim ref","claim #"],
            "Claimant Name":     ["claimant name","claimant","insured name","name","injured party","employee name","driver name"],
            "Loss Date":         ["date of loss","loss date","loss dt","date of accident","incident date","date of injury","injury date","date of incident"],
            "Date Reported":     ["date reported","reported date","report date"],
            "Total Incurred":    ["total incurred","incurred","total incurred amount"],
            "Total Paid":        ["total paid","amount paid","paid amount"],
            "Reserve":           ["reserve","outstanding reserve","case reserve"],
            "Indemnity Paid":    ["indemnity paid","indemnity","wage loss paid","ttd paid","bi paid"],
            "Medical Paid":      ["medical paid","medical","med paid"],
            "Expense Paid":      ["expense paid","expense","legal expense","defense costs"],
            "Claim Status":      ["status","claim status","open/closed"],
            "Line of Business":  ["line of business","lob","coverage line"],
            "Policy Number":     ["policy number","policy no","policy#","policy id","policy :","policy #"],
            "Insured Name":      ["insured name","insured","employer name"],
            "Description of Loss": ["description of loss","loss description","description","narrative","nature of injury","nature of claim","type of loss"],
            "Cause of Loss":     ["cause of loss","cause","type of loss","peril","nature of injury","nature of claim"],
            "Carrier Name":      ["carrier","carrier name","insurance company"],
            "Deductible Amount": ["deductible","deductible amount"],
            "Jurisdiction":      ["state","state code","jurisdiction"],
            "Days Lost":         ["days lost","days of disability","lost days","disability days","days missed"],
            "Job Title":         ["job title","occupation","position","employee title"],
            "Body Part":         ["body part","body part injured","part of body"],
            "Vehicle ID":        ["vehicle id","vehicle","unit number","vin"],
            "At Fault":          ["at fault","fault","liable","at-fault"],
            "Building Damage":   ["building damage","structure damage","building loss"],
            "Contents Damage":   ["contents damage","contents loss","stock loss"],
            "Business Interruption Loss": ["bi loss","business interruption","business income loss"],
            "Net Paid":          ["net paid","pd paid","property damage paid","net claim payment"],
            "Services Involved": ["services involved","professional services","service type"],
            "Property Location": ["location","property location","site","premises"],
            "Coverage":          ["coverage","coverage type","type of coverage","subject to $50k sir","within policy limits","coverage under review"],
        },
    },
}

SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)

# ==============================
# SCHEMA-AWARE RULE-BASED NORMALIZER
# Zero cost — no external calls, no LLM.
# Handles: dates → schema format, amounts → 2dp decimal,
#          status → schema enum, names → Title Case,
#          states → 2-letter uppercase, booleans → true/false
# ==============================

# Comprehensive date format map keyed by schema date_format string
_DATE_FMT_MAP = {
    "YYYY-MM-DD": "%Y-%m-%d",
    "MM/DD/YYYY": "%m/%d/%Y",
    "MM-DD-YYYY": "%m-%d-%Y",
    "DD/MM/YYYY": "%d/%m/%Y",
    "DD-MM-YYYY": "%d-%m-%Y",
}

def _parse_date_flexible(val: str):
    """Parse any recognizable date string → datetime.date, or None."""
    val = val.strip().replace("\u2013", "-").replace("\u2014", "-")
    for fmt in [
        "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%d-%m-%Y",
        "%m/%d/%y", "%m-%d-%y", "%d/%m/%y", "%Y/%m/%d",
        "%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%b %d %Y",
        "%d %B %Y", "%d %b %Y",
    ]:
        try:
            return datetime.datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None

def _format_date_for_schema(val: str, schema_name: str) -> str:
    """Reformat date string to the schema's required format. Returns original if unparseable."""
    if not val or not val.strip():
        return val
    parsed = _parse_date_flexible(val.strip())
    if parsed is None:
        return val
    fmt_key    = SCHEMAS.get(schema_name, {}).get("date_format", "YYYY-MM-DD")
    strfmt     = _DATE_FMT_MAP.get(fmt_key, "%Y-%m-%d")
    return parsed.strftime(strfmt)

def _format_amount_for_schema(val: str) -> str:
    """Normalize currency/amount → plain 2dp decimal float string. Handles $, commas, (negatives)."""
    if not val or not val.strip():
        return val
    s = val.strip()
    is_neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    try:
        num = float(s)
        if is_neg:
            num = -abs(num)
        return f"{num:.2f}"
    except ValueError:
        return val

def _format_status_for_schema(val: str, schema_name: str) -> str:
    """Map any status synonym → schema's exact accepted enum value (correct casing)."""
    if not val or not val.strip():
        return val
    allowed = SCHEMAS.get(schema_name, {}).get("status_values", [])
    v_lower = val.strip().lower()
    # Exact match first (case-insensitive)
    for sv in allowed:
        if sv.lower() == v_lower:
            return sv
    # Synonym mapping
    _synonyms = {
        "open":     ["open","active","in progress","inprogress","new","opened"],
        "closed":   ["closed","close","completed","done","finalized","resolved"],
        "pending":  ["pending","pend","on hold","hold","waiting","review"],
        "reopened": ["reopen","reopened","re-opened","re open"],
        "reopen":   ["reopen","reopened","re-opened","re open"],
        "denied":   ["denied","deny","rejected","reject","declined"],
        "submitted":["submitted","submit","filed"],
        "draft":    ["draft","drafting"],
        "settled":  ["settled","settlement"],
    }
    for canonical, syns in _synonyms.items():
        if v_lower in syns or v_lower == canonical:
            for sv in allowed:
                if sv.lower() == canonical:
                    return sv
    return val  # Unrecognized — return as-is

def _format_name_for_schema(val: str) -> str:
    """Title case names but preserve all-uppercase abbreviations (LLC, INC, LLP, NA, etc.)."""
    if not val or not val.strip():
        return val
    # Abbreviations that should stay uppercase
    _keep_upper = {"llc","inc","lp","llp","na","n/a","dba","usa","us","uk","ltd","corp","co"}
    parts = val.strip().split()
    result = []
    for part in parts:
        if part.lower() in _keep_upper:
            result.append(part.upper())
        else:
            result.append(part.capitalize())
    return " ".join(result)

def _format_state_for_schema(val: str) -> str:
    if not val or not val.strip(): return val
    v = val.strip()
    if len(v) == 2: return v.upper()
    _st = {
        "alabama":"AL","alaska":"AK","arizona":"AZ","arkansas":"AR","california":"CA",
        "colorado":"CO","connecticut":"CT","delaware":"DE","florida":"FL","georgia":"GA",
        "hawaii":"HI","idaho":"ID","illinois":"IL","indiana":"IN","iowa":"IA","kansas":"KS",
        "kentucky":"KY","louisiana":"LA","maine":"ME","maryland":"MD","massachusetts":"MA",
        "michigan":"MI","minnesota":"MN","mississippi":"MS","missouri":"MO","montana":"MT",
        "nebraska":"NE","nevada":"NV","new hampshire":"NH","new jersey":"NJ","new mexico":"NM",
        "new york":"NY","north carolina":"NC","north dakota":"ND","ohio":"OH","oklahoma":"OK",
        "oregon":"OR","pennsylvania":"PA","rhode island":"RI","south carolina":"SC",
        "south dakota":"SD","tennessee":"TN","texas":"TX","utah":"UT","vermont":"VT",
        "virginia":"VA","washington":"WA","west virginia":"WV","wisconsin":"WI","wyoming":"WY",
        "district of columbia":"DC",
    }
    return _st.get(v.lower(), v)

def _format_boolean_for_schema(val: str) -> str:
    if not val or not val.strip(): return val
    v = val.strip().lower()
    if v in ("yes","y","true","1","x"):  return "true"
    if v in ("no","n","false","0"):      return "false"
    return val

# Field-type classifier patterns
_DATE_FIELD_PATTERNS   = re.compile(r"date|loss\s*dt|incident\s*dt|injury\s*dt|accident\s*dt|closed\s*dt|reopen\s*dt|updated\s*dt|effective\s*dt|expiry\s*dt", re.IGNORECASE)
_AMOUNT_FIELD_PATTERNS = re.compile(r"incurred|paid|reserve|amount|deductible|recovery|subrogation|damage|bi\s*loss|interruption", re.IGNORECASE)
_STATUS_FIELD_PATTERNS = re.compile(r"\bstatus\b|open.?closed", re.IGNORECASE)
_NAME_FIELD_PATTERNS   = re.compile(r"(claimant|insured|adjuster|employer|driver|employee|injured)\s*(name)?$|^name$", re.IGNORECASE)
_STATE_FIELD_PATTERNS  = re.compile(r"\bstate\b|\bjurisdiction\b|\bstate\s*code\b", re.IGNORECASE)
_BOOL_FIELD_PATTERNS   = re.compile(r"\bat\s*fault\b|\blitigation\s*flag\b|\bsubrogation\s*flag\b", re.IGNORECASE)


def auto_normalize_field(field_name: str, value: str, schema_name: str) -> str:
    """
    Normalize one field value to the active schema's format.
    Pure Python — zero cost, zero latency, called on every field when schema activates.
    """
    if not value or not str(value).strip():
        return value
    v  = str(value).strip()
    fn = field_name.strip()
    if _DATE_FIELD_PATTERNS.search(fn):   return _format_date_for_schema(v, schema_name)
    if _AMOUNT_FIELD_PATTERNS.search(fn): return _format_amount_for_schema(v)
    if _STATUS_FIELD_PATTERNS.search(fn): return _format_status_for_schema(v, schema_name)
    if _NAME_FIELD_PATTERNS.search(fn):   return _format_name_for_schema(v)
    if _STATE_FIELD_PATTERNS.search(fn):  return _format_state_for_schema(v)
    if _BOOL_FIELD_PATTERNS.search(fn):   return _format_boolean_for_schema(v)
    return v


def auto_normalize_claim(claim_data: dict, schema_name: str) -> dict:
    """Returns {field: new_value} for every field that changed after normalization."""
    changes = {}
    for field, info in claim_data.items():
        original = info.get("modified") or info.get("value", "")
        if not original:
            continue
        normalized = auto_normalize_field(field, str(original), schema_name)
        if normalized != original:
            changes[field] = normalized
    return changes


def auto_normalize_on_schema_activate(data: list, schema_name: str, selected_sheet: str):
    """
    Batch-normalize all claims when schema activates or sheet loads.
    Writes to session_state only if user has not manually edited the field yet.
    Updates in-memory claim["modified"] so exports pick up normalized values.
    """
    for i, claim in enumerate(data):
        claim_id = detect_claim_id(claim, i)
        changes  = auto_normalize_claim(claim, schema_name)
        for field, new_val in changes.items():
            mk_schema = f"mod_{selected_sheet}_{claim_id}_schema_{field}"
            mk_plain  = f"mod_{selected_sheet}_{claim_id}_{field}"
            if mk_schema not in st.session_state:
                st.session_state[mk_schema] = new_val
            if mk_plain not in st.session_state:
                st.session_state[mk_plain] = new_val
            if field in claim:
                claim[field]["modified"] = new_val


# ==============================
# LLM ENGINE — Single focused task: Cause of Loss extraction
#
# WHY LLM HERE AND NOWHERE ELSE:
# Cause of Loss comes from free-text narrative descriptions like
# "claimant slipped on wet floor in aisle 4 and fractured wrist".
# Rules cannot reliably map arbitrary prose to a standard taxonomy.
# LLM reads the narrative once per claim and outputs a standardised label.
#
# COST CONTROL:
# • One call per claim, only when description fields exist
# • Only when Cause of Loss is currently empty
# • Result cached in session_state — never called twice for same claim
# • ~120 tokens output, temperature=0 (deterministic, cheap)
# • Silent fallback on any error — never breaks the UI
# ==============================

import urllib.request
import urllib.error

def _llm_available() -> bool:
    return (bool(os.environ.get("OPENAI_API_KEY","").strip()) and
            bool(os.environ.get("OPENAI_DEPLOYMENT_ENDPOINT","").strip()))

def _llm_call(prompt: str, max_tokens: int = 150) -> str:
    """Internal Azure OpenAI call. Raises on failure. Not referenced in UI."""
    endpoint = os.environ.get("OPENAI_DEPLOYMENT_ENDPOINT","").rstrip("/")
    api_key  = os.environ.get("OPENAI_API_KEY","")
    api_ver  = os.environ.get("OPENAI_API_VERSION","2024-12-01-preview")
    model    = os.environ.get("OPENAI_DEPLOYMENT_NAME","gpt-4o-mini")
    url      = f"{endpoint}/openai/deployments/{model}/chat/completions?api-version={api_ver}"
    payload  = json.dumps({
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": max_tokens,
        "temperature": 0.0,
    }).encode()
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type":"application/json","api-key":api_key},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read().decode())["choices"][0]["message"]["content"]

# Cause-of-loss taxonomies by line of business
_COL_TAXONOMY_GENERAL = [
    "Slip and Fall","Trip and Fall","Vehicle Collision","Rear-End Collision",
    "Fire - Electrical","Fire - Arson","Fire - Unknown","Theft","Burglary",
    "Vandalism","Water Damage - Flood","Water Damage - Pipe","Wind / Hail",
    "Lightning","Earthquake","Product Liability","Professional Error",
    "Medical Malpractice","Assault / Battery","Equipment Failure",
    "Explosion","Repetitive Stress Injury","Strain / Sprain",
    "Animal Bite","Falling Object","Other",
]
_COL_TAXONOMY_PROF_LIABILITY = [
    "Negligent Advice","Unsuitable Product Recommendation","Misrepresentation",
    "Failure to Execute Instructions","Breach of Fiduciary Duty",
    "Conflict of Interest","Unauthorized Trading","Due Diligence Failure",
    "Inadequate Risk Assessment","Failure to Disclose","Elder Financial Abuse",
    "Fraud / Intentional Misrepresentation","Omission of Material Fact",
    "Portfolio Mismanagement","Failure to Supervise","Other - Professional Liability",
]
_COL_TAXONOMY_WORKERS_COMP = [
    "Strain / Sprain - Back","Strain / Sprain - Shoulder","Strain / Sprain - Knee",
    "Laceration","Fracture","Contusion / Bruising","Slip and Fall",
    "Trip and Fall","Repetitive Stress Injury","Occupational Disease",
    "Heat / Chemical Exposure","Electrical Injury","Crush Injury",
    "Vehicle Accident - Work Related","Falling Object","Equipment Failure",
    "Other - Workers Comp",
]
_COL_TAXONOMY_AUTO = [
    "Rear-End Collision","Side-Impact Collision","Head-On Collision",
    "Single Vehicle Accident","Pedestrian Strike","Uninsured Motorist",
    "Hit and Run","Backing Accident","Vehicle Rollover","Weather-Related Collision",
    "DUI-Related Accident","Vehicle Theft","Vandalism","Other - Auto",
]
_COL_TAXONOMY_PROPERTY = [
    "Fire - Electrical","Fire - Arson","Fire - Unknown","Water Damage - Pipe",
    "Water Damage - Flood","Wind / Hail Damage","Lightning Strike",
    "Earthquake","Theft / Burglary","Vandalism","Collapse",
    "Equipment Breakdown","Mold","Sinkhole","Other - Property",
]

def _pick_taxonomy(sheet_name: str, claim_text: str) -> list:
    """Pick the right taxonomy based on sheet name and claim text context."""
    s = (sheet_name + " " + claim_text).lower()
    if any(x in s for x in ["prof liab","professional liab","e&o","errors","fiduciary","advisory","malpractice"]):
        return _COL_TAXONOMY_PROF_LIABILITY
    if any(x in s for x in ["workers comp","work comp","wc loss","injury","strain","sprain","lacerat"]):
        return _COL_TAXONOMY_WORKERS_COMP
    if any(x in s for x in ["auto","vehicle","collision","motor","driving","fleet"]):
        return _COL_TAXONOMY_AUTO
    if any(x in s for x in ["property","building","premises","fire","water damage","theft","hail"]):
        return _COL_TAXONOMY_PROPERTY
    return _COL_TAXONOMY_GENERAL

def _llm_extract_cause_of_loss(description_text: str, sheet_name: str = "") -> dict:
    """One focused LLM call: narrative → standardised cause of loss + one-sentence summary."""
    taxonomy = _pick_taxonomy(sheet_name, description_text)
    taxonomy_str = "\n".join(f"- {t}" for t in taxonomy)
    prompt = (
        "You are an insurance claims analyst. Read the loss description and:\n"
        "1. Pick the SINGLE best-matching cause of loss from the taxonomy below.\n"
        "   You MUST choose exactly one label from the list — do not invent new labels.\n"
        "   If nothing fits well, choose 'Other' or the most similar 'Other - ...' entry.\n"
        "2. Write one factual plain-English sentence summarizing what the claimant alleges happened.\n"
        "   Base the summary ONLY on the text provided — do not invent facts.\n\n"
        f"TAXONOMY (choose one exactly as written):\n{taxonomy_str}\n\n"
        f"LOSS DESCRIPTION:\n{description_text[:1200]}\n\n"
        'Reply ONLY with valid JSON, no markdown, no explanation:\n'
        '{"cause_of_loss": "<exact taxonomy label>", "summary": "<one sentence>"}'
    )
    raw = _llm_call(prompt, max_tokens=120)
    raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
    return json.loads(raw)


def enrich_claim_cause_of_loss(claim_data: dict, claim_id: str, selected_sheet: str):
    """
    Silently enriches Cause of Loss for a claim when:
      - LLM is configured in .env
      - Claim has description/narrative text
      - Cause of Loss field is currently empty or is the raw description text
      - This claim hasn't been enriched yet this session
    Returns True if enrichment ran (so caller can trigger rerun to surface result).
    Errors are caught silently — never disrupts the UI.
    """
    if not _llm_available():
        return False
    cache_key = f"_col_enriched_{selected_sheet}_{claim_id}"
    if st.session_state.get(cache_key):
        return False  # Already processed this session

    # Find description/narrative fields — broad match covering all loss run column naming conventions
    _DESC_PAT = re.compile(
        r"desc|narr|detail|note|comment|descript|loss\s*desc|allegation"
        r"|nature|nature.of.claim|nature.of.loss|nature.of.injury"
        r"|type.of.claim|type.of.loss|cause.of.loss"
        r"|services.involved|service.type"
        r"|injury.type|accident.type|incident.type"
        r"|claim.type|peril|event.type",
        re.IGNORECASE
    )
    # Explicitly EXCLUDE fields that are dates, IDs, amounts, or names
    _EXCL_PAT = re.compile(
        r"^(date|claim.?id|claim.?num|claimant|adjuster|insured|"
        r"paid|incurred|reserve|total|cost|amount|number|id|#)$",
        re.IGNORECASE
    )

    desc_keys = []
    for k in claim_data:
        if _DESC_PAT.search(k) and not _EXCL_PAT.match(k.strip()):
            val = str(claim_data[k].get("modified") or claim_data[k].get("value","")).strip()
            # Must have actual text — not just a date, number, or very short value
            if val and len(val) > 6 and not re.match(r"^\d{1,4}[-/]\d{1,2}[-/]\d{2,4}$", val):
                desc_keys.append(k)

    if not desc_keys:
        st.session_state[cache_key] = True
        return False

    texts = [str(claim_data[k].get("modified") or claim_data[k].get("value","")).strip()
             for k in desc_keys if claim_data[k].get("modified") or claim_data[k].get("value","")]
    combined = " | ".join(t for t in texts if t and len(t) > 4)
    if not combined:
        st.session_state[cache_key] = True
        return False

    # Skip ONLY if a dedicated Cause of Loss field already has a short standardised value
    # (not raw narrative text — those are often long sentences, not taxonomy labels)
    for k, info in claim_data.items():
        if re.search(r"^cause\s*of\s*loss$|^cause_of_loss$", k.strip(), re.IGNORECASE):
            existing = str(info.get("modified") or info.get("value","")).strip()
            # Only skip if it looks like a taxonomy label (short, not a sentence)
            if existing and len(existing) > 3 and len(existing) < 60 and "." not in existing:
                st.session_state[cache_key] = True
                return False

    try:
        result  = _llm_extract_cause_of_loss(combined, sheet_name=selected_sheet)
        col_val = result.get("cause_of_loss","")
        summary = result.get("summary","")

        # Validate: reject result if it doesn't match the chosen taxonomy at all
        taxonomy = _pick_taxonomy(selected_sheet, combined)
        if col_val and col_val not in taxonomy:
            # LLM invented a label outside the taxonomy — use "Other"
            col_val = "Other"

        if col_val:
            # Write to both schema and plain mod keys for Cause of Loss field
            for field_key in ["Cause of Loss","Cause Of Loss","cause_of_loss","Cause_of_Loss"]:
                mk_s = f"mod_{selected_sheet}_{claim_id}_schema_{field_key}"
                mk_p = f"mod_{selected_sheet}_{claim_id}_{field_key}"
                for mk in (mk_s, mk_p):
                    st.session_state[mk] = col_val
            # Update in-memory for export
            for k, info in claim_data.items():
                if re.search(r"cause.?of.?loss", k, re.IGNORECASE):
                    claim_data[k]["modified"] = col_val

        if summary:
            st.session_state[f"_col_summary_{selected_sheet}_{claim_id}"] = summary

        # Store which fields were used as source (for transparency)
        st.session_state[f"_col_source_fields_{selected_sheet}_{claim_id}"] = desc_keys

        st.session_state[cache_key] = True
        _append_audit({"event":"LLM_CAUSE_ENRICHED","timestamp":datetime.datetime.now().isoformat(),
                       "sheet":selected_sheet,"claim_id":claim_id,
                       "source_fields":desc_keys,"input_text":combined[:200],
                       "cause_of_loss":col_val,"summary":summary})
        return True  # Signal to caller: rerun needed
    except Exception:
        st.session_state[cache_key] = True  # Prevent retry on error
        return False


# ==============================
# SETTINGS DIALOG
# ==============================
@st.dialog("Settings", width="large")
def show_settings_dialog():
    global SCHEMAS
    st.markdown("### Configuration")
    st.markdown("---")
    st.markdown("#### Confidence Settings")
    use_conf = st.checkbox("Enable configurable confidence threshold", value=st.session_state.get("use_conf_threshold",True), key="use_conf_toggle")
    st.session_state["use_conf_threshold"] = use_conf
    if use_conf:
        conf = st.slider("Confidence threshold", 0, 100, value=st.session_state.get("conf_threshold",80), step=5, format="%d%%")
        st.session_state["conf_threshold"] = conf
        bar_color = "#22c55e" if conf>=70 else "#f59e0b" if conf>=40 else "#ef4444"
        level_txt = "High confidence — minimal manual review needed" if conf>=70 else "Medium — review flagged fields carefully" if conf>=40 else "Low — most fields will require manual review"
        st.markdown(f"<div class='conf-bar-wrap'><div class='conf-bar-fill' style='width:{conf}%;background:{bar_color};'></div></div><div style='color:{bar_color};font-size:12px;margin-top:5px;'>{level_txt}</div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("#### Export Schema")
    active_schema = st.session_state.get("active_schema",None)
    for schema_name, schema_def in SCHEMAS.items():
        is_active  = active_schema == schema_name
        border_col = schema_def["color"] if is_active else "#2a2a3e"
        bg_col     = "#1a1a2e" if is_active else "#16161e"
        active_tag = f"<span style='font-size:10px;color:{schema_def['color']};margin-left:8px;font-weight:bold;'>● ACTIVE</span>" if is_active else ""
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}",[]))
        st.markdown(f"<div style='background:{bg_col};border:1px solid {border_col};border-radius:8px;padding:12px 14px;margin-bottom:4px;'><div style='display:flex;align-items:center;'><span style='font-size:var(--sz-body);font-weight:700;color:var(--t0);font-family:var(--font);'>{schema_def['icon']} {schema_name}</span><span style='font-size:var(--sz-sm);color:var(--t3);margin-left:8px;font-family:var(--font);'>{schema_def['version']}</span>{active_tag}</div><div style='font-size:var(--sz-sm);color:var(--t2);margin-top:4px;font-family:var(--font);'>{schema_def['description']}</div></div>", unsafe_allow_html=True)
        bc1,bc2,bc3 = st.columns([1,1,1])
        with bc1:
            if st.button("✓ Deactivate" if is_active else "Activate", key=f"activate_{schema_name}", use_container_width=True):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(f"Custom Fields ({custom_count})", key=f"custom_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:10px;font-family:var(--font);'>Config directory: <code>{CONFIG_DIR}</code></div>", unsafe_allow_html=True)
    for schema_name, status in _CONFIG_LOAD_STATUS.items():
        sc     = SCHEMAS.get(schema_name,{})
        col_st = sc.get("color","#64748b")
        badge  = ("<span style='background:#0f2d1f;border:1px solid #22c55e;border-radius:4px;padding:1px 7px;font-size:10px;color:#22c55e;'>✓ Loaded</span>" if status["loaded"] else "<span style='background:#2d0f0f;border:1px solid #ef4444;border-radius:4px;padding:1px 7px;font-size:10px;color:#ef4444;'>✗ Not found — using defaults</span>")
        st.markdown(f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:6px;padding:10px 14px;margin-bottom:6px;'><div style='display:flex;align-items:center;gap:10px;'><span style='color:{col_st};font-weight:700;font-size:var(--sz-body);font-family:var(--font);'>{sc.get('icon','')} {schema_name}</span>{badge}</div><div style='font-size:var(--sz-xs);color:var(--t3);margin-top:4px;font-family:var(--font);'>📄 {status['file']}</div></div>", unsafe_allow_html=True)
    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.session_state["sheet_cache"] = {}
        st.success("✅ Configs reloaded")
        st.rerun()
    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("Reset Defaults", use_container_width=True):
            st.session_state["conf_threshold"]     = 80
            st.session_state["use_conf_threshold"] = True
            st.session_state["active_schema"]      = None
            for s in SCHEMAS: st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ==============================
# SCHEMA FIELD MANAGER DIALOG
# ==============================
@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name):
    schema     = SCHEMAS[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state: st.session_state[custom_key] = []
    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:14px;font-family:var(--font);'>{schema['description']}</div>", unsafe_allow_html=True)
    tab_req, tab_accepted, tab_custom = st.tabs(["Mandatory Fields","All Accepted Fields","My Custom Fields"])
    with tab_req:
        pills = "".join(f"<span class='field-pill field-pill-required'>✓ {f}</span>" for f in schema["required_fields"])
        st.markdown(f"<div style='margin:12px 0;'>{pills}</div>", unsafe_allow_html=True)
    with tab_accepted:
        optional  = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(f"<span class='field-pill field-pill-required'>✓ {f}</span>" for f in schema["required_fields"])
        opt_pills = "".join(f"<span class='field-pill'>{f}</span>" for f in optional)
        st.markdown(f"<div style='margin:12px 0;'><b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;text-transform:uppercase;'>MANDATORY</b><br><div style='margin-top:6px;'>{req_pills}</div></div><div style='margin:12px 0;'><b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;text-transform:uppercase;'>OPTIONAL</b><br><div style='margin-top:6px;'>{opt_pills}</div></div>", unsafe_allow_html=True)
    with tab_custom:
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]
        if available:
            sel_col, add_col = st.columns([4,1])
            with sel_col:
                chosen = st.selectbox("Pick field", ["— select a field —"]+available, key=f"new_field_sel_{schema_name}", label_visibility="collapsed")
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}", use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()
        if not custom_fields:
            st.markdown("<div style='color:var(--t2);font-size:var(--sz-body);padding:10px 0;font-family:var(--font);'>No optional fields added yet.</div>", unsafe_allow_html=True)
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5,1])
                with cf1:
                    cls = "field-pill-required" if cf in schema["required_fields"] else "field-pill-custom"
                    st.markdown(f"<span class='field-pill {cls}'>{'✓' if cf in schema['required_fields'] else '+'} {cf}</span>", unsafe_allow_html=True)
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}", use_container_width=True):
                        st.session_state[custom_key].pop(idx); st.rerun()
            st.markdown("---")
            if st.button("Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []; st.rerun()
        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:8px;padding:10px 16px;'><span style='color:var(--t2);font-size:var(--sz-body);font-family:var(--font);'>Mandatory: <b style='color:var(--blue);'>{len(schema['required_fields'])}</b> &nbsp;|&nbsp; Custom: <b style='color:var(--green);'>{len(custom_fields)}</b> &nbsp;|&nbsp; Total: <b style='color:var(--t0);'>{total}</b></span></div>", unsafe_allow_html=True)


# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Loss Run Parser", page_icon="🛡️")
if "focus_field" not in st.session_state: st.session_state.focus_field = None

# ==============================
# GLOBAL CSS — DARK THEME
# ==============================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;600&family=Source+Sans+3:wght@300;400;600;700&display=swap');

:root {
    --bg:       #0d0d14;
    --surface:  #12121c;
    --s0:       #17172a;
    --s1:       #1e1e32;
    --s2:       #252540;
    --b0:       #2a2a45;
    --b1:       #343458;

    --blue:     #4f9cf9;
    --blue-lt:  #7ab8ff;
    --blue-dk:  #2563eb;
    --blue-g:   rgba(79,156,249,0.08);
    --blue-mid: rgba(79,156,249,0.15);
    --green:    #34d399;
    --green-lt: #6ee7b7;
    --green-g:  rgba(52,211,153,0.08);
    --yellow:   #f5c842;
    --yellow-lt:#fde68a;
    --yellow-g: rgba(245,200,66,0.08);
    --red:      #f87171;
    --red-lt:   #fca5a5;
    --red-g:    rgba(248,113,113,0.08);
    --purple:   #a78bfa;
    --purple-g: rgba(167,139,250,0.08);

    --t0:  #ffffff;
    --t1:  #f0efff;
    --t2:  #e8e7ff;
    --t3:  #c8c7f0;
    --t4:  #a0a0c8;

    --font-head: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    --font: 'Source Sans 3', 'Source Sans Pro', 'Segoe UI', system-ui, sans-serif;
    --mono: 'JetBrains Mono', 'Cascadia Code', 'Consolas', monospace;

    --sz-xl:   16px;
    --sz-lg:   15px;
    --sz-body: 14px;
    --sz-sm:   13px;
    --sz-xs:   12px;

    --shadow-sm: 0 1px 6px rgba(0,0,0,0.5), 0 0 1px rgba(79,156,249,0.08);
    --shadow:    0 4px 20px rgba(0,0,0,0.6), 0 0 2px rgba(79,156,249,0.10);
    --shadow-lg: 0 8px 40px rgba(0,0,0,0.7), 0 0 4px rgba(79,156,249,0.12);

    --radius-sm: 4px;
    --radius:    7px;
    --radius-lg: 11px;
    --radius-xl: 16px;

    --fw-h: 700;
    --fw-s: 600;
    --fw-b: 400;
}

*, *::before, *::after { box-sizing: border-box; }

.stApp {
    background: var(--bg) !important;
    color: var(--t1);
    font-family: var(--font);
    font-size: var(--sz-body);
    line-height: 1.6;
    -webkit-font-smoothing: antialiased;
}

.stApp::before {
    content: '';
    position: fixed;
    inset: 0;
    background: repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,0,0,0.03) 2px,rgba(0,0,0,0.03) 4px);
    pointer-events: none;
    z-index: 0;
}

h1,h2,h3,h4 { font-family: var(--font-head) !important; color: var(--t0) !important; }
h1 { font-size: var(--sz-xl) !important; font-weight: 700 !important; letter-spacing: -0.3px; }
h2 { font-size: var(--sz-lg) !important; font-weight: 700 !important; }
h3 { font-size: var(--sz-body) !important; font-weight: 600 !important; }

p, li { font-size: var(--sz-body) !important; color: var(--t0); font-family: var(--font) !important; }

code {
    background: var(--s1) !important;
    border: 1px solid var(--b0) !important;
    border-radius: var(--radius-sm) !important;
    padding: 2px 6px !important;
    font-family: var(--mono) !important;
    font-size: var(--sz-xs) !important;
    color: var(--blue) !important;
}

/* Hide Streamlit's default top chrome */
#MainMenu { visibility: hidden; }
header[data-testid="stHeader"] { display: none !important; }
div[data-testid="stToolbar"] { display: none !important; }
div[data-testid="stDecoration"] { display: none !important; }
footer { display: none !important; }

.block-container {
    padding-top: 0 !important;
    padding-left: 1.5rem !important;
    padding-right: 1.5rem !important;
    max-width: 100% !important;
}

/* Top bar title row */
.topbar-title-row {
    display: flex;
    align-items: center;
    padding: 10px 0 6px 0;
}
.topbar-divider {
    border: none;
    border-top: 1px solid var(--b0);
    margin: 4px 0 20px 0;
}
.navbar-title-wrap {
    display: flex;
    flex-direction: column;
    justify-content: center;
    gap: 1px;
}
.navbar-title {
    font-size: 15px;
    font-weight: 700;
    color: var(--t0);
    font-family: var(--font-head);
    letter-spacing: -0.2px;
    white-space: nowrap;
    line-height: 1.2;
}
.navbar-subtitle {
    font-size: 10px;
    font-weight: 400;
    color: var(--t3);
    font-family: var(--mono);
    letter-spacing: 0.4px;
    white-space: nowrap;
}
.navbar-divider {
    width: 1px;
    height: 24px;
    background: var(--b1);
    margin: 0 2px;
}
.navbar-schema-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    border-radius: 6px;
    padding: 5px 13px;
    font-size: 12px;
    font-weight: 700;
    font-family: var(--mono);
    border: 1px solid;
    white-space: nowrap;
    letter-spacing: 0.2px;
}
.navbar-right {
    display: flex;
    align-items: center;
    gap: 10px;
}

.main-title { display: none; }

.section-lbl {
    font-size: var(--sz-xs);
    font-weight: 600;
    color: var(--t2);
    text-transform: uppercase;
    letter-spacing: 2px;
    font-family: var(--mono);
    margin-bottom: 10px;
    margin-top: 2px;
}

.file-card {
    background: var(--surface);
    border: 1px solid var(--b0);
    border-top: 2px solid var(--blue);
    border-radius: var(--radius-xl);
    margin-bottom: 18px;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.file-card-header {
    background: var(--s0);
    border-bottom: 1px solid var(--b0);
    padding: 13px 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.file-card-title {
    font-size: var(--sz-body);
    font-weight: 700;
    color: var(--t0);
    display: flex;
    align-items: center;
    gap: 10px;
    font-family: var(--font-head);
}
.file-badge {
    font-family: var(--mono);
    font-size: 10px;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 20px;
    text-transform: uppercase;
    letter-spacing: 1px;
}
.badge-unique    { background: var(--green-g); color: var(--green); border: 1px solid rgba(52,211,153,0.3); }
.badge-duplicate { background: var(--yellow-g); color: var(--yellow); border: 1px solid rgba(245,200,66,0.3); }
.file-card-body {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    padding: 16px 20px;
    gap: 16px;
    background: var(--surface);
}
.file-stat { display: flex; flex-direction: column; gap: 5px; }
.file-stat-lbl { font-size: var(--sz-xs); font-weight: 600; color: var(--t2); text-transform: uppercase; letter-spacing: 1.4px; font-family: var(--mono); }
.file-stat-val { font-size: var(--sz-body); font-weight: 600; color: var(--t0); font-family: var(--font); }
.file-stat-val.accent { color: var(--blue); font-weight: 700; }
.file-stat-val.mono-sm { font-size: var(--sz-xs); color: var(--t2); letter-spacing: 0.3px; word-break: break-all; font-weight: 400; font-family: var(--mono); }
.file-card-sheets { padding: 10px 20px 14px; border-top: 1px solid var(--b0); background: var(--s0); }
.sheet-pill-sm { display: inline-block; background: var(--s1); border: 1px solid var(--b0); border-radius: 4px; padding: 3px 10px; font-family: var(--mono); font-size: var(--sz-xs); color: var(--t1); margin: 3px 4px 3px 0; }

.sheet-card {
    background: var(--surface);
    border: 1px solid var(--b0);
    border-left: 3px solid var(--blue);
    border-radius: var(--radius-lg);
    margin-bottom: 16px;
    overflow: hidden;
    box-shadow: var(--shadow-sm);
}
.sheet-card-hdr {
    padding: 12px 18px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid var(--b0);
    background: var(--s0);
}
.sheet-card-name { font-size: var(--sz-body); font-weight: 700; color: var(--t0); display: flex; align-items: center; gap: 10px; font-family: var(--font-head); }
.sheet-type-tag { font-family: var(--mono); font-size: 10px; padding: 3px 10px; border-radius: 20px; text-transform: uppercase; letter-spacing: 0.8px; font-weight: 600; background: var(--blue-g); border: 1px solid rgba(79,156,249,0.2); color: var(--blue); }
.sheet-type-tag.unk { background: var(--s1); border-color: var(--b0); color: var(--t3); }
.sheet-stats-grid { display: grid; grid-template-columns: repeat(6, 1fr); padding: 14px 18px; gap: 12px; background: var(--surface); }
.sh-stat { display: flex; flex-direction: column; gap: 5px; }
.sh-stat-lbl { font-size: var(--sz-xs); font-weight: 600; color: var(--t2); text-transform: uppercase; letter-spacing: 1.4px; font-family: var(--mono); }
.sh-stat-val { font-size: var(--sz-body); font-weight: 600; color: var(--t0); font-family: var(--font); }
.sh-stat-val.hi  { color: var(--green); font-weight: 700; }
.sh-stat-val.mid { color: var(--yellow); font-weight: 700; }
.sh-stat-val.hash-sm { font-size: 10px; color: var(--t3); letter-spacing: 0.3px; word-break: break-all; font-weight: 400; font-family: var(--mono); }

.claim-card {
    background: var(--surface);
    border: 1px solid var(--b0);
    border-radius: var(--radius);
    padding: 12px 14px;
    margin-bottom: 6px;
    cursor: pointer;
    transition: border-color 0.15s, box-shadow 0.15s, background 0.15s;
}
.claim-card:hover { border-color: var(--blue); background: var(--blue-g); box-shadow: var(--shadow-sm); }
.selected-card { border-left: 3px solid var(--blue) !important; background: var(--blue-g) !important; box-shadow: 0 0 12px rgba(79,156,249,0.15) !important; }
.status-text { font-size: var(--sz-xs); color: var(--green); margin-top: 4px; font-family: var(--mono); font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; }
.status-progress { font-size: var(--sz-xs); color: var(--yellow); margin-top: 4px; font-family: var(--mono); font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; }

.sheet-title-banner { background: var(--blue-g); border: 1px solid rgba(79,156,249,0.2); border-left: 3px solid var(--blue); border-radius: var(--radius); padding: 11px 16px; margin-bottom: 12px; }
.sheet-title-label { font-size: var(--sz-xs); color: var(--t2); text-transform: uppercase; font-weight: 600; letter-spacing: 1.4px; margin-bottom: 4px; font-family: var(--mono); }
.sheet-title-value { font-size: var(--sz-body); color: var(--t0); font-weight: 700; font-family: var(--font-head); }
.sheet-subtitle-val { font-size: var(--sz-sm); color: var(--t1); margin-top: 3px; font-family: var(--font); }

.mid-header-title { font-size: var(--sz-lg); font-weight: 700; color: var(--t0); margin-bottom: 2px; letter-spacing: -0.2px; font-family: var(--font-head); }
.mid-header-sub { font-size: var(--sz-body); color: var(--t1); margin-top: 2px; margin-bottom: 3px; font-family: var(--font); }
.mid-header-status { font-size: var(--sz-xs); color: var(--green); margin-bottom: 12px; font-family: var(--mono); font-weight: 600; letter-spacing: 0.8px; text-transform: uppercase; }
.incurred-label { font-size: var(--sz-xs); color: var(--t2); margin-bottom: 2px; text-transform: uppercase; letter-spacing: 1.4px; font-weight: 600; font-family: var(--mono); }
.incurred-amount { font-size: var(--sz-lg); font-weight: 700; color: var(--green); margin-top: 2px; margin-bottom: 14px; font-family: var(--font-head); text-shadow: 0 0 20px rgba(52,211,153,0.3); }

.mandatory-asterisk { display: inline-block; font-size: var(--sz-body); color: var(--blue); font-weight: 700; margin-left: 3px; vertical-align: middle; }
.optional-badge { display: inline-block; background: var(--s1); border: 1px solid var(--b0); border-radius: 3px; font-size: var(--sz-xs); color: var(--t1); padding: 0 5px; margin-left: 4px; vertical-align: middle; font-family: var(--mono); }

.custom-field-badge {
    display: inline-block;
    background: var(--purple-g);
    border: 1px solid rgba(167,139,250,0.3);
    border-radius: 3px;
    font-size: 10px;
    color: var(--purple);
    padding: 0 5px;
    margin-left: 4px;
    vertical-align: middle;
    font-family: var(--mono);
}

.add-field-panel {
    background: var(--s0);
    border: 1px dashed var(--b1);
    border-radius: var(--radius-lg);
    padding: 14px 16px;
    margin-top: 16px;
}
.add-field-panel:hover { border-color: var(--purple); }

div[data-baseweb="input"], div[data-baseweb="base-input"], div[data-baseweb="select"] {
    background-color: var(--s1) !important;
    border: 1px solid var(--b1) !important;
    border-radius: var(--radius) !important;
}
div[data-baseweb="input"]:focus-within, div[data-baseweb="base-input"]:focus-within {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(79,156,249,0.12) !important;
}
div[data-baseweb="input"] input {
    color: var(--t0) !important;
    -webkit-text-fill-color: var(--t0) !important;
    background-color: transparent !important;
    font-size: var(--sz-body) !important;
    padding: 8px 12px !important;
    font-family: var(--font) !important;
}
div[data-baseweb="input"]:has(input:disabled),
div[data-baseweb="base-input"]:has(input:disabled) {
    background-color: transparent !important;
    border: none !important;
}
div[data-baseweb="input"] input:disabled {
    color: var(--t0) !important;
    -webkit-text-fill-color: var(--t0) !important;
    cursor: default !important;
    padding-left: 0 !important;
    font-size: var(--sz-body) !important;
}

div[data-testid="stButton"] button {
    background-color: var(--s1) !important;
    color: var(--t0) !important;
    border: 1px solid var(--b1) !important;
    border-radius: var(--radius) !important;
    padding: 7px 14px !important;
    transition: all 0.15s ease !important;
    font-family: var(--font) !important;
    font-size: var(--sz-body) !important;
    font-weight: 600 !important;
}
div[data-testid="stButton"] button:hover {
    border-color: var(--blue) !important;
    color: var(--blue) !important;
    background-color: var(--blue-g) !important;
    box-shadow: 0 0 12px rgba(79,156,249,0.15) !important;
}
div[data-testid="stButton"] button[kind="primary"] {
    background: linear-gradient(135deg, var(--blue-dk) 0%, var(--blue) 100%) !important;
    color: #ffffff !important;
    border-color: transparent !important;
    font-weight: 700 !important;
    box-shadow: 0 2px 12px rgba(79,156,249,0.35) !important;
}
div[data-testid="stButton"] button[kind="primary"]:hover {
    box-shadow: 0 4px 20px rgba(79,156,249,0.50) !important;
    transform: translateY(-1px);
}
div[data-testid="stButton"] button:disabled { opacity: 0.3 !important; }

.navbar-gear-btn {
    background: transparent;
    border: 1px solid rgba(255,255,255,0.18);
    border-radius: 6px;
    color: rgba(255,255,255,0.65);
    font-size: 18px;
    height: 34px;
    width: 34px;
    padding: 0;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: background 0.15s, border-color 0.15s, color 0.15s;
    flex-shrink: 0;
}
.navbar-gear-btn:hover {
    background: rgba(79,156,249,0.12);
    border-color: #4f9cf9;
    color: #4f9cf9;
}


div[role="dialog"] {
    background-color: var(--surface) !important;
    border: 1px solid var(--b0) !important;
    border-radius: var(--radius-xl) !important;
    box-shadow: var(--shadow-lg) !important;
}
div[role="dialog"] * { color: var(--t1) !important; }
div[role="dialog"] h1, div[role="dialog"] h2, div[role="dialog"] h3 { color: var(--t0) !important; }
div[role="dialog"] button {
    background: var(--s1) !important;
    border: 1px solid var(--b1) !important;
    color: var(--t0) !important;
    border-radius: var(--radius) !important;
    font-size: var(--sz-body) !important;
    font-family: var(--font) !important;
}
div[role="dialog"] button:hover { border-color: var(--blue) !important; color: var(--blue) !important; }
div[data-baseweb="select"] > div { font-size: var(--sz-body) !important; font-family: var(--font) !important; }

.conf-bar-wrap { background: var(--s1); border-radius: 4px; height: 5px; width: 100%; margin-top: 4px; overflow: hidden; }
.conf-bar-fill { height: 100%; border-radius: 4px; transition: width 0.4s ease; }

.field-pill { display: inline-block; background: var(--s1); border: 1px solid var(--b0); border-radius: 4px; padding: 4px 12px; font-size: var(--sz-sm); color: var(--t1); margin: 3px 4px; font-family: var(--font); }
.field-pill-required { border-color: rgba(79,156,249,0.35) !important; color: var(--blue) !important; background: var(--blue-g) !important; }
.field-pill-custom { border-color: rgba(52,211,153,0.35) !important; color: var(--green) !important; background: var(--green-g) !important; }

.schema-badge { display: inline-flex; align-items: center; gap: 6px; background: var(--blue-g); border: 1px solid rgba(79,156,249,0.25); border-radius: var(--radius); padding: 3px 12px; font-size: var(--sz-sm); font-weight: 600; color: var(--blue); font-family: var(--mono); }
.schema-badge-duck  { background: var(--yellow-g) !important; border-color: rgba(245,200,66,0.25) !important; color: var(--yellow) !important; }
.schema-badge-guide { background: var(--blue-g) !important; border-color: rgba(79,156,249,0.25) !important; color: var(--blue) !important; }

div[data-baseweb="tab-list"] { background: var(--s0) !important; border-radius: var(--radius) var(--radius) 0 0 !important; border-bottom: 2px solid var(--b0) !important; padding: 0 6px !important; }
div[data-baseweb="tab"] { color: var(--t3) !important; font-family: var(--mono) !important; font-weight: 600 !important; font-size: var(--sz-sm) !important; padding: 11px 18px !important; border-bottom: 2px solid transparent !important; transition: all 0.15s !important; margin-bottom: -2px !important; }
div[data-baseweb="tab"]:hover { color: var(--t1) !important; }
div[data-baseweb="tab"][aria-selected="true"] { color: var(--blue) !important; border-bottom-color: var(--blue) !important; font-weight: 700 !important; }
div[data-baseweb="tab-panel"] { background: var(--surface) !important; border: 1px solid var(--b0) !important; border-top: none !important; border-radius: 0 0 var(--radius) var(--radius) !important; padding: 18px !important; }

.stDataFrame thead th { background: var(--s0) !important; color: var(--blue) !important; font-family: var(--mono) !important; font-size: var(--sz-xs) !important; text-transform: uppercase !important; letter-spacing: 0.9px !important; border-color: var(--b0) !important; font-weight: 600 !important; }
.stDataFrame tbody td { color: var(--t1) !important; font-family: var(--font) !important; font-size: var(--sz-body) !important; border-color: var(--b0) !important; }

div[data-testid="stFileUploader"] { background: var(--s0) !important; border: 2px dashed var(--b1) !important; border-radius: var(--radius-lg) !important; }
div[data-testid="stFileUploader"]:hover { border-color: var(--blue) !important; }

::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--b1); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: var(--blue); }

hr { border-color: var(--b0) !important; margin: 16px 0 !important; }

div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] { display: none !important; }
div[data-testid="stForm"] { border: none !important; padding: 0 !important; }

details { background: var(--s0) !important; border: 1px solid var(--b0) !important; border-radius: var(--radius) !important; margin-bottom: 8px !important; }
details summary { color: var(--t2) !important; font-family: var(--font) !important; font-weight: 600 !important; font-size: var(--sz-body) !important; padding: 10px 14px !important; }

div[data-testid="stAlert"] { font-family: var(--font) !important; font-size: var(--sz-body) !important; border-radius: var(--radius) !important; }
div[data-testid="stMarkdownContainer"] p, div[data-testid="stMarkdownContainer"] li { font-family: var(--font) !important; font-size: var(--sz-body) !important; color: var(--t0) !important; }
div[data-baseweb="select"] span, div[data-baseweb="select"] div { font-family: var(--font) !important; font-size: var(--sz-body) !important; }
div[data-testid="stWidgetLabel"] p, div[data-testid="stWidgetLabel"] label { font-family: var(--font) !important; font-size: var(--sz-sm) !important; font-weight: 600 !important; color: var(--t1) !important; }
div[data-testid="stCheckbox"] label { font-family: var(--font) !important; font-size: var(--sz-body) !important; color: var(--t0) !important; }

.json-live-panel {
    background: var(--s0);
    border: 1px solid var(--b0);
    border-radius: var(--radius-lg);
    padding: 0;
    overflow: hidden;
    margin-top: 12px;
}
.json-live-header {
    background: var(--s1);
    border-bottom: 1px solid var(--b0);
    padding: 8px 14px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.json-live-dot { width: 8px; height: 8px; background: var(--green); border-radius: 50%; animation: pulse-dot 2s infinite; display: inline-block; margin-right: 6px; }
@keyframes pulse-dot { 0%,100%{opacity:1;} 50%{opacity:0.3;} }
.json-live-body { padding: 12px 14px; font-family: var(--mono); font-size: var(--sz-xs); color: var(--t2); max-height: 320px; overflow-y: auto; white-space: pre-wrap; word-break: break-all; line-height: 1.7; }

/* Validation panel */
.validation-badge { display: inline-block; background: var(--green-g); border: 1px solid rgba(52,211,153,0.3); border-radius: 4px; font-size: 10px; color: var(--green); padding: 1px 7px; font-family: var(--mono); }
/* Loss summary panel */
.col-summary-panel { background: var(--s0); border: 1px solid var(--b1); border-left: 3px solid var(--green); border-radius: var(--radius); padding: 7px 12px; margin-top: 4px; margin-bottom: 6px; }
.col-summary-text  { font-size: var(--sz-xs); color: var(--t2); font-family: var(--font); line-height: 1.5; }
</style>
""", unsafe_allow_html=True)

# ==============================
# SHEET HELPERS
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv": return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others

def classify_sheet(rows):
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)
    if "line of business" in text: return "SUMMARY"
    has_claim = any(x in text for x in ["claim number","claim no","claim #","claim id","claim_id","claim ref","claimant","file number","file no"])
    has_loss  = any(x in text for x in ["loss date","date of loss","loss dt","accident date","occurrence date","incident date","date of injury","date of incident","injury date"])
    has_fin   = any(x in text for x in ["incurred","paid","reserve","outstanding","total paid","total incurred","indemnity","expense"])
    if has_claim and (has_loss or has_fin): return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text): return "COMMERCIAL_LOSS_RUN"
    if has_claim: return "LOSS_RUN"
    return "UNKNOWN"

def get_sheet_dimensions(file_path: str, sheet_name: str) -> tuple:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        return len(rows), max((len(r) for r in rows), default=0)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r == 0 or max_c == 0:
        actual_rows = 0
        actual_cols = 0
        for row in ws.iter_rows():
            row_has_data = any(cell.value is not None for cell in row)
            if row_has_data:
                actual_rows += 1
                row_col = max((cell.column for cell in row if cell.value is not None), default=0)
                actual_cols = max(actual_cols, row_col)
        max_r, max_c = actual_rows, actual_cols
    wb.close()
    return max_r, max_c

def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv": return {}
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    merged_info = {}
    for mr in ws.merged_cells.ranges:
        mn_r,mn_c,mx_r,mx_c = mr.min_row,mr.min_col,mr.max_row,mr.max_col
        cell = ws.cell(mn_r,mn_c)
        val  = str(cell.value).strip() if cell.value else ""
        span_cols, span_rows = mx_c-mn_c+1, mx_r-mn_r+1
        region_type = "TITLE" if mn_r<=3 and span_cols>=3 else "HEADER" if span_cols>=2 and span_rows==1 else "DATA"
        merged_info[f"R{mn_r}C{mn_c}"] = {"value":val,"type":region_type,"row_start":mn_r,"col_start":mn_c,"row_end":mx_r,"col_end":mx_c,"span_cols":span_cols,"span_rows":span_rows,"excel_row":mn_r,"excel_col":mn_c}
    wb.close()
    return merged_info

def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()
    totals = {}
    if ext == ".csv":
        with open(file_path,"r",encoding="utf-8-sig") as f: rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows = raw_rows
        wb.close()
    if not rows: return totals
    header_row_index, headers = None, []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and ("date" in row_text or "incurred" in row_text or "paid" in row_text):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}" for j,h in enumerate(row)]
            break
    if header_row_index is None or not headers: return totals
    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index+1:]):
        r_idx = header_row_index+2+r_idx_rel
        if not any(raw_row): continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total","subtotal","grand total","sum","totals"]):
            row_data = {}
            cell_row = cell_rows[header_row_index+1+r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers): continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column if hasattr(cell_row[c_idx_0],'column') else c_idx_0+1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0+1
                if clean_val: row_data[headers[c_idx_0]] = {"value":clean_val,"excel_row":r_idx,"excel_col":real_col}
            if row_data: totals_rows.append(row_data)
    if totals_rows:
        totals["rows"]      = totals_rows
        totals["excel_row"] = totals_rows[0].get(list(totals_rows[0].keys())[0],{}).get("excel_row",9999)
        agg = {}
        for row_data in totals_rows:
            for field,info in row_data.items():
                try:
                    num = float(str(info["value"]).replace(",","").replace("$",""))
                    agg[field] = agg.get(field,0.0) + num
                except: pass
        totals["aggregated"] = {k:round(v,2) for k,v in agg.items()}
    return totals

# ==============================
# CELL FORMATTING
# ==============================
_THEME_COLORS = {0:"FFFFFF",1:"000000",2:"EEECE1",3:"1F497D",4:"4F81BD",5:"C0504D",6:"9BBB59",7:"8064A2",8:"4BACC6",9:"F79646"}

def _resolve_color(color_obj, default="FFFFFF") -> str:
    if color_obj is None: return default
    t = color_obj.type
    if t == "rgb":
        rgb = color_obj.rgb or ""
        if len(rgb)==8 and rgb not in ("00000000","FF000000"): return rgb[2:]
        if len(rgb)==6: return rgb
        return default
    if t == "theme":
        base = _THEME_COLORS.get(color_obj.theme, default)
        tint = color_obj.tint or 0.0
        if tint != 0.0:
            r,g,b = int(base[0:2],16),int(base[2:4],16),int(base[4:6],16)
            if tint>0: r,g,b = int(r+(255-r)*tint),int(g+(255-g)*tint),int(b+(255-b)*tint)
            else: r,g,b = int(r*(1+tint)),int(g*(1+tint)),int(b*(1+tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        return base
    if t == "indexed":
        indexed_map = {0:"000000",1:"FFFFFF",2:"FF0000",3:"00FF00",4:"0000FF",5:"FFFF00",6:"FF00FF",7:"00FFFF",64:"000000",65:"FFFFFF"}
        return indexed_map.get(color_obj.indexed, default)
    return default

def format_cell_value(value) -> str:
    if value is None: return ""
    if isinstance(value, datetime.datetime): return value.strftime("%Y-%m-%d %H:%M:%S") if (value.hour or value.minute) else value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date): return value.strftime("%Y-%m-%d")
    if isinstance(value, bool): return str(value)
    if isinstance(value, int): return str(value)
    if isinstance(value, float):
        if value == int(value): return f"{int(value)}.0"
        formatted = f"{value:.10f}".rstrip('0')
        if '.' not in formatted: formatted += '.0'
        return formatted
    return normalize_str(str(value).strip())

def _apply_date_number_format(dt, nf: str) -> str:
    if not nf or nf.lower() in ("general","@",""): return dt.strftime("%m-%d-%Y")
    fmt = re.sub(r'\[.*?\]','',nf); fmt = re.sub(r'["_*\\]','',fmt)
    result = fmt
    result = re.sub(r'(?i)(?<=h)mm','__MIN__',result); result = re.sub(r'(?i)mm(?=ss)','__MIN__',result)
    def _tok(m):
        tok = m.group(0).lower()
        return {'yyyy':'%Y','yy':'%y','mmmm':'%B','mmm':'%b','mm':'%m','__min__':'%M','m':'%m','dd':'%d','d':'%d','hh':'%H','h':'%H','ss':'%S','s':'%S','am/pm':'%p','a/p':'%p'}.get(tok,m.group(0))
    result = re.sub(r'(?i)yyyy|yy|mmmm|mmm|__min__|mm|dd|hh|ss|am/pm|a/p|d|h|s|m',_tok,result)
    try: return dt.strftime(result)
    except: return dt.strftime("%m-%d-%Y")

def format_cell_value_with_fmt(cell) -> str:
    value = cell.value
    if value is None: return ""
    nf = (cell.number_format or "").strip()
    if isinstance(value,(datetime.datetime,datetime.date)): return _apply_date_number_format(value,nf)
    if isinstance(value,bool): return str(value)
    if isinstance(value,(int,float)):
        decimal_places = None
        if nf and nf.lower() not in ("general","@",""):
            clean_nf = re.sub(r'[$€£¥"_*\\]','',nf)
            is_date_fmt = any(x in clean_nf.lower() for x in ['yy','mm','dd','hh','ss']) and not any(ch in clean_nf for ch in ['0','#'])
            if not is_date_fmt:
                if '.' in clean_nf:
                    after_dot = re.sub(r'\[.*?\]','',clean_nf.split('.')[1])
                    decimal_places = sum(1 for ch in after_dot if ch in '0#')
                else: decimal_places = 0
        if decimal_places is not None:
            fval = float(value)
            return str(int(round(fval))) if decimal_places==0 else f"{fval:.{decimal_places}f}"
        if isinstance(value,int): return str(value)
        fval = float(value); remainder = fval-int(fval)
        if remainder==0.0: return f"{fval:.2f}"
        formatted = f"{fval:.10f}".rstrip('0')
        if '.' not in formatted: formatted += '.00'
        elif len(formatted.split('.')[1]) < 2: formatted = f"{fval:.2f}"
        return formatted
    return normalize_str(str(value).strip())

# ==============================
# EXCEL RENDERER
# ==============================
def _col_px(ws,c,scale=1.0):
    letter = get_column_letter(c); cd = ws.column_dimensions.get(letter)
    w = cd.width if (cd and cd.width and cd.width>0) else 8.43
    return max(20,int(w*8*scale))

def _row_px(ws,r,scale=1.0):
    rd = ws.row_dimensions.get(r); h = rd.height if (rd and rd.height and rd.height>0) else 15.0
    return max(14,int(h*1.5*scale))

def render_excel_sheet(excel_path,sheet_name,scale=1.0):
    wb = openpyxl.load_workbook(excel_path, data_only=True); ws = wb[sheet_name]
    max_col = ws.max_column or 1; max_row = ws.max_row or 1
    col_starts = [0]
    for c in range(1,max_col+1): col_starts.append(col_starts[-1]+_col_px(ws,c,scale))
    row_starts = [0]
    for r in range(1,max_row+1): row_starts.append(row_starts[-1]+_row_px(ws,r,scale))
    img = Image.new("RGB",(col_starts[-1],row_starts[-1]),"white"); draw = ImageDraw.Draw(img,"RGBA")
    merged_master = {}
    for mr in ws.merged_cells.ranges:
        mn_r,mn_c,mx_r,mx_c = mr.min_row,mr.min_col,mr.max_row,mr.max_col
        for rr in range(mn_r,mx_r+1):
            for cc in range(mn_c,mx_c+1): merged_master[(rr,cc)] = (mn_r,mn_c,mx_r,mx_c)
    drawn_merges = set()
    for r in range(1,max_row+1):
        for c in range(1,max_col+1):
            merge_info = merged_master.get((r,c))
            if merge_info:
                mn_r,mn_c,mx_r,mx_c = merge_info
                if (mn_r,mn_c) in drawn_merges: continue
                drawn_merges.add((mn_r,mn_c))
                x1,y1 = col_starts[mn_c-1],row_starts[mn_r-1]; x2,y2 = col_starts[mx_c],row_starts[mx_r]
                cell = ws.cell(mn_r,mn_c)
            else:
                x1,y1 = col_starts[c-1],row_starts[r-1]; x2,y2 = col_starts[c],row_starts[r]
                cell = ws.cell(r,c)
            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type=="solid": bg_hex = _resolve_color(cell.fill.fgColor,"FFFFFF")
            draw.rectangle([x1,y1,x2-1,y2-1],fill=f"#{bg_hex}",outline="#CCCCCC",width=1)
            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color,"000000")
                    if fc.upper() != bg_hex.upper(): txt_color = f"#{fc}"
                bold = bool(cell.font and cell.font.bold)
                text = format_cell_value_with_fmt(cell) if cell.value is not None else ""
                cell_w = x2-x1; ch_w = 8 if bold else 7; max_chars = max(1,(cell_w-8)//ch_w)
                if len(text)>max_chars: text = text[:max_chars-1]+"…"
                draw.text((x1+4,y1+4),text,fill=txt_color)
    wb.close()
    return img,col_starts,row_starts,merged_master

def get_cell_pixel_bbox(col_starts,row_starts,target_row,target_col,merged_master=None):
    c = max(1,min(target_col,len(col_starts)-1)); r = max(1,min(target_row,len(row_starts)-1))
    if merged_master:
        info = merged_master.get((r,c))
        if info:
            mn_r,mn_c,mx_r,mx_c = info
            return (col_starts[mn_c-1],row_starts[mn_r-1],col_starts[min(mx_c,len(col_starts)-1)],row_starts[min(mx_r,len(row_starts)-1)])
    return (col_starts[c-1],row_starts[r-1],col_starts[min(c,len(col_starts)-1)],row_starts[min(r,len(row_starts)-1)])

def crop_context(img,x1,y1,x2,y2,pad_x=220,pad_y=160):
    iw,ih = img.size; cx1,cy1 = max(0,x1-pad_x),max(0,y1-pad_y); cx2,cy2 = min(iw,x2+pad_x),min(ih,y2+pad_y)
    return img.crop((cx1,cy1,cx2,cy2)),x1-cx1,y1-cy1,x2-cx1,y2-cy1

# ==============================
# EYE POPUP
# ==============================
@st.dialog("Cell View", width="large")
def show_eye_popup(field,info,excel_path,sheet_name):
    st.markdown(f"### 📍 {field}")
    value = info.get("modified",info["value"]); target_row = info.get("excel_row"); target_col = info.get("excel_col")
    col_a,col_b = st.columns([1,1])
    with col_a: st.markdown("**Extracted Value**"); st.code(value if value else "(empty)")
    with col_b:
        col_letter = get_column_letter(target_col) if target_col else "?"
        st.markdown(f"<div style='padding:10px 0;color:var(--t2);font-size:var(--sz-body);font-family:var(--font);'>Cell: <span style='color:var(--blue);font-weight:bold;'>{col_letter}{target_row or '?'}</span> &nbsp;|&nbsp; Row <span style='color:var(--t0);'>{target_row or '?'}</span> · Col <span style='color:var(--t0);'>{target_col or '?'}</span></div>", unsafe_allow_html=True)
    if not target_row or not target_col: st.warning("No cell location recorded for this field."); return
    ext = os.path.splitext(excel_path)[1].lower()
    if ext == ".csv":
        st.markdown("---")
        try:
            with open(excel_path,"r",encoding="utf-8-sig") as f: all_rows = list(csv.reader(f))
            if not all_rows: return
            n_rows,n_cols = len(all_rows),max(len(r) for r in all_rows)
            r0,r1 = max(0,target_row-6),min(n_rows,target_row+5)
            col_headers = "".join(f"<th style='background:var(--s0);color:var(--t3);font-size:var(--sz-xs);padding:4px 8px;border:1px solid var(--b0);font-family:var(--mono);font-weight:600;'>{get_column_letter(c+1)}</th>" for c in range(n_cols))
            thead = f"<thead><tr><th style='background:var(--s0);color:var(--t3);font-size:var(--sz-xs);padding:4px 8px;border:1px solid var(--b0);font-family:var(--mono);font-weight:600;'>#</th>{col_headers}</tr></thead>"
            tbody = ""
            for r_idx in range(r0,r1):
                row_data = all_rows[r_idx] if r_idx<len(all_rows) else []
                is_tr    = (r_idx+1==target_row)
                rn_style = "background:#1a2540;color:#4f9cf9;font-weight:bold;" if is_tr else "background:var(--s0);color:var(--t3);"
                cells    = f"<td style='{rn_style}font-size:var(--sz-xs);padding:4px 8px;border:1px solid var(--b0);text-align:center;font-family:var(--mono);'>{r_idx+1}</td>"
                for c_idx in range(n_cols):
                    cell_val = row_data[c_idx] if c_idx<len(row_data) else ""
                    is_tc    = is_tr and (c_idx+1==target_col)
                    if is_tc:    style = "background:#2a2010;border:2px solid var(--yellow);color:var(--t0);font-weight:bold;"
                    elif is_tr:  style = "background:var(--blue-g);border:1px solid rgba(79,156,249,0.2);color:var(--t1);"
                    else:        style = "background:var(--surface);border:1px solid var(--b0);color:var(--t2);"
                    cells += f"<td style='{style}font-size:var(--sz-sm);padding:5px 10px;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-family:var(--font);'>{cell_val}</td>"
                tbody += f"<tr>{cells}</tr>"
            st.markdown(f"<div style='overflow-x:auto;border-radius:var(--radius);border:1px solid var(--b0);'><table style='border-collapse:collapse;width:100%;font-family:var(--font);'>{thead}<tbody>{tbody}</tbody></table></div>",unsafe_allow_html=True)
        except Exception as e: st.error(f"CSV preview error: {e}")
        return
    st.markdown("---")
    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img,col_starts,row_starts,merged_master = render_excel_sheet(excel_path,sheet_name,scale=1.0)
            st.session_state[cache_key] = (rendered_img,col_starts,row_starts,merged_master)
        else:
            rendered_img,col_starts,row_starts,merged_master = st.session_state[cache_key]
    try:
        img = rendered_img.copy(); draw = ImageDraw.Draw(img,"RGBA")
        x1,y1,x2,y2 = get_cell_pixel_bbox(col_starts,row_starts,target_row,target_col,merged_master)
        draw.rectangle([x1+1,y1+1,x2-1,y2-1],fill=(255,230,0,80))
        draw.rectangle([x1,y1,x2,y2],outline=(245,158,11,255),width=3)
        draw.rectangle([x1+3,y1+3,x2-3,y2-3],outline=(255,255,255,160),width=1)
        cropped,_,_,_,_ = crop_context(img,x1,y1,x2,y2,pad_x=300,pad_y=200)
        col_letter = get_column_letter(target_col)
        st.image(cropped,use_container_width=True,caption=f"Cell {col_letter}{target_row}  ·  Value: {value or '(empty)'}")
    except Exception as e: st.error(f"Rendering error: {e}")

# ==============================
# FORMAT CONVERTERS
# ==============================
def to_duck_creek_json(mapped_records,sheet_meta):
    transactions = []
    for rec in mapped_records:
        claim_obj = {}
        for sf,fd in rec.items():
            if sf.startswith("_"): continue
            claim_obj[sf] = {"value":fd.get("value",""),"confidence":fd.get("confidence",0),"edited":fd.get("edited",False)}
            if fd.get("edited"): claim_obj[sf]["originalValue"] = fd.get("original","")
        transactions.append({"transactionType":"UPDATE","avgConfidence":rec.get("_avg_confidence",0),"claim":claim_obj})
    return {"schema":"DuckCreek.Claims.Transaction.v6","exportDate":datetime.datetime.now().isoformat(),"source":"TPA_Loss_Run_Parser","sheetName":sheet_meta.get("sheet_name",""),"recordCount":len(transactions),"transactions":transactions}

def to_guidewire_json(mapped_records,sheet_meta):
    _GW = {"Claim Number":"claimNumber","Claimant Name":"claimantName","Loss Date":"lossDate","Date Reported":"reportedDate","Total Incurred":"totalIncurredAmount","Total Paid":"totalPaidAmount","Reserve":"reserveAmount","Status":"status","Line of Business":"lineOfBusinessCode","Policy Number":"policyNumber","Insured Name":"insuredName","Description of Loss":"lossDescription","Cause of Loss":"causeOfLoss"}
    claims = []
    for rec in mapped_records:
        claim_obj = {"_type":"cc.Claim","_confidence":rec.get("_avg_confidence",0)}; financials = {}
        for sf,fd in rec.items():
            if sf.startswith("_"): continue
            gw_key = _GW.get(sf, sf[0].lower()+sf[1:].replace(" ",""))
            val    = fd.get("value","")
            if any(x in sf.lower() for x in ["paid","reserve","incurred","deductible","recovery","subrogation"]):
                financials[gw_key] = {"amount":val,"currency":"USD","confidence":fd.get("confidence",0)}
                if fd.get("edited"): financials[gw_key]["originalValue"] = fd.get("original","")
            else:
                claim_obj[gw_key] = {"value":val,"confidence":fd.get("confidence",0)}
                if fd.get("edited"): claim_obj[gw_key]["originalValue"] = fd.get("original","")
        if financials: claim_obj["financials"] = financials
        claims.append(claim_obj)
    return {"schema":"Guidewire.ClaimCenter.REST.v1","exportDate":datetime.datetime.now().isoformat(),"source":"TPA_Loss_Run_Parser","sheetName":sheet_meta.get("sheet_name",""),"recordCount":len(claims),"data":{"claims":claims}}

def to_standard_json(export_data,sheet_meta,totals,merged_meta):
    titles_section = []
    for _,m in sorted([(k,v) for k,v in merged_meta.items() if v.get("value")],key=lambda x:(x[1]["row_start"],x[1]["col_start"])):
        titles_section.append({"type":m["type"],"value":m["value"],"excel_row":m["excel_row"],"excel_col":m["excel_col"],"span_cols":m["span_cols"],"span_rows":m["span_rows"]})
    totals_section = {}
    if totals: totals_section = {"excel_row":totals.get("excel_row"),"rows":totals.get("rows",[]),"aggregated":totals.get("aggregated",{})}
    return {"exportDate":datetime.datetime.now().isoformat(),"sheetMeta":{"sheet_name":sheet_meta.get("sheet_name"),"record_count":sheet_meta.get("record_count")},"titleRows":titles_section,"records":export_data,"totals":totals_section,"recordCount":len(export_data)}

def build_mapped_records_for_export(data,schema_name,selected_sheet):
    schema = SCHEMAS[schema_name]; custom_flds = st.session_state.get(f"custom_fields_{schema_name}",[])
    export_flds = list(schema["required_fields"]) + [f for f in custom_flds if f not in schema["required_fields"]]
    title_fields = st.session_state.get("sheet_cache",{}).get(selected_sheet,{}).get("title_fields",{})
    records = []
    for i,row in enumerate(data):
        c_id   = detect_claim_id(row,i)
        mapped = map_claim_to_schema(row,schema_name,title_fields)
        rec    = {}; confs = []
        for sf in export_flds:
            if sf not in mapped:
                rec[sf] = {"value":"","confidence":0,"edited":False,"original":""}; confs.append(0); continue
            m       = mapped[sf]; mk_key = f"mod_{selected_sheet}_{c_id}_schema_{sf}"
            live_val = st.session_state.get(mk_key,None); orig = m["info"].get("value",""); final = live_val if live_val is not None else m["value"]
            rec[sf] = {"value":final,"original":orig,"edited":final!=orig,"confidence":m["confidence"],"excel_row":m["info"].get("excel_row"),"excel_col":m["info"].get("excel_col")}
            confs.append(m["confidence"])
        rec["_avg_confidence"] = round(sum(confs)/len(confs)) if confs else 0
        rec["_claim_id"]       = c_id
        records.append(rec)
    return records

# ==============================
# CONFIDENCE ENGINE
# ==============================
def _word_tokens(s: str) -> set:
    stopwords = {"of","the","a","an","and","or","to","in","for"}
    words = re.sub(r"[_/#+]", " ", s.lower()).split()
    return {w for w in words if len(w) > 1 and w not in stopwords}

def _str_similarity(a: str, b: str) -> float:
    a_tok, b_tok = _word_tokens(a), _word_tokens(b)
    if not a_tok or not b_tok: return 0.0
    if a_tok == b_tok: return 1.0
    return len(a_tok & b_tok) / len(a_tok | b_tok)

def _header_match_score(excel_col, schema_field, aliases):
    ec_norm = excel_col.lower().replace("_"," ").strip()
    for alias in aliases:
        if ec_norm == alias.lower(): return 1.0
    best = max((_str_similarity(ec_norm, a.lower()) for a in aliases), default=0.0)
    return max(best, _str_similarity(ec_norm, schema_field.lower()))

def _value_quality_score(value, schema_field):
    if not value or not value.strip(): return 0.0
    v, sf = value.strip(), schema_field.lower()
    if any(x in sf for x in ["date","loss dt"]):
        for p in [r"\d{2}-\d{2}-\d{4}",r"\d{4}-\d{2}-\d{2}",r"\d{2}/\d{2}/\d{4}",r"\d{1,2}/\d{1,2}/\d{2,4}"]:
            if re.fullmatch(p, v): return 1.0
        return 0.4
    if any(x in sf for x in ["incurred","paid","reserve","amount","deductible","recovery"]):
        try:
            float(v.replace(",","").replace("$","").replace("(","-").replace(")",""))
            return 1.0
        except ValueError: return 0.3
    if any(x in sf for x in ["id","number","no","code"]):
        return 0.9 if len(v) >= 2 else 0.5
    if "status" in sf:
        return 1.0 if v.lower() in {"open","closed","pending","reopened","denied","settled"} else 0.7
    return 0.85 if len(v) > 0 else 0.0

_MIN_HEADER_MATCH = 0.70

def map_claim_to_schema(claim, schema_name, title_fields=None):
    if schema_name not in SCHEMAS: return {}
    schema, aliases, accepted = SCHEMAS[schema_name], SCHEMAS[schema_name].get("field_aliases",{}), SCHEMAS[schema_name]["accepted_fields"]
    title_fields = title_fields or {}
    result = {}
    for schema_field in accepted:
        field_aliases = aliases.get(schema_field, [schema_field.lower()])
        best_excel_col, best_header_sc, best_info = None, 0.0, None
        for excel_col, info in claim.items():
            h_sc = _header_match_score(excel_col, schema_field, field_aliases)
            if h_sc > best_header_sc:
                best_header_sc, best_excel_col, best_info = h_sc, excel_col, info
        if best_header_sc >= _MIN_HEADER_MATCH and best_info is not None:
            val  = best_info.get("modified", best_info.get("value",""))
            v_sc = _value_quality_score(val, schema_field)
            conf = round(best_header_sc*0.40*100 + v_sc*0.60*100)
            result[schema_field] = {"excel_field":best_excel_col,"value":val,"header_score":round(best_header_sc*100),"value_score":round(v_sc*100),"confidence":conf,"is_required":schema_field in schema["required_fields"],"info":best_info,"from_title":False}
        elif schema_field in title_fields:
            tf   = title_fields[schema_field]
            val  = tf.get("value","")
            v_sc = _value_quality_score(val, schema_field)
            conf = min(95, round(1.0*0.40*100 + v_sc*0.60*100))
            result[schema_field] = {"excel_field":f"[title row {tf['excel_row']}]","value":val,"header_score":100,"value_score":round(v_sc*100),"confidence":conf,"is_required":schema_field in schema["required_fields"],"info":tf,"from_title":True}
    return result

def extract_title_fields(merged_meta):
    found = {}
    title_rows = sorted([v for v in merged_meta.values() if v.get("value") and v["type"] in ("TITLE","HEADER")], key=lambda x:(x["row_start"],x["col_start"]))
    for m in title_rows:
        text, r, c = str(m["value"]).strip(), m["excel_row"], m["excel_col"]
        def _info(val): return {"value":val,"original":val,"modified":val,"source":"title_row","excel_row":r,"excel_col":c,"title_text":text}
        pol = re.search(r'Policy\s*(?:#|No\.?|Number)?\s*[:\-]\s*([A-Z0-9][A-Z0-9\-/\.]+)', text, re.IGNORECASE)
        if pol and "Policy Number" not in found: found["Policy Number"] = _info(pol.group(1).strip())
        ins = re.search(r'Insured\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if ins and "Insured Name" not in found: found["Insured Name"] = _info(ins.group(1).strip())
        carr = re.search(r'Carrier\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if carr:
            val = carr.group(1).strip()
            if "Carrier" not in found: found["Carrier"] = _info(val)
            if "Carrier Name" not in found: found["Carrier Name"] = _info(val)
        state = re.search(r'State\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if state:
            val = state.group(1).strip()
            for k in ["State","Jurisdiction","State Code"]:
                if k not in found: found[k] = _info(val)
        period = re.search(r'Period\s*[:\-]?\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[\s\u2013\u2014\-to]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', text, re.IGNORECASE)
        if period:
            s, e = period.group(1).strip(), period.group(2).strip()
            for k,v in [("Policy Period Start",s),("Policy Period End",e),("Policy Effective Date",s),("Policy Expiry Date",e)]:
                if k not in found: found[k] = _info(v)
        lob_map = [
            (r"workers[\'\'\u2019\s\-]*compensation","Workers Compensation"),
            (r"workers[\s\-]*comp\b","Workers Compensation"),
            (r"\bW\.?C\.?\b(?:\s+loss|\s+claim|\s+run)?","Workers Compensation"),
            (r"commercial\s+general\s+liability","Commercial General Liability"),
            (r"\bC\.?G\.?L\.?\b","Commercial General Liability"),
            (r"commercial\s+auto(?:mobile|motive)?","Commercial Auto"),
            (r"commercial\s+prop(?:erty)?","Commercial Property"),
            (r"professional\s+liability","Professional Liability"),
            (r"\bE\.?\s*&\s*O\.?\b","Professional Liability"),
            (r"general\s+liability|\bG\.?L\.?\b","General Liability"),
        ]
        for pattern, lob_val in lob_map:
            if re.search(pattern, text, re.IGNORECASE) and "Line of Business" not in found:
                found["Line of Business"] = _info(lob_val)
                break
    return found

# ==============================
# UTILS
# ==============================
def get_val(claim,keys,default=""):
    for pk in keys:
        for k,v in claim.items():
            if pk.lower() in str(k).lower(): return v["value"] or default
    return default

def detect_claim_id(row,index=None):
    keys = ["claim id","claim_id","claimid","claim number","claim no","claim #","claim ref","claim reference","file number","record id"]
    for k,v in row.items():
        name = str(k).lower().replace("_"," ").strip()
        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip(): return str(val)
    if index is not None: return str(index+1)
    return ""

def clean_duplicate_fields(record):
    seen,out = set(),{}
    for k,v in record.items():
        if k.strip() not in seen: seen.add(k.strip()); out[k.strip()] = v
    return out

def save_feature_store(sheet_name,data):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S"); path = os.path.join(FEATURE_STORE_PATH,f"{sheet_name}_{ts}.json")
    def _san(obj):
        if isinstance(obj,dict): return {k:_san(v) for k,v in obj.items()}
        if isinstance(obj,list): return [_san(i) for i in obj]
        if isinstance(obj,str):  return normalize_str(obj)
        return obj
    with open(path,"w") as f: json.dump(_san(data),f,indent=2,ensure_ascii=False)
    return path

def _sanitize_for_json(obj):
    if isinstance(obj,dict): return {k:_sanitize_for_json(v) for k,v in obj.items()}
    if isinstance(obj,list): return [_sanitize_for_json(i) for i in obj]
    if isinstance(obj,str):  return normalize_str(obj)
    return obj

def extract_from_excel(file_path,sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path,"r",encoding="utf-8-sig") as f: rows = list(csv.reader(f))
        if not rows: return [],"UNKNOWN"
        return parse_rows(classify_sheet(rows),rows)
    else:
        wb = openpyxl.load_workbook(file_path,data_only=True); ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows: return [],"UNKNOWN"
        return parse_rows_with_cells(classify_sheet(raw_rows),raw_rows,cell_rows)

def parse_rows_with_cells(sheet_type,rows,cell_rows):
    if sheet_type == "SUMMARY":
        hri = None
        for i,row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt: hri = i; break
        if hri is None: return [],sheet_type
        headers = [str(h).strip() if h is not None else f"Column_{i}" for i,h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel,(raw_row,cell_row) in enumerate(zip(rows[hri+1:],cell_rows[hri+1:])):
            r_idx = hri+2+r_idx_rel
            if not any(raw_row): continue
            row_data = {}
            for c_idx_0,(raw_val,cell) in enumerate(zip(raw_row,cell_row)):
                if c_idx_0>=len(headers): continue
                clean_val = format_cell_value_with_fmt(cell); real_col = cell.column if hasattr(cell,'column') and cell.column else c_idx_0+1
                row_data[headers[c_idx_0]] = {"value":clean_val,"modified":clean_val,"excel_row":r_idx,"excel_col":real_col}
            if any(v["value"] for v in row_data.values()): extracted.append(row_data)
        return extracted,sheet_type
    hri = None
    for i,row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if ("claim" in rt or "employee name" in rt or "driver name" in rt) and ("date" in rt or "incurred" in rt or "paid" in rt or "injury" in rt or "incident" in rt):
            hri = i; break
    if hri is None: return [],sheet_type
    headers = [str(h).strip() if h is not None else f"Column_{i}" for i,h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel,(raw_row,cell_row) in enumerate(zip(rows[hri+1:],cell_rows[hri+1:])):
        r_idx = hri+2+r_idx_rel
        if not any(raw_row): continue
        if any(str(c).lower().strip() in ["totals","total","grand total"] for c in raw_row if c): break
        row_data = {}
        for c_idx_0,(raw_val,cell) in enumerate(zip(raw_row,cell_row)):
            if c_idx_0>=len(headers): continue
            clean_val = format_cell_value_with_fmt(cell); real_col = cell.column if hasattr(cell,'column') and cell.column else c_idx_0+1
            row_data[headers[c_idx_0]] = {"value":clean_val,"modified":clean_val,"excel_row":r_idx,"excel_col":real_col}
        if any(v["value"] for v in row_data.values()): extracted.append(row_data)
    return extracted,sheet_type

def parse_rows(sheet_type,rows):
    if sheet_type == "SUMMARY":
        hri = None
        for i,row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt: hri = i; break
        if hri is None: return [],sheet_type
        headers = [str(h).strip() if h is not None else f"Column_{i}" for i,h in enumerate(rows[hri])]
        extracted = []
        for r_idx,row in enumerate(rows[hri+1:],start=hri+2):
            if not any(row): continue
            row_data = {}
            for c_idx,value in enumerate(row,start=1):
                if c_idx-1>=len(headers): continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx-1]] = {"value":clean_val,"modified":clean_val,"excel_row":r_idx,"excel_col":c_idx}
            if any(v["value"] for v in row_data.values()): extracted.append(row_data)
        return extracted,sheet_type
    hri = None
    for i,row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if ("claim" in rt or "employee name" in rt or "driver name" in rt) and ("date" in rt or "incurred" in rt or "paid" in rt or "injury" in rt or "incident" in rt):
            hri = i; break
    if hri is None: return [],sheet_type
    headers = [str(h).strip() if h is not None else f"Column_{i}" for i,h in enumerate(rows[hri])]
    extracted = []
    for r_idx,row in enumerate(rows[hri+1:],start=hri+2):
        if not any(row): continue
        if any(str(cell).lower().strip() in ["totals","total","grand total"] for cell in row if cell): break
        row_data = {}
        for c_idx,value in enumerate(row,start=1):
            if c_idx-1>=len(headers): continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx-1]] = {"value":clean_val,"modified":clean_val,"excel_row":r_idx,"excel_col":c_idx}
        if any(v["value"] for v in row_data.values()): extracted.append(row_data)
    return extracted,sheet_type

# ==============================
# SESSION STATE DEFAULTS
# ==============================
for _k,_v in [
    ("conf_threshold",80),("use_conf_threshold",True),("active_schema",None),
    ("schema_popup_target",None),("schema_popup_tab","required"),
]:
    if _k not in st.session_state: st.session_state[_k] = _v

# ==============================
# TOP BAR — title row with working settings button
# ==============================
active_schema = st.session_state.get("active_schema", None)

def _navbar_badge_html():
    if not active_schema or active_schema not in SCHEMAS: return ""
    sc = SCHEMAS[active_schema]
    return (
        f'<span class="navbar-schema-badge" '
        f'style="border-color:{sc["color"]}44;color:{sc["color"]};background:{sc["color"]}11;'
        f'display:inline-flex;align-items:center;gap:6px;border-radius:6px;padding:4px 12px;'
        f'font-size:12px;font-weight:700;font-family:monospace;border:1px solid;white-space:nowrap;">'
        f'{sc["icon"]} {active_schema} &nbsp;&middot;&nbsp; {sc["version"]}</span>'
    )

_logo_nav  = _logo_img_tag(height=34)
_nav_badge = _navbar_badge_html()

# Title row: logo + title on left, gear button on far right — all native Streamlit columns
_col_title, _col_gear = st.columns([11, 1])

with _col_title:
    st.markdown(
        '<div class="topbar-title-row">'
        + _logo_nav
        + '<div style="display:inline-flex;flex-direction:column;vertical-align:middle;margin-left:14px;">'
        '<span class="navbar-title">&#128737; TPA Loss Run Parser</span>'
        '<span class="navbar-subtitle">Insurance Loss Run Parsing &amp; Schema Export Platform</span>'
        '</div>'
        + ('&nbsp;&nbsp;' + _nav_badge if _nav_badge else '')
        + '</div>',
        unsafe_allow_html=True
    )

with _col_gear:
    if st.button("⚙", key="open_settings", help="Settings", use_container_width=True):
        show_settings_dialog()

st.markdown('<hr class="topbar-divider">', unsafe_allow_html=True)

# Sheet dropdown column
_, col_sheet_dropdown = st.columns([6.8, 1.2])

if st.session_state.get("schema_popup_target"):
    _target = st.session_state["schema_popup_target"]; st.session_state["schema_popup_target"] = None
    show_schema_fields_dialog(_target)

# ==============================
# FILE UPLOAD
# ==============================
uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx","csv"])

if uploaded:
    if "tmpdir" not in st.session_state: st.session_state.tmpdir = tempfile.mkdtemp()
    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path,"wb") as f: f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None
        for key in list(st.session_state.keys()):
            if key.startswith("_rendered_"): del st.session_state[key]

        file_hash  = _compute_file_sha256(excel_path)
        st.session_state["current_file_hash"] = file_hash

        sheet_hashes = {}
        for sn in st.session_state.sheet_names:
            sheet_hashes[sn] = _compute_sheet_sha256(excel_path, sn)
        st.session_state["sheet_hashes"] = sheet_hashes

        hash_store = _load_hash_store()

        # ── File-level duplicate check ──
        if file_hash in hash_store:
            st.session_state["is_duplicate_file"]    = True
            st.session_state["duplicate_first_seen"] = hash_store[file_hash]["first_seen"]
            st.session_state["duplicate_orig_name"]  = hash_store[file_hash]["filename"]
        else:
            st.session_state["is_duplicate_file"]    = False
            st.session_state["duplicate_first_seen"] = None
            hash_store[file_hash] = {
                "filename":    uploaded.name,
                "first_seen":  datetime.datetime.now().isoformat(),
                "sheet_hashes": sheet_hashes,           # {sheet_name: sha256}
            }
            _save_hash_store(hash_store)
            _append_audit({"event":"FILE_INGESTED","timestamp":datetime.datetime.now().isoformat(),
                           "filename":uploaded.name,"file_hash":file_hash,
                           "sheets":st.session_state.sheet_names})

        # ── Sheet-level duplicate check ──
        # Build reverse map: sheet_sha256 → {filename, sheet_name, first_seen}
        # from every OTHER file already in the store
        _sheet_hash_index = {}
        for _fh, _fdata in hash_store.items():
            if _fh == file_hash:
                continue  # exclude current file — we only want cross-file matches
            if not isinstance(_fdata, dict):
                continue
            for _sn, _sh in _fdata.get("sheet_hashes", {}).items():
                _sheet_hash_index[_sh] = {
                    "filename":   _fdata.get("filename", "unknown"),
                    "sheet_name": _sn,
                    "first_seen": _fdata.get("first_seen", "unknown"),
                    "file_hash":  _fh,
                }

        # Per-sheet: None = new, dict = seen before in a different file
        sheet_dup_info = {}
        for sn, sh in sheet_hashes.items():
            sheet_dup_info[sn] = _sheet_hash_index.get(sh)  # None if not a dup
        st.session_state["sheet_dup_info"] = sheet_dup_info
    else:
        file_hash    = st.session_state.get("current_file_hash","")
        sheet_hashes = st.session_state.get("sheet_hashes",{})

    is_dup       = st.session_state.get("is_duplicate_file", False)
    sheet_dup_info = st.session_state.get("sheet_dup_info", {})
    # Count how many sheets are duplicates from other files
    n_dup_sheets = sum(1 for v in sheet_dup_info.values() if v is not None)

    file_format = os.path.splitext(uploaded.name)[1].upper().lstrip(".")
    n_sheets    = len(st.session_state.sheet_names)
    file_size_b = os.path.getsize(excel_path)
    file_size   = f"{file_size_b/1024:.1f} KB" if file_size_b < 1_000_000 else f"{file_size_b/1_048_576:.2f} MB"
    badge_cls   = "badge-duplicate" if is_dup else "badge-unique"
    badge_lbl   = "DUPLICATE" if is_dup else "UNIQUE"
    hash_short  = file_hash[:18]+"…" if file_hash else "—"
    dup_note    = "<span style='font-size:11px;color:#f5c842;font-family:var(--mono);'>⚠ Already processed</span>" if is_dup else "<span style='font-size:11px;color:#34d399;font-family:var(--mono);'>✓ New file ingested</span>"

    # Sheet pills — mark each with ⚠ if it's a sheet-level duplicate from another file
    def _sheet_pill(sn):
        info = sheet_dup_info.get(sn)
        if info:
            title = f"Sheet already seen in {info['filename']} (sheet: {info['sheet_name']}) on {info['first_seen'][:10]}"
            return (f"<span class='sheet-pill-sm' title='{title}' "
                    f"style='border-color:#f5c842;color:#f5c842;'>⚠ {sn}</span>")
        return f"<span class='sheet-pill-sm'>{sn}</span>"

    sheet_pills = "".join(_sheet_pill(sn) for sn in st.session_state.sheet_names)

    # File status line
    _file_status_color = '#f5c842' if is_dup else ('#f5c842' if n_dup_sheets > 0 else '#34d399')
    _file_status_text  = 'Duplicate' if is_dup else (f'{n_dup_sheets} sheet(s) duplicate' if n_dup_sheets > 0 else 'New')

    st.markdown(f"""
    <div class="file-card">
      <div class="file-card-header">
        <div class="file-card-title">📄 {uploaded.name} <span class="file-badge {badge_cls}">{badge_lbl}</span></div>
        {dup_note}
      </div>
      <div class="file-card-body">
        <div class="file-stat"><div class="file-stat-lbl">Format</div><div class="file-stat-val accent">{file_format}</div></div>
        <div class="file-stat"><div class="file-stat-lbl">Sheets</div><div class="file-stat-val accent">{n_sheets}</div></div>
        <div class="file-stat"><div class="file-stat-lbl">File Size</div><div class="file-stat-val">{file_size}</div></div>
        <div class="file-stat"><div class="file-stat-lbl">Status</div><div class="file-stat-val" style="color:{_file_status_color};">{_file_status_text}</div></div>
        <div class="file-stat"><div class="file-stat-lbl">SHA-256</div><div class="file-stat-val mono-sm">{hash_short}</div></div>
      </div>
      <div class="file-card-sheets"><div class="section-lbl">Worksheets</div>{sheet_pills}</div>
    </div>
    """, unsafe_allow_html=True)

    if is_dup:
        st.warning(
            f"⚠ **Duplicate file detected.** First processed on "
            f"**{st.session_state.get('duplicate_first_seen','unknown')}** "
            f"(original: `{st.session_state.get('duplicate_orig_name', uploaded.name)}`).",
           
        )
    elif n_dup_sheets > 0:
        _dup_sheet_names = [sn for sn, v in sheet_dup_info.items() if v is not None]
        _dup_details = "; ".join(
            f"**{sn}** → seen in `{sheet_dup_info[sn]['filename']}` on {sheet_dup_info[sn]['first_seen'][:10]}"
            for sn in _dup_sheet_names
        )
        st.warning(
            f"⚠ **{n_dup_sheets} sheet(s) already processed in a different file.** {_dup_details}",
            icon="📋"
        )
    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox("Sheet", st.session_state.sheet_names, index=0, label_visibility="collapsed")

    st.markdown("<hr>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Reading '{selected_sheet}'..."):
            data,sheet_type = extract_from_excel(excel_path,selected_sheet)
            merged_meta     = extract_merged_cell_metadata(excel_path,selected_sheet)
            totals_data     = extract_totals_row(excel_path,selected_sheet)
            total_rows,total_cols = get_sheet_dimensions(excel_path,selected_sheet)
            if not data:
                st.warning(f"No data found in sheet '{selected_sheet}'."); st.stop()
            for row in data:
                for fld,inf in row.items():
                    for key in ("value","modified"):
                        if key in inf and isinstance(inf[key],str): inf[key] = normalize_str(inf[key])
            _title_flds = extract_title_fields(merged_meta)
            sh_hash     = sheet_hashes.get(selected_sheet,"")
            st.session_state.sheet_cache[selected_sheet] = {
                "data":data,"merged_meta":merged_meta,"totals":totals_data,
                "title_fields":_title_flds,"sheet_type":sheet_type,
                "total_rows":total_rows,"total_cols":total_cols,"sheet_hash":sh_hash,
            }
            st.session_state.selected_idx = 0; st.session_state.focus_field = None
            _append_audit({"event":"SHEET_PARSED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"sheet_hash":sh_hash,"claim_rows":len(data),"sheet_type":sheet_type,"total_rows":total_rows,"total_cols":total_cols})

            # ── AUTO-NORMALIZE on load if a schema is already active ──
            _cur_schema = st.session_state.get("active_schema")
            if _cur_schema and _cur_schema in SCHEMAS:
                auto_normalize_on_schema_activate(data, _cur_schema, selected_sheet)
                st.session_state.sheet_cache[selected_sheet]["_normalized_for"] = _cur_schema

    active       = st.session_state.sheet_cache[selected_sheet]
    data         = active["data"]
    merged_meta  = active.get("merged_meta",{})
    totals_data  = active.get("totals",{})
    title_fields = active.get("title_fields",{})
    sheet_type   = active.get("sheet_type","UNKNOWN")
    total_rows   = active.get("total_rows",0)
    total_cols   = active.get("total_cols",0)
    sh_hash      = active.get("sheet_hash","")
    sh_hash_short = sh_hash[:18]+"…" if sh_hash else "—"

    # ── AUTO-NORMALIZE when schema changes ──
    _active_schema_now = st.session_state.get("active_schema")
    _normalized_for    = active.get("_normalized_for")
    if _active_schema_now and _active_schema_now in SCHEMAS and _normalized_for != _active_schema_now:
        auto_normalize_on_schema_activate(data, _active_schema_now, selected_sheet)
        active["_normalized_for"] = _active_schema_now

    totals_cls   = "hi" if totals_data else "mid"
    totals_found = "Found" if totals_data else "None"

    st.markdown(f"""
    <div class="sheet-card">
      <div class="sheet-card-hdr">
        <div class="sheet-card-name">⊞ {selected_sheet} <span class="sheet-type-tag {'unk' if sheet_type=='UNKNOWN' else ''}">{sheet_type}</span></div>
        <span style="font-size:10px;color:var(--t3);font-family:var(--mono);">SHA-256: <span style="color:var(--green);font-size:9px;">{sh_hash_short}</span></span>
      </div>
      <div class="sheet-stats-grid">
        <div class="sh-stat"><div class="sh-stat-lbl">Claim Rows</div><div class="sh-stat-val hi">{len(data)}</div></div>
        <div class="sh-stat"><div class="sh-stat-lbl">Total Rows</div><div class="sh-stat-val">{total_rows}</div></div>
        <div class="sh-stat"><div class="sh-stat-lbl">Columns</div><div class="sh-stat-val">{total_cols}</div></div>
        <div class="sh-stat"><div class="sh-stat-lbl">Merged Regions</div><div class="sh-stat-val">{len(merged_meta)}</div></div>
        <div class="sh-stat"><div class="sh-stat-lbl">Totals Row</div><div class="sh-stat-val {totals_cls}">{totals_found}</div></div>
        <div class="sh-stat"><div class="sh-stat-lbl">Title Fields</div><div class="sh-stat-val">{len(title_fields)}</div></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Per-sheet duplicate warning ──
    _selected_sheet_dup = sheet_dup_info.get(selected_sheet)
    if _selected_sheet_dup:
        _orig_file  = _selected_sheet_dup.get("filename","unknown file")
        _orig_sheet = _selected_sheet_dup.get("sheet_name", selected_sheet)
        _orig_date  = _selected_sheet_dup.get("first_seen","")[:10]
        _same_name  = _orig_sheet == selected_sheet
        _sheet_ref  = f"sheet **{_orig_sheet}**" if not _same_name else f"the same sheet name"
        st.warning(
            f"⚠ **This sheet was already processed** — {_sheet_ref} in `{_orig_file}` on **{_orig_date}**. "
            ,
            
        )
    curr_claim = data[st.session_state.selected_idx]

    col_nav,col_main,col_fmt = st.columns([1.2,3.2,1.4],gap="large")

    # ══════════════════════════════════════════
    # LEFT NAV
    # ══════════════════════════════════════════
    with col_nav:
        with st.container(height=700,border=False):
            st.markdown("<p class='section-lbl'>Claim Records</p>", unsafe_allow_html=True)

            # ── Pure text search — no external calls ──
            _search_k = f"search_{selected_sheet}"
            _search_q = st.text_input("", key=_search_k, placeholder="🔍 Filter claims…", label_visibility="collapsed")
            _q_lower  = _search_q.strip().lower()
            if _q_lower:
                _hit_indices = [
                    i for i, row in enumerate(data)
                    if any(_q_lower in str(v.get("modified", v.get("value",""))).lower()
                           for v in row.values())
                ]
                st.markdown(f"<div style='font-size:var(--sz-xs);color:var(--green);font-family:var(--mono);margin:3px 0 6px;'>● {len(_hit_indices)} match{'es' if len(_hit_indices)!=1 else ''}</div>", unsafe_allow_html=True)
            else:
                _hit_indices = list(range(len(data)))
            _display_indices = _hit_indices

            for i in _display_indices:
                row_data = data[i]
                is_sel   = "selected-card" if st.session_state.selected_idx==i else ""
                c_id     = detect_claim_id(row_data,i)
                c_name   = get_val(row_data,["Insured Name","Name","Company","Claimant","TPA_NAME"],"Unknown Entity")
                raw_st   = get_val(row_data,["Status","CLAIM_STATUS"],"")
                c_status = raw_st or ("Yet to Review" if i==0 else "In Progress" if i==1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or c_status.lower()=="open" else "status-text"
                st.markdown(f"""<div class="claim-card {is_sel}">
                    <div style="font-weight:700;color:var(--t0);font-size:var(--sz-body);font-family:var(--font-head);">{c_id}</div>
                    <div style="color:var(--t1);font-size:var(--sz-xs);margin-top:3px;font-family:var(--font);">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Select",key=f"sel_{selected_sheet}_{i}",use_container_width=True):
                    st.session_state.selected_idx = i; st.session_state.focus_field = None; st.rerun()

    # ══════════════════════════════════════════
    # MIDDLE PANEL
    # ══════════════════════════════════════════
    with col_main:
        # ── Silently enrich Cause of Loss for current claim via LLM ──
        # Returns True if it ran and wrote new data → rerun surfaces the result immediately
        curr_claim_id = detect_claim_id(curr_claim)
        if enrich_claim_cause_of_loss(curr_claim, curr_claim_id, selected_sheet):
            st.rerun()  # Surface the enriched cause + summary without user action

        sorted_titles = sorted([(k,v) for k,v in merged_meta.items() if v.get("value")],key=lambda x:(x[1]["row_start"],x[1]["col_start"]))
        if sorted_titles:
            main_title_val,sub_title_val = "",""
            for _,m in sorted_titles:
                if m["type"]=="TITLE":
                    if not main_title_val: main_title_val = m["value"]
                    elif not sub_title_val: sub_title_val = m["value"]
            if main_title_val or sub_title_val:
                st.markdown(f"""<div class="sheet-title-banner">
                    <div class="sheet-title-label">Sheet Title</div>
                    <div class="sheet-title-value">{main_title_val}</div>
                    {"" if not sub_title_val else f'<div class="sheet-subtitle-val">{sub_title_val}</div>'}
                </div>""", unsafe_allow_html=True)

        head_left,head_right = st.columns([3,1])
        with head_left:
            st.markdown("<p class='section-lbl'>Review Details</p>", unsafe_allow_html=True)
            h_name   = get_val(curr_claim,["Insured Name","Name","Claimant","TPA_NAME"],"Unknown Entity")
            h_date   = get_val(curr_claim,["Loss Date","Date","LOSS_DATE"],"N/A")
            h_status = get_val(curr_claim,["Status","CLAIM_STATUS"],"Submitted")
            h_total  = get_val(curr_claim,["Total Incurred","Incurred","Total","Amount","TOTAL_INCURRED"],"$0")
            st.markdown(f"""<div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} — {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
            """, unsafe_allow_html=True)
        with head_right:
            st.markdown("<p class='section-lbl' style='text-align:right;'>Export Selection</p>", unsafe_allow_html=True)
            b1,b2 = st.columns([1,1])
            with b1:
                if st.button("✔ All",key=f"all_{selected_sheet}_{curr_claim_id}",use_container_width=True):
                    for fld in curr_claim: st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{fld}"] = True
                    st.rerun()
            with b2:
                if st.button("✘ None",key=f"none_{selected_sheet}_{curr_claim_id}",use_container_width=True):
                    for fld in curr_claim: st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{fld}"] = False
                    st.rerun()

        # ── Cause of Loss enrichment runs silently — result shown in right panel only ──
        st.markdown("<hr>", unsafe_allow_html=True)

        _active_schema = st.session_state.get("active_schema",None)
        _use_conf      = st.session_state.get("use_conf_threshold",True)
        _conf_thresh   = st.session_state.get("conf_threshold",80) if _use_conf else 0
        _show_conf     = _use_conf

        if _active_schema and _active_schema in SCHEMAS:
            _schema_def  = SCHEMAS[_active_schema]
            _mapped      = map_claim_to_schema(curr_claim,_active_schema,title_fields)
            _custom_flds = st.session_state.get(f"custom_fields_{_active_schema}",[])
            _display_fields = list(_schema_def["required_fields"]) + [f for f in _custom_flds if f not in _schema_def["required_fields"]]
            _low_conf = [sf for sf in _display_fields if sf in _mapped and _mapped[sf]["confidence"]<_conf_thresh and _use_conf]
            _missing  = [sf for sf in _schema_def["required_fields"] if sf not in _mapped]

            if _missing:
                st.markdown(f"<div style='background:var(--red-g);border:1px solid rgba(248,113,113,0.3);border-radius:6px;padding:8px 12px;margin-bottom:8px;font-size:var(--sz-body);color:var(--red);font-family:var(--font);'>⚠ {len(_missing)} mandatory field(s) not mapped: {', '.join(_missing)}</div>", unsafe_allow_html=True)
            if _low_conf:
                st.markdown(f"<div style='background:var(--yellow-g);border:1px solid rgba(245,200,66,0.3);border-radius:6px;padding:8px 12px;margin-bottom:8px;font-size:var(--sz-body);color:var(--yellow);font-family:var(--font);'>⚡ {len(_low_conf)} field(s) below threshold ({_conf_thresh}%): {', '.join(_low_conf)}</div>", unsafe_allow_html=True)

            _col_hdr = "<div style='font-size:var(--sz-xs);font-weight:700;color:var(--t1);text-transform:uppercase;letter-spacing:1.4px;font-family:var(--font-head);padding-bottom:5px;border-bottom:1px solid var(--b0);margin-bottom:5px;'>{}</div>"
            if _show_conf:
                hc = st.columns([1.8,1.5,1.8,1.8,0.55,0.55,0.45])
                for ci,lbl in enumerate(["Schema Field","Confidence","Extracted Value","Modified Value"]):
                    with hc[ci]: st.markdown(_col_hdr.format(lbl), unsafe_allow_html=True)
            else:
                hc = st.columns([1.8,1.8,1.8,0.55,0.55,0.45])
                for ci,lbl in enumerate(["Schema Field","Extracted Value","Modified Value"]):
                    with hc[ci]: st.markdown(_col_hdr.format(lbl), unsafe_allow_html=True)

            for schema_field in _display_fields:
                if schema_field not in _mapped:
                    is_req = schema_field in _schema_def["required_fields"]
                    if is_req: _bg,_br,_fc,_lbl = "var(--red-g)","rgba(248,113,113,0.3)","var(--red)","MANDATORY · NOT FOUND"
                    else:      _bg,_br,_fc,_lbl = "var(--s0)","var(--b0)","var(--t3)","OPTIONAL · NOT IN SHEET"
                    st.markdown(f"<div style='display:flex;align-items:center;gap:8px;background:{_bg};border:1px solid {_br};border-radius:6px;padding:6px 10px;margin:2px 0;'><span style='color:{_fc};font-size:var(--sz-sm);font-weight:700;text-transform:uppercase;font-family:var(--font);'>{schema_field}</span><span style='background:var(--s1);color:{_fc};font-size:9px;border-radius:4px;padding:1px 5px;border:1px solid {_br};font-family:var(--mono);'>{_lbl}</span></div>", unsafe_allow_html=True)
                    continue

                m = _mapped[schema_field]; conf = m["confidence"]; excel_f = m["excel_field"]; info = m["info"]; is_req = m["is_required"]; is_title_sourced = m.get("from_title",False)

                if not _use_conf:        conf_col,row_border,row_bg = "var(--t3)","var(--b0)","var(--bg)"
                elif conf < _conf_thresh: conf_col,row_border,row_bg = "var(--red)","rgba(248,113,113,0.3)","var(--red-g)"
                elif conf < 75:          conf_col,row_border,row_bg = "var(--yellow)","rgba(245,200,66,0.3)","var(--yellow-g)"
                elif conf < 88:          conf_col,row_border,row_bg = "var(--yellow)","var(--b0)","var(--bg)"
                else:                    conf_col,row_border,row_bg = "var(--green)","var(--b0)","var(--bg)"

                ek = f"edit_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state:
                    # Initialize with auto-normalized value
                    raw_val = info.get("modified", info["value"])
                    normalized_init = auto_normalize_field(schema_field, raw_val, _active_schema)
                    st.session_state[mk] = normalized_init

                st.markdown(f"<div style='border-left:2px solid {row_border};background:{row_bg};border-radius:0 4px 4px 0;padding:2px 0 2px 4px;margin:1px 0;'></div>", unsafe_allow_html=True)

                _cur_val = st.session_state.get(mk, info.get("modified",info["value"]))
                _edited  = _cur_val != info["value"]
                _dot     = "<span style='color:var(--yellow);font-size:8px;'>●</span> " if _edited else ""
                _badge_html = "<span class='mandatory-asterisk' title='Mandatory'>*</span>" if is_req else "<span class='optional-badge'>OPT</span>"
                _ink = "var(--t0)" if is_req else "var(--t1)"
                _field_label_html = f"<div style='min-height:40px;display:flex;flex-direction:column;justify-content:center;color:{_ink};font-size:var(--sz-body);font-weight:600;text-transform:uppercase;letter-spacing:0.8px;font-family:var(--font-head);'><div style='display:flex;align-items:center;gap:3px;'>{_dot}{schema_field}{_badge_html}</div></div>"
                _conf_html = f"<div style='min-height:40px;display:flex;flex-direction:column;justify-content:center;gap:4px;'><span style='background:{conf_col}20;border:1px solid {conf_col};border-radius:20px;padding:2px 10px;font-size:var(--sz-body);color:{conf_col};font-weight:600;font-family:var(--mono);'>{conf}%</span><div style='background:var(--s1);border-radius:4px;height:4px;width:80%;'><div style='background:{conf_col};height:4px;border-radius:4px;width:{conf}%;'></div></div></div>"

                def _render_edit_col(ek,mk,info,is_title_sourced,selected_sheet,curr_claim_id,schema_field,active):
                    # Always resolve display value from session_state explicitly.
                    # Do NOT rely on key= for disabled widgets — Streamlit reads widget
                    # internal state which is blank on first render in the same cycle.
                    _display_val = st.session_state.get(mk, info.get("modified", info["value"])) or ""
                    if st.session_state[ek]:
                        with st.form(key=f"form_s_{selected_sheet}_{curr_claim_id}_{schema_field}",border=False):
                            nv = st.text_input("m", value=_display_val, label_visibility="collapsed")
                            submitted = st.form_submit_button("",use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv
                                if not is_title_sourced and excel_f in active["data"][st.session_state.selected_idx]:
                                    active["data"][st.session_state.selected_idx][excel_f]["modified"] = nv
                                st.session_state[ek] = False
                                _append_audit({"event":"FIELD_EDITED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"claim_id":curr_claim_id,"field":schema_field,"original":info["value"],"new_value":nv})
                                st.rerun()
                    else:
                        # Use value= not key= so display is always correct on first render
                        st.text_input("m", value=_display_val, key=f"disp_{mk}", label_visibility="collapsed", disabled=True)
                    if not is_title_sourced and excel_f in active["data"][st.session_state.selected_idx]:
                        active["data"][st.session_state.selected_idx][excel_f]["modified"] = _display_val

                def _render_edit_btn(ek,selected_sheet,curr_claim_id,schema_field):
                    if not st.session_state[ek]:
                        if st.button("✏",key=f"ed_s_{selected_sheet}_{curr_claim_id}_{schema_field}",use_container_width=True,help="Edit field"):
                            st.session_state[ek] = True; st.rerun()
                    else:
                        st.markdown("<div style='height:38px;display:flex;align-items:center;justify-content:center;color:var(--green);font-size:11px;border:1px solid var(--b0);border-radius:6px;'>↵</div>", unsafe_allow_html=True)

                if _show_conf:
                    cl,cc,co,cm,ce,cb,cx = st.columns([1.8,1.5,1.8,1.8,0.55,0.55,0.45],gap="small")
                    with cl: st.markdown(_field_label_html,unsafe_allow_html=True)
                    with cc: st.markdown(_conf_html,unsafe_allow_html=True)
                    with co: st.text_input("o",value=info["value"],key=f"orig_{selected_sheet}_{curr_claim_id}_schema_{schema_field}",label_visibility="collapsed",disabled=True)
                    with cm: _render_edit_col(ek,mk,info,is_title_sourced,selected_sheet,curr_claim_id,schema_field,active)
                    with ce:
                        if st.button("👁",key=f"eye_s_{selected_sheet}_{curr_claim_id}_{schema_field}",use_container_width=True):
                            show_eye_popup(schema_field,info,excel_path,selected_sheet)
                    with cb: _render_edit_btn(ek,selected_sheet,curr_claim_id,schema_field)
                    with cx: st.markdown("<div style='height:8px;'></div>",unsafe_allow_html=True); st.checkbox("",key=xk,label_visibility="collapsed")
                else:
                    cl,co,cm,ce,cb,cx = st.columns([1.8,1.8,1.8,0.55,0.55,0.45],gap="small")
                    with cl: st.markdown(_field_label_html,unsafe_allow_html=True)
                    with co: st.text_input("o",value=info["value"],key=f"orig_{selected_sheet}_{curr_claim_id}_schema_{schema_field}",label_visibility="collapsed",disabled=True)
                    with cm: _render_edit_col(ek,mk,info,is_title_sourced,selected_sheet,curr_claim_id,schema_field,active)
                    with ce:
                        if st.button("👁",key=f"eye_s_{selected_sheet}_{curr_claim_id}_{schema_field}",use_container_width=True):
                            show_eye_popup(schema_field,info,excel_path,selected_sheet)
                    with cb: _render_edit_btn(ek,selected_sheet,curr_claim_id,schema_field)
                    with cx: st.markdown("<div style='height:8px;'></div>",unsafe_allow_html=True); st.checkbox("",key=xk,label_visibility="collapsed")

        else:
            # PLAIN MODE
            _col_hdr2 = "<div style='font-size:var(--sz-xs);font-weight:700;color:var(--t1);text-transform:uppercase;letter-spacing:1.4px;font-family:var(--font-head);padding-bottom:5px;border-bottom:1px solid var(--b0);margin-bottom:5px;'>{}</div>"
            hc = st.columns([2,2.6,2.6,0.6,0.6,0.5])
            with hc[0]: st.markdown(_col_hdr2.format("Field"), unsafe_allow_html=True)
            with hc[1]: st.markdown(_col_hdr2.format("Extracted Value"), unsafe_allow_html=True)
            with hc[2]: st.markdown(_col_hdr2.format("Modified Value"), unsafe_allow_html=True)

            for field,info in curr_claim.items():
                ek = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_{field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_{field}"
                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state: st.session_state[mk] = info.get("modified",info["value"])

                cl,co,cm,ce,cb,cx = st.columns([2,2.6,2.6,0.9,0.9,0.5],gap="small")
                with cl:
                    _cv  = st.session_state.get(mk,info.get("modified",info["value"]))
                    _dot = "<span style='color:var(--yellow);margin-left:4px;font-size:8px;'>●</span>" if _cv!=info["value"] else ""
                    st.markdown(f"<div style='height:40px;display:flex;align-items:center;color:var(--t0);font-size:var(--sz-body);font-weight:600;text-transform:uppercase;letter-spacing:0.8px;font-family:var(--font-head);'>{field}{_dot}</div>", unsafe_allow_html=True)
                with co:
                    st.text_input("o",value=info["value"],key=f"orig_{selected_sheet}_{curr_claim_id}_{field}",label_visibility="collapsed",disabled=True)
                with cm:
                    _plain_display_val = st.session_state.get(mk, info.get("modified", info["value"])) or ""
                    if st.session_state[ek]:
                        with st.form(key=f"form_{selected_sheet}_{curr_claim_id}_{field}",border=False):
                            nv = st.text_input("m",value=_plain_display_val,label_visibility="collapsed")
                            submitted = st.form_submit_button("",use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv; active["data"][st.session_state.selected_idx][field]["modified"] = nv; st.session_state[ek] = False
                                _append_audit({"event":"FIELD_EDITED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"claim_id":curr_claim_id,"field":field,"original":info["value"],"new_value":nv})
                                st.rerun()
                    else:
                        st.text_input("m", value=_plain_display_val, key=f"disp_plain_{mk}", label_visibility="collapsed", disabled=True)
                    active["data"][st.session_state.selected_idx][field]["modified"] = _plain_display_val
                with ce:
                    if st.button("👁",key=f"eye_{selected_sheet}_{curr_claim_id}_{field}",use_container_width=True):
                        show_eye_popup(field,info,excel_path,selected_sheet)
                with cb:
                    if not st.session_state[ek]:
                        if st.button("✏",key=f"ed_{selected_sheet}_{curr_claim_id}_{field}",use_container_width=True,help="Edit field"):
                            st.session_state[ek] = True; st.rerun()
                    else:
                        st.markdown("<div style='height:38px;display:flex;align-items:center;justify-content:center;color:var(--green);font-size:11px;border:1px solid var(--b0);border-radius:6px;'>↵</div>", unsafe_allow_html=True)
                with cx:
                    st.markdown("<div style='height:8px;'></div>",unsafe_allow_html=True)
                    st.checkbox("",key=xk,label_visibility="collapsed")

        # ══════════════════════════════════════════
        # ✦ ADD CUSTOM FIELD
        # ══════════════════════════════════════════
        _user_fields_key = f"user_added_fields_{selected_sheet}_{curr_claim_id}"
        if _user_fields_key not in st.session_state:
            st.session_state[_user_fields_key] = []

        _add_counter_key = f"add_field_counter_{selected_sheet}_{curr_claim_id}"
        if _add_counter_key not in st.session_state:
            st.session_state[_add_counter_key] = 0
        _ctr = st.session_state[_add_counter_key]

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("""
        <div style='display:flex;align-items:center;gap:10px;margin-bottom:10px;'>
          <span style='font-size:var(--sz-xs);font-weight:600;color:var(--purple);text-transform:uppercase;letter-spacing:2px;font-family:var(--mono);'>+ Add Custom Field</span>
          <span style='flex:1;height:1px;background:linear-gradient(90deg,rgba(167,139,250,0.4),transparent);'></span>
        </div>
        """, unsafe_allow_html=True)

        _user_fields = st.session_state[_user_fields_key]
        if _user_fields:
            _col_hdr_cf = "<div style='font-size:var(--sz-xs);font-weight:600;color:var(--t3);text-transform:uppercase;letter-spacing:1.6px;font-family:var(--mono);padding-bottom:5px;border-bottom:1px solid var(--b0);margin-bottom:6px;'>{}</div>"
            uc1,uc2,uc3,uc4 = st.columns([2,3,0.6,0.6])
            with uc1: st.markdown(_col_hdr_cf.format("Custom Field"), unsafe_allow_html=True)
            with uc2: st.markdown(_col_hdr_cf.format("Value"), unsafe_allow_html=True)

            for uf_idx, uf in enumerate(_user_fields):
                uf_name  = uf["name"]
                uf_mk    = f"uf_mod_{selected_sheet}_{curr_claim_id}_{uf_name}_{uf_idx}"
                uf_ek    = f"uf_edit_{selected_sheet}_{curr_claim_id}_{uf_name}_{uf_idx}"
                if uf_mk not in st.session_state: st.session_state[uf_mk] = uf.get("value","")
                if uf_ek not in st.session_state: st.session_state[uf_ek] = False

                uc1b,uc2b,uc3b,uc4b = st.columns([2,3,0.6,0.6],gap="small")
                with uc1b:
                    st.markdown(f"<div style='min-height:40px;display:flex;align-items:center;gap:4px;color:var(--purple);font-size:var(--sz-xs);font-weight:600;text-transform:uppercase;letter-spacing:1px;font-family:var(--mono);'>{uf_name}</div>", unsafe_allow_html=True)
                with uc2b:
                    if st.session_state[uf_ek]:
                        with st.form(key=f"uf_form_{selected_sheet}_{curr_claim_id}_{uf_name}_{uf_idx}",border=False):
                            new_uf_val = st.text_input("v", value=st.session_state[uf_mk], label_visibility="collapsed")
                            if st.form_submit_button("", use_container_width=False):
                                st.session_state[uf_mk] = new_uf_val
                                st.session_state[_user_fields_key][uf_idx]["value"] = new_uf_val
                                st.session_state[uf_ek] = False
                                st.rerun()
                    else:
                        st.text_input("v", key=uf_mk, label_visibility="collapsed", disabled=True)
                        st.session_state[_user_fields_key][uf_idx]["value"] = st.session_state.get(uf_mk,"")
                with uc3b:
                    if not st.session_state[uf_ek]:
                        if st.button("✏", key=f"uf_ed_{selected_sheet}_{curr_claim_id}_{uf_name}_{uf_idx}", use_container_width=True):
                            st.session_state[uf_ek] = True; st.rerun()
                    else:
                        st.markdown("<div style='height:38px;display:flex;align-items:center;justify-content:center;color:var(--green);font-size:11px;border:1px solid var(--b0);border-radius:6px;'>↵</div>", unsafe_allow_html=True)
                with uc4b:
                    if st.button("🗑", key=f"uf_del_{selected_sheet}_{curr_claim_id}_{uf_name}_{uf_idx}", use_container_width=True, help="Remove field"):
                        st.session_state[_user_fields_key].pop(uf_idx); st.rerun()

            st.markdown("<div style='height:4px;'></div>", unsafe_allow_html=True)

        st.markdown("<div class='add-field-panel'>", unsafe_allow_html=True)
        _new_name_k  = f"nf_name_{selected_sheet}_{curr_claim_id}_{_ctr}"
        _new_val_k   = f"nf_val_{selected_sheet}_{curr_claim_id}_{_ctr}"
        af1, af2, af3 = st.columns([1.8, 2.5, 0.8], gap="small")
        with af1:
            new_field_name = st.text_input("Field name", key=_new_name_k, placeholder="e.g. Internal Notes", label_visibility="collapsed")
        with af2:
            new_field_value = st.text_input("Field value", key=_new_val_k, placeholder="Enter value…", label_visibility="collapsed")
        with af3:
            if st.button("＋ Add", key=f"add_field_go_{selected_sheet}_{curr_claim_id}_{_ctr}", use_container_width=True, type="primary"):
                fname_stripped = new_field_name.strip()
                if fname_stripped:
                    existing_names = {f["name"] for f in st.session_state[_user_fields_key]}
                    if fname_stripped not in existing_names:
                        st.session_state[_user_fields_key].append({"name": fname_stripped, "value": new_field_value.strip()})
                        st.session_state[_add_counter_key] = _ctr + 1
                        _append_audit({"event":"FIELD_ADDED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"claim_id":curr_claim_id,"field":fname_stripped,"value":new_field_value.strip()})
                        st.rerun()
                    else:
                        st.warning(f"Field '{fname_stripped}' already exists.")
                else:
                    st.warning("Please enter a field name.")
        st.markdown("</div>", unsafe_allow_html=True)

        if totals_data:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("<p class='section-lbl'>Sheet Totals</p>", unsafe_allow_html=True)
            agg = totals_data.get("aggregated",{})
            if agg:
                t_cols = st.columns(min(4,len(agg)))
                for idx,(k,v) in enumerate(agg.items()):
                    with t_cols[idx%len(t_cols)]:
                        st.markdown(f"<div style='background:var(--s0);border:1px solid var(--b0);border-top:2px solid var(--green);border-radius:8px;padding:10px 14px;margin-bottom:8px;'><div style='font-size:var(--sz-xs);color:var(--t2);text-transform:uppercase;font-family:var(--mono);letter-spacing:0.8px;'>{k}</div><div style='font-size:var(--sz-body);font-weight:700;color:var(--green);font-family:var(--mono);margin-top:2px;'>{v:,.2f}</div></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # RIGHT PANEL — Export
    # ══════════════════════════════════════════
    with col_fmt:
        st.markdown("<p class='section-lbl'>Export Format</p>", unsafe_allow_html=True)

        _active = st.session_state.get("active_schema",None)
        if _active and _active in SCHEMAS:
            _sc = SCHEMAS[_active]; _cf_count = len(st.session_state.get(f"custom_fields_{_active}",[]))
            date_fmt_right = _sc.get("date_format","YYYY-MM-DD")
            st.markdown(f"<div style='background:var(--s0);border:1px solid {_sc['color']}44;border-left:2px solid {_sc['color']};border-radius:7px;padding:10px 12px;margin-bottom:8px;'><div style='font-size:var(--sz-body);font-weight:700;color:{_sc['color']};'>{_sc['icon']} {_active}</div><div style='font-size:var(--sz-xs);color:var(--t2);margin-top:2px;font-family:var(--mono);'>{_sc['version']}</div><div style='font-size:var(--sz-xs);color:var(--t3);margin-top:2px;font-family:var(--mono);'>Date: {date_fmt_right} · Amounts: 2dp</div><div style='font-size:var(--sz-xs);color:var(--t2);margin-top:2px;font-family:var(--mono);'>Fields: {len(_sc['required_fields'])} req · {_cf_count} custom</div></div>", unsafe_allow_html=True)

        _use_conf_disp = st.session_state.get("use_conf_threshold",True); _conf = st.session_state.get("conf_threshold",80)
        if _use_conf_disp:
            _bc = "var(--green)" if _conf>=70 else "var(--yellow)" if _conf>=40 else "var(--red)"
            st.markdown(f"<div style='margin-bottom:10px;'><div style='font-size:var(--sz-xs);color:var(--t3);text-transform:uppercase;font-weight:600;margin-bottom:3px;font-family:var(--mono);letter-spacing:1px;'>Confidence Threshold</div><div style='display:flex;align-items:center;gap:8px;'><div class='conf-bar-wrap' style='flex:1;'><div class='conf-bar-fill' style='width:{_conf}%;background:{_bc};'></div></div><span style='color:{_bc};font-size:var(--sz-body);font-weight:700;'>{_conf}%</span></div></div>", unsafe_allow_html=True)

        # ── Cause of Loss enrichment status — shown first, most prominent ──
        _col_enriched   = st.session_state.get(f"_col_enriched_{selected_sheet}_{curr_claim_id}", False)
        _col_summary_rp = st.session_state.get(f"_col_summary_{selected_sheet}_{curr_claim_id}")
        _col_val_rp     = st.session_state.get(f"mod_{selected_sheet}_{curr_claim_id}_schema_Cause of Loss") or \
                          st.session_state.get(f"mod_{selected_sheet}_{curr_claim_id}_Cause of Loss") or \
                          st.session_state.get(f"mod_{selected_sheet}_{curr_claim_id}_Cause_of_Loss")

        if _col_enriched and (_col_summary_rp or _col_val_rp):
            st.markdown(
                f"<div style='background:rgba(52,211,153,0.07);border:1px solid rgba(52,211,153,0.3);"
                f"border-left:3px solid var(--green);border-radius:7px;padding:10px 12px;margin-bottom:10px;'>"
                f"<div style='font-size:10px;font-weight:700;color:var(--green);font-family:var(--mono);"
                f"text-transform:uppercase;letter-spacing:1px;margin-bottom:5px;'>✓ Cause of Loss Identified</div>"
                + (f"<div style='font-size:var(--sz-sm);color:var(--blue);font-family:var(--mono);"
                   f"font-weight:700;margin-bottom:4px;'>{_col_val_rp}</div>" if _col_val_rp else "")
                + (f"<div style='font-size:var(--sz-xs);color:var(--t2);font-family:var(--font);"
                   f"line-height:1.5;'>{_col_summary_rp}</div>" if _col_summary_rp else "")
                + f"</div>",
                unsafe_allow_html=True
            )
        elif _llm_available() and not _col_enriched:
            # Show spinner-like indicator while LLM is processing
            st.markdown(
                "<div style='background:rgba(79,156,249,0.06);border:1px solid rgba(79,156,249,0.2);"
                "border-radius:6px;padding:8px 10px;margin-bottom:10px;'>"
                "<div style='font-size:10px;color:var(--blue);font-family:var(--mono);"
                "text-transform:uppercase;letter-spacing:1px;'>⏳ Analysing cause of loss…</div>"
                "</div>",
                unsafe_allow_html=True
            )

        # ── Live JSON toggle ──
        st.markdown("<hr>", unsafe_allow_html=True)
        _json_toggle_key = f"show_live_json_{selected_sheet}_{curr_claim_id}"
        if _json_toggle_key not in st.session_state:
            st.session_state[_json_toggle_key] = False

        _json_btn_label = "▲ Hide Live JSON" if st.session_state[_json_toggle_key] else "{ } Preview JSON"
        if st.button(_json_btn_label, key=f"json_toggle_btn_{selected_sheet}_{curr_claim_id}", use_container_width=True):
            st.session_state[_json_toggle_key] = not st.session_state[_json_toggle_key]
            st.rerun()

        if st.session_state[_json_toggle_key]:
            # Build the live record
            _rp_live_record = {}
            _rp_schema = st.session_state.get("active_schema", None)
            if _rp_schema and _rp_schema in SCHEMAS:
                _rp_schema_def = SCHEMAS[_rp_schema]
                _rp_mapped     = map_claim_to_schema(curr_claim, _rp_schema, title_fields)
                _rp_cf         = st.session_state.get(f"custom_fields_{_rp_schema}", [])
                _rp_disp       = list(_rp_schema_def["required_fields"]) + [f for f in _rp_cf if f not in _rp_schema_def["required_fields"]]
                for sf in _rp_disp:
                    mk_rp = f"mod_{selected_sheet}_{curr_claim_id}_schema_{sf}"
                    if sf in _rp_mapped:
                        _rp_live_record[sf] = st.session_state.get(mk_rp, _rp_mapped[sf]["value"])
                    elif st.session_state.get(mk_rp):
                        _rp_live_record[sf] = st.session_state[mk_rp]
            else:
                for fld, inf in curr_claim.items():
                    mk_rp = f"mod_{selected_sheet}_{curr_claim_id}_{fld}"
                    _rp_live_record[fld] = st.session_state.get(mk_rp, inf.get("modified", inf["value"]))
            # Include user ad-hoc fields
            _rp_uf_key = f"user_added_fields_{selected_sheet}_{curr_claim_id}"
            for uf in st.session_state.get(_rp_uf_key, []):
                uf_idx_rp = st.session_state.get(_rp_uf_key, []).index(uf)
                uf_mk_rp  = f"uf_mod_{selected_sheet}_{curr_claim_id}_{uf['name']}_{uf_idx_rp}"
                _rp_live_record[uf["name"]] = st.session_state.get(uf_mk_rp, uf["value"])

            _rp_json = json.dumps(_sanitize_for_json(_rp_live_record), indent=2, ensure_ascii=False)
            st.markdown(
                f"<div class='json-live-panel' style='margin-top:6px;'>"
                f"<div class='json-live-header'>"
                f"<span style='font-size:var(--sz-xs);font-weight:600;color:var(--t2);font-family:var(--mono);'>"
                f"<span class='json-live-dot'></span>{curr_claim_id}</span>"
                f"<span style='font-size:10px;color:var(--t3);font-family:var(--mono);'>{len(_rp_live_record)} fields</span>"
                f"</div>"
                f"<div class='json-live-body' style='max-height:420px;'>{_rp_json}</div>"
                f"</div>",
                unsafe_allow_html=True
            )

        _sheet_meta = {"sheet_name":selected_sheet,"record_count":len(data)}

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<p class='section-lbl'>📄 Standard Export</p>", unsafe_allow_html=True)

        if st.button("⬇ Download Standard JSON",use_container_width=True,type="primary",key=f"export_std_json_{selected_sheet}"):
            _std_export_data = {}
            for i,row in enumerate(data):
                c_id = detect_claim_id(row,i); rec = {}
                for fld,inf in row.items():
                    if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}",True):
                        mk_key = f"mod_{selected_sheet}_{c_id}_{fld}"; live_val = st.session_state.get(mk_key,None)
                        orig = inf.get("value",""); final = live_val if live_val is not None else inf.get("modified",orig)
                        rec[fld] = {"value":final,"original":orig,"edited":final!=orig,"excel_row":inf.get("excel_row"),"excel_col":inf.get("excel_col"),"record_index":i}
                _uf_key = f"user_added_fields_{selected_sheet}_{c_id}"
                for uf in st.session_state.get(_uf_key, []):
                    uf_idx_e = st.session_state.get(_uf_key,[]).index(uf)
                    uf_mk_e = f"uf_mod_{selected_sheet}_{c_id}_{uf['name']}_{uf_idx_e}"
                    rec[uf["name"]] = {"value": st.session_state.get(uf_mk_e, uf["value"]), "original": "", "edited": True, "user_added": True, "excel_row": None, "excel_col": None, "record_index": i}
                _std_export_data[c_id] = clean_duplicate_fields(rec)
            output   = _sanitize_for_json(to_standard_json(_std_export_data,_sheet_meta,totals_data,merged_meta))
            json_str = json.dumps(output,indent=2,ensure_ascii=False)
            save_feature_store(selected_sheet,output)
            st.session_state[f"_std_json_ready_{selected_sheet}"] = json_str
            _append_json_export({"filename":uploaded.name,"sheet":selected_sheet,"timestamp":datetime.datetime.now().isoformat(),"type":"Standard","record_count":len(_std_export_data),"json":json_str})
            _append_audit({"event":"EXPORT_GENERATED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"export_type":"Standard JSON","records":len(_std_export_data)})

        if st.session_state.get(f"_std_json_ready_{selected_sheet}"):
            st.download_button("📥 Save Standard JSON",data=st.session_state[f"_std_json_ready_{selected_sheet}"],file_name=f"{selected_sheet}_standard.json",mime="application/json",use_container_width=True,key=f"dl_std_json_{selected_sheet}")

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<p class='section-lbl'>🔌 Schema Export</p>", unsafe_allow_html=True)
        _schema_sel = st.selectbox("Schema export format",options=["— Select schema format —","🔵 Guidewire (JSON)","🟡 Duck Creek (JSON)"],key=f"schema_export_sel_{selected_sheet}",label_visibility="collapsed")

        if _schema_sel and _schema_sel != "— Select schema format —":
            if st.button("⬇ Generate Export",use_container_width=True,key=f"schema_export_go_{selected_sheet}"):
                if "Guidewire" in _schema_sel:
                    recs = build_mapped_records_for_export(data,"Guidewire",selected_sheet)
                    for rec in recs:
                        _uf_key = f"user_added_fields_{selected_sheet}_{rec.get('_claim_id','')}"
                        for uf in st.session_state.get(_uf_key,[]):
                            uf_idx = st.session_state.get(_uf_key,[]).index(uf)
                            uf_mk_e = f"uf_mod_{selected_sheet}_{rec.get('_claim_id','')}_{uf['name']}_{uf_idx}"
                            rec[uf["name"]] = {"value": st.session_state.get(uf_mk_e, uf["value"]), "confidence": 100, "edited": True, "original": "", "user_added": True}
                    gj   = _sanitize_for_json(to_guidewire_json(recs,_sheet_meta))
                    json_str = json.dumps(gj,indent=2,ensure_ascii=False)
                    save_feature_store(selected_sheet,gj)
                    st.session_state[f"_schema_export_data_{selected_sheet}"] = {"data":json_str,"filename":f"{selected_sheet}_Guidewire_ClaimCenter.json","mime":"application/json","label":"📥 Save Guidewire JSON"}
                    etype = "Guidewire JSON"; rec_count = len(recs)
                elif "Duck Creek" in _schema_sel:
                    recs = build_mapped_records_for_export(data,"Duck Creek",selected_sheet)
                    for rec in recs:
                        _uf_key = f"user_added_fields_{selected_sheet}_{rec.get('_claim_id','')}"
                        for uf in st.session_state.get(_uf_key,[]):
                            uf_idx = st.session_state.get(_uf_key,[]).index(uf)
                            uf_mk_e = f"uf_mod_{selected_sheet}_{rec.get('_claim_id','')}_{uf['name']}_{uf_idx}"
                            rec[uf["name"]] = {"value": st.session_state.get(uf_mk_e, uf["value"]), "confidence": 100, "edited": True, "original": "", "user_added": True}
                    dj   = _sanitize_for_json(to_duck_creek_json(recs,_sheet_meta))
                    json_str = json.dumps(dj,indent=2,ensure_ascii=False)
                    save_feature_store(selected_sheet,dj)
                    st.session_state[f"_schema_export_data_{selected_sheet}"] = {"data":json_str,"filename":f"{selected_sheet}_DuckCreek.json","mime":"application/json","label":"📥 Save Duck Creek JSON"}
                    etype = "Duck Creek JSON"; rec_count = len(recs)
                else:
                    etype, json_str, rec_count = "Unknown", "{}", 0

                _append_json_export({"filename":uploaded.name,"sheet":selected_sheet,"timestamp":datetime.datetime.now().isoformat(),"type":etype,"record_count":rec_count,"json":json_str})
                _append_audit({"event":"EXPORT_GENERATED","timestamp":datetime.datetime.now().isoformat(),"filename":uploaded.name,"sheet":selected_sheet,"export_type":etype,"records":len(data)})
                st.success("✅ Ready!")

        _exp_ready = st.session_state.get(f"_schema_export_data_{selected_sheet}")
        if _exp_ready:
            st.download_button(_exp_ready["label"],data=_exp_ready["data"],file_name=_exp_ready["filename"],mime=_exp_ready["mime"],use_container_width=True,key=f"dl_schema_export_{selected_sheet}")

        st.markdown("<hr>", unsafe_allow_html=True)
        if merged_meta:
            st.markdown("<p class='section-lbl'>Merged Regions</p>", unsafe_allow_html=True)
            sorted_merges = sorted([(k,v) for k,v in merged_meta.items() if v["value"]],key=lambda x:(x[1]["row_start"],x[1]["col_start"]))
            for key,m in sorted_merges[:8]:
                type_color = "var(--blue)" if m["type"]=="TITLE" else "var(--yellow)" if m["type"]=="HEADER" else "var(--t3)"
                st.markdown(f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:6px;padding:6px 10px;margin-bottom:4px;'><div style='font-size:var(--sz-xs);color:{type_color};font-family:var(--mono);'>{m['type']} · R{m['row_start']}C{m['col_start']}→R{m['row_end']}C{m['col_end']}</div><div style='font-size:var(--sz-body);color:var(--t0);margin-top:2px;'>{m['value'][:35]}</div></div>", unsafe_allow_html=True)
