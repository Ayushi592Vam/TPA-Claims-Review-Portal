import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import re
import base64
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)

# ==============================
# LOGO HELPER  — always embedded as base64 so it shows in every env
# ==============================
def _load_logo_b64() -> str:
    """
    Try several locations for the ValueMomentum logo PNG.
    Falls back to the hardcoded base64 embedded at build time.
    """
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "valuemomentum_logo.png"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "valuemomentum_logo.png"),
        "valuemomentum_logo.png",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, "rb") as f:
                return base64.b64encode(f.read()).decode()
    return _LOGO_B64_EMBEDDED

# ── Logo embedded at build time — always available even without the PNG file ──
_LOGO_B64_EMBEDDED = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABBAKcDASIAAhEBAxEB/8QAHAABAAMAAwEBAAAAAAAAAAAAAAUGBwMECAEC/8QAPBAAAQMDAgMGAQoFBAMAAAAAAQIDBAAFBhESBxMhFCIxQVFhFQgWMlJicYGRodEjQnKxwRczNfA3dbL/xAAbAQEBAQADAQEAAAAAAAAAAAAAAQUDBAYCB//EACwRAAEEAQIEBQMFAAAAAAAAAAABAgMRBAUSITFBUQYTInGhYdHhBxWRscH/2gAMAwEAAhEDEQA/APZTaENtpbbQlCEgBKUjQADwAFfaUoBSlKAUpSgFKUoBSlKAUJ0oazniPmoYDlptDoLpG155J+j9ke9drExJMuRI40/BwZGQyBm5xYjmthTlbeNqmJTMdSS2T9BShpqgH161ZK8Z8S3nWpVtfbcWh1JWpK0nQg93rrWz8CuK7V/aax/IXkN3RCQlh5R0Eken9X969JqfhZ+PiNyse3JXqTqn19v6MbB1xss6wS8F6fY2WlRN+vsK0bEvErcWR3E+IHrXacucFu2G4rktpipTvLhPQCvDMzsd8r4WvRXN5p2PSPiexiSOSmr1OWfLjwYjkqU8lpltO5SlHoBXTx+9wb3DEmE4SnzSoaKHp0rEuImaSMklmPHK2bc2ruN69Vn6yv2qRxe4ybY3FlRV7VBA1HkoaeBroJqzXTU1PShjN1JHyq1qelDb6VGY9eI14hB5k7XAP4jZ8Un9qk62GuRyWhqNcjktBSlK+iivi9SkhJ0OnQ+lfaUBQ+FmUXW63G/WDIVtG6WqWUaoRsC2j9FWn4frXf4s5Q5ieGyLjF2qnOKSxEQoa7nFHQdPPzNVXLUHFeNdlyJPchXtHYJR8g4PoE/p+VfrN0jLOMVixpB3w7Oj4hNHlv8A5En9PzoQ0Owuy27BBXen2u2qYSZCtAhO8jU6Cu3Le0gPvMrBKW1KSoHUagVjGbtYmrP5rmc39+6JCQIdohJdUI6PMrCPPw8/OunwuucKLxCvVhx6RPNhftypDceWlaSyseIAX10q0U0LgnfrjkGBsXK8SufKXIdQVkBOoCiAOlXgkAak6CvMUB11rhhg5adWjdkqgdqiNRuPQ1vHFJSkcO78tCilQgukEHQju1CFk3Dbu1GnrX4Q+y4opbdbWoeISoEivPd8yN5jh7gWPPXKZBh3KPvnyI6VKdLSf5RpqetdbKVYDbrP8QwV++QL/E0WwtLMjR8g9Ur3DQ61aFmm8R8ufZ5lptAc5nVLzyUnu/ZH71lxZfJ1LLpP9Jrf8SnLu+M225yGOU9JjIccQU6aKI69PvqU2I+qPyrcwtZZiR7GRe63z+DLyNOdkP3Of8fk8acU23ELt5W2pIIc01Gn1apjLjjLqHWlqQ4hQUlSToQR4EV7ezzELTmFictlyaAOhLLyQNzS/JQP+K8f57iN1w2+uWy5tHbqSw+B3HkeoP8Aiv0fwzr0GfF5C+l6dO6fT/TyGs6VLiP81OLV69jTMEzpjI2kxb7NQxdGk9X3Tol9AHj/AFgeXnXLk2QOT2xboi1otzatQknq4r6x/asexr/nIv8AX/itCr8O/UDw7gaPrSy4bNqyt3L2tVVFrsi1ampHr+bmYTcaV1tb/K9r9hV5tjTptschtenLH8p9K7vC/A1XNbd4u7RTCSdWmlDQu+5+z/etnQ02hAQhCUpSNAAOgFY2n6Y97N71q+RoafgPVu93CzHrTLn2yYmTFS4lQ8RodFD0Nalj91ausMPISW3B0cbV4pP7VI7U+goAB4VtY2M6DhutDYhgWLrwPtKUruHYFKVX+JCinAL8pJIIgOkEHw7poCO4xY4/k2DS4cJG6ewRIiadDzEHUAff1FQ3BHHb5BTeMiyqOWb1dZAK0q01S2kAD/vsKqPyW7zKjGXjNxcUecyifDKjrqlXRQH6VbvlCXmTCw9uyW5RE+8vpitBJ6hPio/l0/GqQgrTGyfA84yOSnE5eQRLvI7QzKikFaOpIQrXwHX9K/eK2LM5PFGZlF+tQitTLWtttCFBQZ8ktk+aump++u78nOW1b+FLsm4ykttR5b/MddX0SAR5mpH/AFmwztGhXcOy79vbOyL5Ovru9KAqEfAMlc4OW6G3D5N8tVzVOYjukd/vE6fiDUtk94z/ADDE5dii4VLtL77CkyX5Lg26adUtjxJV4D760DIcwsNjs8S7zZgMGW6hpp5sbkkq8DqPL3quHjBhqbiiMp+YmOtzlpmqjKDBPsv096Ar0jCcjbwjDrpa4yE5Bj7feiOkDmpP0ka+tSMzJeIV+aZtllw2TYZS1pD86aUltlPntH81Sl54t4hbZzkbnTJiWTo8/FjqcabPuodPyq1N5DZXMeGQJuLHwwt83tBVokJ/75UBIRUONxmm3Xea4lACl6abjp1OnlXJWdM8ZMOXJShargzGWvamW5EWlk++7096smRZlj9hctouU0Nt3HXs7oGqCANdSfIaGoUsNV3PsRtOY2JdsubI18WXkjvsr8lA/wCKgIvF/D3rm1DW5OjNvL2NSX4qkMrPso+XvUNxp4mqxqZAtdoffZmCYyqUvswWhUdSSSEkggk9PDrXLDLJC9JI1pyclPiSNkjVY9LRTCLxiN2w3OWLXc2+m8qZeSO66jr1FbNwwwNdzW3d7u0pMIHcy0R1e9z9n+9WG7ZjgWQ2Ri5XWFJdjMTm2GufFUlSXleGnnp6+VaDMkR7bbHpTg2R4zRWoJT4JSNeg+4Vy69k/vWVFkzpxa1Gr2VUVVv55GLh6JHjSKt227RPuc6EJQgIQkJSkaAAeAr9VUXuIuMM4/ab6/LcagXV4tR3FtkdRrru9B3TUbC4vYdJujMJT8yMh9exmS/GU2ys+yj5e5rqm6aBSs+uHGDDoc5ccvTXmG18tyW1GUphKtdPpftV7gS40+GzMhvofjvIC23EHVKknwIoDmpSlAKr3Er/AMf3/wD9e9/8mrDUfktt+MY/PtXN5Pa462eZprt3DTXSgMKht/N7FuHeeNApRGQmHPI82Vk9T93WrSXG8y4zy5SdHbbjMIobI6pVIWNSfwB/SrUnBI6+FqcHky+ahMbkpkbNCFA6hWnsacM8FZwzHZNtEwzJMpxTj8lSdCokaDpr5CrZDFA64OA8aMVqRFlZKWpagdP4RUSdfbUCvRLlptHzdVbOyR/h3Zy3y9g27Nv7VV7Dw1t0TAZeI3SQZ0aS+48XAnYpJUdQR49R61CnhxmaoPwNXEOR8F02bBGHOLf1d+vp0oDMFBczgrFtzy1LhoyYR45J1/hknoPbrW18XbTb0cIrxDbiMoYjQiWUJQAEFOmmlL7w3t0vEbVjdsf7BGt0tuSlWzeVlJ1Ovh1OvjVkzCz/ADgxe4WXn8jtjCmuZt126+elLFFd4V2m3s8JrVGRFa5ciCFvApHfUoakn18axFTstPB6BbGEociqyhTCm3VlKFJ1JCFHySTXpHGLR8FxiDZedzuyx0s8zbpu0GmulVS08MLczhFwxa5ylS2ZcpckOoTsU0pR1BHj1FQETdmuJE6wP2iRiGKphOMFop7WrahOmmo6dNPGqVdsfmR4HDPH8iUxJUi4Otq5bnMQpvcCka+fTpV4Xw6zKVDFluHECQ7ZdNi0IjBLy2/qlevp0qdvPD+JKlYsqBJ7HGx93e2zs3cwaDprr08PGrYo6vHu3w3uE933MNjszSXGdEgbFJUNNPSqXxWUXcD4ePOaKcXOi7lHxPcrWc6sPzmxO4WLtPZu1tbObt3beoOun4VDZNgEW/YNbscfnOsPW8NKjS2095DiBoFafn0oKIL5RYbaxC1LCUoQm8RiogaAdTV2zlxDeE3la1BKRAd1JP2DVeewGZd8Jm45lOQPXZT7gWzK5QQpkgd3QefWoccNcquFuNmyHOn5lpSjYllpgIU5oO6Fq11IHTp56VClAkRWZ3CThjEkIC2XrwULSfMFbmorSPlFwYh4Rz9I7aezqZLW1IGzvpHT06HSuVvhlsxnFbL8V6WCb2rmcr/e6qO3TXp9KrLxExv524jLsPauy9oKP4uzdt2qCvD8KtkojodntzPCRFuRFa7P8J6p2jqS3qT9+vXWuh8ndxbnCW071FW3mJGvkAs6Crgm2bcbFn5vhE7PzNPs7ddKjuHONfNHEoth7X2vs5Uebs27tyifD8ahSxUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKA//2Q=="

_LOGO_B64 = _load_logo_b64()


def _logo_img_tag(height: int = 38) -> str:
    """Return an <img> tag with the logo embedded, or empty string."""
    if not _LOGO_B64:
        return ""
    return (
        f'<img src="data:image/png;base64,{_LOGO_B64}" '
        f'style="height:{height}px;margin-right:14px;vertical-align:middle;'
        f'border-radius:4px;background:white;padding:3px 6px;" />'
    )


# ==============================
# YAML CONFIG LOADER
# ==============================
CONFIG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")


def _parse_yaml_simple(text: str) -> dict:
    def _cast(v: str):
        v = v.strip()
        if not v or v.lower() in ("null", "~", ""):
            return None
        if v.lower() == "true":  return True
        if v.lower() == "false": return False
        try:    return int(v)
        except: pass
        try:    return float(v)
        except: pass
        return v.strip('"').strip("'")

    lines   = text.splitlines()
    root    = {}
    stack   = [(0, root)]
    cur_key = None

    for raw in lines:
        if not raw.strip() or raw.strip().startswith("#"):
            continue
        indent = len(raw) - len(raw.lstrip())
        line   = raw.strip()

        while len(stack) > 1 and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1]

        if line.startswith("- "):
            val = line[2:].strip()
            if cur_key and isinstance(parent, dict):
                if not isinstance(parent.get(cur_key), list):
                    parent[cur_key] = []
                parent[cur_key].append(_cast(val))
        elif ":" in line:
            parts = line.split(":", 1)
            key   = parts[0].strip().strip('"').strip("'")
            val   = parts[1].strip() if len(parts) > 1 else ""
            if " #" in val:
                val = val[:val.index(" #")].strip()
            cur_key = key
            if val:
                parent[key] = _cast(val)
            else:
                parent[key] = {}
                stack.append((indent + 2, parent[key]))

    return root


def load_schema_config(schema_filename: str) -> dict | None:
    path = os.path.join(CONFIG_DIR, schema_filename)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = _parse_yaml_simple(f.read())
        return raw
    except Exception:
        return None


def _merge_schema_from_config(hardcoded: dict, cfg: dict | None) -> dict:
    if not cfg:
        return hardcoded

    merged = dict(hardcoded)

    schema_block = cfg.get("schema", {})
    for k in ("version", "description"):
        if schema_block.get(k):
            merged[k] = schema_block[k]

    if cfg.get("required_fields"):
        rf = cfg["required_fields"]
        if isinstance(rf, dict):
            rf = list(rf.keys())
        if isinstance(rf, list):
            merged["required_fields"] = [str(f) for f in rf if f]

    if cfg.get("accepted_fields"):
        af = cfg["accepted_fields"]
        if isinstance(af, dict):
            af = list(af.keys())
        if isinstance(af, list):
            merged["accepted_fields"] = [str(f) for f in af if f]

    if cfg.get("field_aliases") and isinstance(cfg["field_aliases"], dict):
        aliases = {}
        for field, vals in cfg["field_aliases"].items():
            if isinstance(vals, list):
                aliases[field] = [str(v) for v in vals if v]
            elif isinstance(vals, str):
                aliases[field] = [vals]
        if aliases:
            merged["field_aliases"] = aliases

    conf_block = cfg.get("confidence", {})
    if isinstance(conf_block, dict):
        if conf_block.get("field_thresholds") and isinstance(conf_block["field_thresholds"], dict):
            merged["field_thresholds"] = {
                k: int(v) for k, v in conf_block["field_thresholds"].items()
                if v is not None
            }
        if conf_block.get("global_threshold") is not None:
            merged["config_threshold"] = int(conf_block["global_threshold"])

    if cfg.get("export") and isinstance(cfg["export"], dict):
        merged["export_config"] = cfg["export"]

    return merged


_CONFIG_LOAD_STATUS = {}


def _load_all_configs(hardcoded_schemas: dict) -> dict:
    filemap = {"Guidewire": "guidewire.yaml", "Duck Creek": "duck_creek.yaml"}
    result  = {}
    for name, schema in hardcoded_schemas.items():
        fname = filemap.get(name)
        cfg   = load_schema_config(fname) if fname else None
        result[name] = _merge_schema_from_config(schema, cfg)
        _CONFIG_LOAD_STATUS[name] = {
            "file":   fname,
            "loaded": cfg is not None,
            "path":   os.path.join(CONFIG_DIR, fname) if fname else "",
        }
    return result


# ==============================
# UNICODE NORMALIZER
# ==============================
_DASH_TABLE = str.maketrans({
    '\u2013': '-', '\u2014': '-', '\u2012': '-', '\u2015': '-',
    '\u2212': '-', '\ufe58': '-', '\ufe63': '-', '\uff0d': '-',
    '\u2018': "'", '\u2019': "'", '\u201c': '"', '\u201d': '"',
    '\u00a0': ' ', '\u202f': ' ',
})


def normalize_str(s: str) -> str:
    if not s:
        return s
    return s.translate(_DASH_TABLE)


# ==============================
# SCHEMA DEFINITIONS
# ==============================
_HARDCODED_SCHEMAS = {
    "Guidewire": {
        "color":   "#58a6ff",
        "icon":    "🔵",
        "css_cls": "guide",
        "version": "ClaimCenter 10.x",
        "description": "Guidewire ClaimCenter 10.x compatible format",
        "required_fields": [
            "Claim Number", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Status", "Line of Business", "Policy Number",
        ],
        "accepted_fields": [
            "Claim Number", "Claimant Name", "Loss Date", "Date Reported",
            "Total Incurred", "Total Paid", "Reserve", "Indemnity Paid",
            "Medical Paid", "Expense Paid", "Status", "Line of Business",
            "Policy Number", "Policy Period Start", "Policy Period End",
            "Carrier", "Insured Name", "Description of Loss",
            "Cause of Loss", "Litigation Flag", "Adjuster Name",
            "Adjuster Phone", "Branch Code", "Department Code",
            "Coverage Type", "Deductible", "Subrogation Amount",
            "Recovery Amount", "Open/Closed", "Reopen Date",
            "Last Activity Date", "Days Lost", "State", "Notes",
            "Job Title", "Body Part", "Vehicle ID", "At Fault",
            "Building Damage", "Contents Damage", "Business Interruption Loss",
            "Net Paid", "Services Involved", "Location",
        ],
        "field_aliases": {
            "Claim Number":      ["claim_id", "claim number", "claim no", "claim#",
                                  "claimid", "claim ref", "claim #"],
            "Claimant Name":     ["claimant name", "claimant", "insured name", "name",
                                  "injured party", "employee name", "driver name"],
            "Loss Date":         ["date of loss", "loss date", "loss dt", "date of accident",
                                  "incident date", "date of injury", "injury date",
                                  "date of incident"],
            "Date Reported":     ["date reported", "reported date", "report date"],
            "Total Incurred":    ["total incurred", "incurred", "total incurred amount"],
            "Total Paid":        ["total paid", "amount paid", "paid amount", "net paid"],
            "Reserve":           ["reserve", "outstanding reserve", "case reserve"],
            "Indemnity Paid":    ["indemnity paid", "indemnity", "wage loss paid",
                                  "ttd paid", "bi paid"],
            "Medical Paid":      ["medical paid", "medical", "med paid"],
            "Expense Paid":      ["expense paid", "expense", "legal expense", "defense costs"],
            "Status":            ["status", "claim status", "open/closed"],
            "Line of Business":  ["line of business", "lob", "coverage line"],
            "Policy Number":     ["policy number", "policy no", "policy#",
                                  "policy id", "policy :", "policy #"],
            "Insured Name":      ["insured name", "insured", "employer name"],
            "Description of Loss": ["description of loss", "loss description", "description",
                                    "narrative", "nature of injury", "nature of claim",
                                    "type of loss", "cause of loss"],
            "Cause of Loss":     ["cause of loss", "cause", "type of loss", "peril",
                                  "nature of injury", "nature of claim"],
            "Adjuster Name":     ["adjuster name", "adjuster", "examiner"],
            "Coverage Type":     ["coverage", "coverage type"],
            "Deductible":        ["deductible", "deductible amount"],
            "Days Lost":         ["days lost", "days of disability", "lost days",
                                  "disability days", "days missed"],
            "Job Title":         ["job title", "occupation", "position", "employee title"],
            "Body Part":         ["body part", "body part injured", "part of body"],
            "Vehicle ID":        ["vehicle id", "vehicle", "unit number", "vin"],
            "At Fault":          ["at fault", "fault", "liable", "at-fault"],
            "Building Damage":   ["building damage", "structure damage", "building loss"],
            "Contents Damage":   ["contents damage", "contents loss", "stock loss"],
            "Business Interruption Loss": ["bi loss", "business interruption",
                                           "business income loss"],
            "Net Paid":          ["net paid", "pd paid", "property damage paid",
                                  "net claim payment"],
            "Services Involved": ["services involved", "professional services", "service type"],
            "Location":          ["location", "property location", "site", "premises"],
        },
    },
    "Duck Creek": {
        "color":   "#f0883e",
        "icon":    "🟠",
        "css_cls": "duck",
        "version": "Claims 6.x",
        "description": "Duck Creek Claims 6.x transaction format",
        "required_fields": [
            "Claim Id", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Policy Number", "Claim Status",
        ],
        "accepted_fields": [
            "Claim Id", "Transaction Id", "Claimant Name", "Loss Date",
            "Date Reported", "Total Incurred", "Total Paid", "Reserve",
            "Indemnity Paid", "Medical Paid", "Expense Paid",
            "Policy Number", "Policy Effective Date", "Policy Expiry Date",
            "Claim Status", "Cause of Loss", "Description of Loss",
            "Insured Name", "Carrier Name", "Line of Business",
            "Adjuster Id", "Adjuster Name", "Office Code",
            "Jurisdiction", "State Code", "Deductible Amount",
            "Subrogation Flag", "Recovery Amount", "Litigation Flag",
            "Date Closed", "Date Reopened", "Last Updated Date", "Days Lost",
            "Notes", "Job Title", "Body Part", "Vehicle ID", "At Fault",
            "Building Damage", "Contents Damage", "Business Interruption Loss",
            "Net Paid", "Services Involved", "Property Location", "Coverage",
        ],
        "field_aliases": {
            "Claim Id":          ["claim_id", "claim number", "claim no", "claim#",
                                  "claimid", "claim ref", "claim #"],
            "Claimant Name":     ["claimant name", "claimant", "insured name", "name",
                                  "injured party", "employee name", "driver name"],
            "Loss Date":         ["date of loss", "loss date", "loss dt", "date of accident",
                                  "incident date", "date of injury", "injury date",
                                  "date of incident"],
            "Date Reported":     ["date reported", "reported date", "report date"],
            "Total Incurred":    ["total incurred", "incurred", "total incurred amount"],
            "Total Paid":        ["total paid", "amount paid", "paid amount"],
            "Reserve":           ["reserve", "outstanding reserve", "case reserve"],
            "Indemnity Paid":    ["indemnity paid", "indemnity", "wage loss paid",
                                  "ttd paid", "bi paid"],
            "Medical Paid":      ["medical paid", "medical", "med paid"],
            "Expense Paid":      ["expense paid", "expense", "legal expense", "defense costs"],
            "Claim Status":      ["status", "claim status", "open/closed"],
            "Line of Business":  ["line of business", "lob", "coverage line"],
            "Policy Number":     ["policy number", "policy no", "policy#",
                                  "policy id", "policy :", "policy #"],
            "Insured Name":      ["insured name", "insured", "employer name"],
            "Description of Loss": ["description of loss", "loss description", "description",
                                    "narrative", "nature of injury", "nature of claim",
                                    "type of loss"],
            "Cause of Loss":     ["cause of loss", "cause", "type of loss", "peril",
                                  "nature of injury", "nature of claim"],
            "Carrier Name":      ["carrier", "carrier name", "insurance company"],
            "Deductible Amount": ["deductible", "deductible amount"],
            "Jurisdiction":      ["state", "state code", "jurisdiction"],
            "Days Lost":         ["days lost", "days of disability", "lost days",
                                  "disability days", "days missed"],
            "Job Title":         ["job title", "occupation", "position", "employee title"],
            "Body Part":         ["body part", "body part injured", "part of body"],
            "Vehicle ID":        ["vehicle id", "vehicle", "unit number", "vin"],
            "At Fault":          ["at fault", "fault", "liable", "at-fault"],
            "Building Damage":   ["building damage", "structure damage", "building loss"],
            "Contents Damage":   ["contents damage", "contents loss", "stock loss"],
            "Business Interruption Loss": ["bi loss", "business interruption",
                                           "business income loss"],
            "Net Paid":          ["net paid", "pd paid", "property damage paid",
                                  "net claim payment"],
            "Services Involved": ["services involved", "professional services", "service type"],
            "Property Location": ["location", "property location", "site", "premises"],
            "Coverage":          ["coverage", "coverage type", "type of coverage",
                                  "subject to $50k sir", "within policy limits",
                                  "coverage under review"],
        },
    },
}

SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)

# ==============================
# SCHEMA MAPPING + CONFIDENCE ENGINE
# ==============================
def _word_tokens(s: str) -> set:
    stopwords = {"of", "the", "a", "an", "and", "or", "to", "in", "for"}
    words = re.sub(r"[_/#+]", " ", s.lower()).split()
    return {w for w in words if len(w) > 1 and w not in stopwords}


def _str_similarity(a: str, b: str) -> float:
    a_tok = _word_tokens(a)
    b_tok = _word_tokens(b)
    if not a_tok or not b_tok:
        return 0.0
    if a_tok == b_tok:
        return 1.0
    intersection = a_tok & b_tok
    union        = a_tok | b_tok
    return len(intersection) / len(union)


def _header_match_score(excel_col: str, schema_field: str, aliases: list) -> float:
    ec_norm = excel_col.lower().replace("_", " ").strip()
    for alias in aliases:
        if ec_norm == alias.lower():
            return 1.0
    best = max((_str_similarity(ec_norm, a.lower()) for a in aliases), default=0.0)
    return max(best, _str_similarity(ec_norm, schema_field.lower()))


def _value_quality_score(value: str, schema_field: str) -> float:
    if not value or not value.strip():
        return 0.0
    v  = value.strip()
    sf = schema_field.lower()

    if any(x in sf for x in ["date", "loss dt"]):
        date_patterns = [
            r"\d{2}-\d{2}-\d{4}", r"\d{4}-\d{2}-\d{2}",
            r"\d{2}/\d{2}/\d{4}", r"\d{1,2}/\d{1,2}/\d{2,4}",
        ]
        for p in date_patterns:
            if re.fullmatch(p, v):
                return 1.0
        return 0.4

    if any(x in sf for x in ["incurred", "paid", "reserve", "amount",
                               "deductible", "recovery"]):
        clean = v.replace(",", "").replace("$", "").replace("(", "-").replace(")", "")
        try:
            float(clean)
            return 1.0
        except ValueError:
            return 0.3

    if any(x in sf for x in ["id", "number", "no", "code"]):
        return 0.9 if len(v) >= 2 else 0.5

    if "status" in sf:
        known = {"open", "closed", "pending", "reopened", "denied", "settled"}
        return 1.0 if v.lower() in known else 0.7

    return 0.85 if len(v) > 0 else 0.0


_MIN_HEADER_MATCH = 0.70


def map_claim_to_schema(claim: dict, schema_name: str,
                        title_fields: dict = None) -> dict:
    if schema_name not in SCHEMAS:
        return {}

    schema       = SCHEMAS[schema_name]
    aliases      = schema.get("field_aliases", {})
    accepted     = schema["accepted_fields"]
    title_fields = title_fields or {}
    result       = {}

    for schema_field in accepted:
        field_aliases  = aliases.get(schema_field, [schema_field.lower()])
        best_excel_col = None
        best_header_sc = 0.0
        best_info      = None

        for excel_col, info in claim.items():
            h_sc = _header_match_score(excel_col, schema_field, field_aliases)
            if h_sc > best_header_sc:
                best_header_sc = h_sc
                best_excel_col = excel_col
                best_info      = info

        if best_header_sc >= _MIN_HEADER_MATCH and best_info is not None:
            val  = best_info.get("modified", best_info.get("value", ""))
            v_sc = _value_quality_score(val, schema_field)
            conf = round(best_header_sc * 0.40 * 100 + v_sc * 0.60 * 100)
            result[schema_field] = {
                "excel_field":  best_excel_col,
                "value":        val,
                "header_score": round(best_header_sc * 100),
                "value_score":  round(v_sc * 100),
                "confidence":   conf,
                "is_required":  schema_field in schema["required_fields"],
                "info":         best_info,
                "from_title":   False,
            }

        elif schema_field in title_fields:
            tf   = title_fields[schema_field]
            val  = tf.get("value", "")
            v_sc = _value_quality_score(val, schema_field)
            conf = min(95, round(1.0 * 0.40 * 100 + v_sc * 0.60 * 100))
            result[schema_field] = {
                "excel_field":  f"[title row {tf['excel_row']}]",
                "value":        val,
                "header_score": 100,
                "value_score":  round(v_sc * 100),
                "confidence":   conf,
                "is_required":  schema_field in schema["required_fields"],
                "info":         tf,
                "from_title":   True,
            }

    return result


def extract_title_fields(merged_meta: dict) -> dict:
    """
    Extract policy-level fields from merged title/header rows.
    """
    found = {}

    title_rows = sorted(
        [v for v in merged_meta.values()
         if v.get("value") and v["type"] in ("TITLE", "HEADER")],
        key=lambda x: (x["row_start"], x["col_start"])
    )

    for m in title_rows:
        text = str(m["value"]).strip()
        r, c = m["excel_row"], m["excel_col"]

        def _info(val):
            return {"value": val, "original": val, "modified": val,
                    "source": "title_row", "excel_row": r, "excel_col": c,
                    "title_text": text}

        pol = re.search(
            r'Policy\s*(?:#|No\.?|Number)?\s*[:\-]\s*([A-Z0-9][A-Z0-9\-/\.]+)',
            text, re.IGNORECASE)
        if pol and "Policy Number" not in found:
            found["Policy Number"] = _info(pol.group(1).strip())

        ins = re.search(r'Insured\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if ins and "Insured Name" not in found:
            found["Insured Name"] = _info(ins.group(1).strip())

        carr = re.search(r'Carrier\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if carr:
            val = carr.group(1).strip()
            if "Carrier" not in found:
                found["Carrier"] = _info(val)
            if "Carrier Name" not in found:
                found["Carrier Name"] = _info(val)

        state = re.search(r'State\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if state:
            val = state.group(1).strip()
            if "State" not in found:
                found["State"] = _info(val)
            if "Jurisdiction" not in found:
                found["Jurisdiction"] = _info(val)
            if "State Code" not in found:
                found["State Code"] = _info(val)

        period = re.search(
            r'Period\s*[:\-]?\s*'
            r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
            r'[\s\u2013\u2014\-to]+'
            r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})',
            text, re.IGNORECASE
        )
        if period:
            s, e = period.group(1).strip(), period.group(2).strip()
            for k, v in [("Policy Period Start", s), ("Policy Period End", e),
                         ("Policy Effective Date", s), ("Policy Expiry Date", e)]:
                if k not in found:
                    found[k] = _info(v)

        retro = re.search(
            r'Retroactive\s+Date\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if retro and "Reopen Date" not in found:
            found["Reopen Date"] = _info(retro.group(1).strip())

        lob_map = [
            (r"workers[\'\'\u2019\s\-]*compensation",        "Workers Compensation"),
            (r"workers[\s\-]*comp\b",                          "Workers Compensation"),
            (r"\bW\.?C\.?\b(?:\s+loss|\s+claim|\s+run)?",  "Workers Compensation"),
            (r"commercial\s+general\s+liability",               "Commercial General Liability"),
            (r"\bC\.?G\.?L\.?\b",                            "Commercial General Liability"),
            (r"commercial\s+auto(?:mobile|motive)?",             "Commercial Auto"),
            (r"business\s+auto",                                 "Commercial Auto"),
            (r"fleet.*auto|auto.*fleet",                          "Commercial Auto"),
            (r"commercial\s+real\s+estate",                     "Commercial Property"),
            (r"commercial\s+prop(?:erty)?",                      "Commercial Property"),
            (r"building\s+&\s+property",                        "Commercial Property"),
            (r"professional\s+liability",                        "Professional Liability"),
            (r"\bE\.?\s*&\s*O\.?\b",                        "Professional Liability"),
            (r"errors?\s+[&and]+\s+omissions?",                 "Professional Liability"),
            (r"directors?\s+[&and]+\s+officers?|\bD\.?&?O\.?\b", "Directors & Officers"),
            (r"cyber\s+liability|cyber\s+risk",                 "Cyber Liability"),
            (r"employment\s+practices?|\bE\.?P\.?L\.?\b",   "Employment Practices Liability"),
            (r"medical\s+malpractice|\bmed\.?\s*mal\b",      "Medical Malpractice"),
            (r"inland\s+marine",                                 "Inland Marine"),
            (r"products?\s+liability",                           "Products Liability"),
            (r"crime|fidelity",                                   "Crime / Fidelity"),
            (r"\bumbrella\b",                                   "Umbrella"),
            (r"excess\s+liability",                              "Excess Liability"),
            (r"general\s+liability|\bG\.?L\.?\b",            "General Liability"),
        ]
        for pattern, lob_val in lob_map:
            if re.search(pattern, text, re.IGNORECASE) and "Line of Business" not in found:
                found["Line of Business"] = _info(lob_val)
                break

    return found


# ==============================
# SETTINGS DIALOG
# ==============================
@st.dialog("Settings", width="large")
def show_settings_dialog():
    global SCHEMAS
    st.markdown("### Configuration")
    st.markdown("---")

    st.markdown("#### Confidence Settings")
    use_conf = st.checkbox(
        "Enable configurable confidence threshold",
        value=st.session_state.get("use_conf_threshold", True),
        key="use_conf_toggle",
        help="When enabled, fields below the threshold are flagged for manual review."
    )
    st.session_state["use_conf_threshold"] = use_conf

    if use_conf:
        st.markdown(
            "<div style='color:#8b949e;font-size:13px;margin-bottom:8px;'>"
            "Fields with confidence below this threshold will be flagged for manual review."
            "</div>",
            unsafe_allow_html=True
        )
        conf = st.slider(
            "Confidence threshold", min_value=0, max_value=100,
            value=st.session_state.get("conf_threshold", 80),
            step=5, format="%d%%", label_visibility="visible"
        )
        st.session_state["conf_threshold"] = conf

        bar_color = "#3fb950" if conf >= 70 else "#d29922" if conf >= 40 else "#f85149"
        level_txt = (
            "High confidence — minimal manual review needed" if conf >= 70
            else "Medium — review flagged fields carefully" if conf >= 40
            else "Low — most fields will require manual review"
        )
        st.markdown(
            f"<div class=\"conf-bar-wrap\">"
            f"<div class=\"conf-bar-fill\" style=\"width:{conf}%;background:{bar_color};\"></div>"
            f"</div>"
            f"<div style=\"color:{bar_color};font-size:12px;margin-top:5px;\">{level_txt}</div>",
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            "<div style='color:#8b949e;font-size:12px;padding:6px 10px;background:#161b22;"
            "border:1px solid #30363d;border-radius:6px;'>"
            "⚠ Confidence threshold is disabled — no fields will be flagged for review."
            "</div>",
            unsafe_allow_html=True
        )

    st.markdown("---")
    st.markdown("#### Export Schema")
    st.markdown(
        "<div style='color:#8b949e;font-size:13px;margin-bottom:12px;'>"
        "Activate a schema to map extracted fields to a standard format. "
        "Custom fields can be added per schema."
        "</div>",
        unsafe_allow_html=True
    )

    active_schema = st.session_state.get("active_schema", None)

    for schema_name, schema_def in SCHEMAS.items():
        is_active  = active_schema == schema_name
        border_col = schema_def["color"] if is_active else "#30363d"
        bg_col     = "#1c2128" if is_active else "#161b22"
        active_tag = (
            f"<span style=\"font-size:10px;color:{schema_def['color']};"
            f"margin-left:8px;font-weight:bold;\">● ACTIVE</span>"
            if is_active else ""
        )
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}", []))

        st.markdown(
            f"<div style=\"background:{bg_col};border:1px solid {border_col};"
            f"border-radius:8px;padding:12px 14px;margin-bottom:4px;\">"
            f"<div style=\"display:flex;align-items:center;\">"
            f"<span style=\"font-size:15px;font-weight:bold;color:white;\">"
            f"{schema_def['icon']} {schema_name}</span>"
            f"<span style=\"font-size:11px;color:#8b949e;margin-left:8px;\">"
            f"{schema_def['version']}</span>"
            f"{active_tag}</div>"
            f"<div style=\"font-size:12px;color:#8b949e;margin-top:4px;\">"
            f"{schema_def['description']}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            lbl = "✓ Deactivate" if is_active else "Activate"
            if st.button(lbl, key=f"activate_{schema_name}", use_container_width=True):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(f"Custom Fields ({custom_count})",
                         key=f"custom_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()

        st.markdown("<div style=\"height:6px;\"></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(
        f"<div style='color:#8b949e;font-size:13px;margin-bottom:8px;'>"
        f"Config directory: <code>{CONFIG_DIR}</code>"
        f"</div>",
        unsafe_allow_html=True
    )
    for schema_name, status in _CONFIG_LOAD_STATUS.items():
        sc     = SCHEMAS.get(schema_name, {})
        col_st = sc.get("color", "#8b949e")
        if status["loaded"]:
            badge = ("<span style=\"background:#1c2d1c;border:1px solid #3fb950;"
                     "border-radius:4px;padding:1px 7px;font-size:10px;"
                     "color:#3fb950;\">✓ Loaded</span>")
        else:
            badge = ("<span style=\"background:#2d1515;border:1px solid #f85149;"
                     "border-radius:4px;padding:1px 7px;font-size:10px;"
                     "color:#f85149;\">✗ Not found — using defaults</span>")
        st.markdown(
            f"<div style=\"background:#161b22;border:1px solid #30363d;"
            f"border-radius:6px;padding:8px 12px;margin-bottom:6px;\">"
            f"<div style=\"display:flex;align-items:center;gap:8px;\">"
            f"<span style=\"color:{col_st};font-weight:bold;font-size:13px;\">"
            f"{sc.get('icon','')} {schema_name}</span>{badge}</div>"
            f"<div style=\"font-size:10px;color:#8b949e;margin-top:3px;\">"
            f"📄 {status['file']}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.session_state["sheet_cache"] = {}
        st.success("✅ Configs reloaded — sheet cache cleared")
        st.rerun()

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("Reset Defaults", use_container_width=True):
            st.session_state["conf_threshold"]      = 80
            st.session_state["use_conf_threshold"]  = True
            st.session_state["active_schema"]       = None
            for s in SCHEMAS:
                st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ==============================
# SCHEMA FIELD MANAGER DIALOG
# ==============================
@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name):
    schema     = SCHEMAS[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state:
        st.session_state[custom_key] = []

    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(
        f"<div style='color:#8b949e;font-size:13px;margin-bottom:12px;'>"
        f"{schema['description']}</div>",
        unsafe_allow_html=True
    )

    tab_req, tab_accepted, tab_custom = st.tabs([
        "Mandatory Fields", "All Accepted Fields", "My Custom Fields"
    ])

    with tab_req:
        st.markdown(
            "These fields are **mandatory** by the schema and will always be "
            "included when this schema is active."
        )
        pills = "".join(
            f"<span class=\"field-pill field-pill-required\">✓ {f}</span>"
            for f in schema["required_fields"]
        )
        st.markdown(f"<div style=\"margin:12px 0;\">{pills}</div>", unsafe_allow_html=True)

    with tab_accepted:
        st.markdown(
            "All fields **accepted** by this schema. Only these fields can be "
            "added as custom fields."
        )
        optional = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(
            f"<span class=\"field-pill field-pill-required\">✓ {f}</span>"
            for f in schema["required_fields"]
        )
        opt_pills = "".join(
            f"<span class=\"field-pill\">{f}</span>"
            for f in optional
        )
        st.markdown(
            f"<div style=\"margin:10px 0;\">"
            f"<b style=\"color:#8b949e;font-size:11px;\">MANDATORY</b><br>{req_pills}</div>"
            f"<div style=\"margin:10px 0;\">"
            f"<b style=\"color:#8b949e;font-size:11px;\">OPTIONAL</b><br>{opt_pills}</div>",
            unsafe_allow_html=True
        )

    with tab_custom:
        st.markdown(
            "Select **optional fields** from the accepted list to include "
            "alongside mandatory fields in the export."
        )
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]

        if available:
            st.markdown("#### Add Optional Field")
            sel_col, add_col = st.columns([4, 1])
            with sel_col:
                chosen = st.selectbox(
                    "Pick field",
                    options=["— select a field —"] + available,
                    key=f"new_field_sel_{schema_name}",
                    label_visibility="collapsed"
                )
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}",
                             use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()
        else:
            st.info("All accepted optional fields have already been added.")

        st.markdown("#### Active Custom Fields")
        if not custom_fields:
            st.markdown(
                "<div style=\"color:#8b949e;font-size:13px;padding:8px 0;\">"
                "No optional fields added yet."
                "</div>",
                unsafe_allow_html=True
            )
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5, 1])
                with cf1:
                    is_req = cf in schema["required_fields"]
                    cls    = "field-pill-required" if is_req else "field-pill-custom"
                    st.markdown(
                        f"<span class=\"field-pill {cls}\">{'✓' if is_req else '+'} {cf}</span>",
                        unsafe_allow_html=True
                    )
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}",
                                 use_container_width=True):
                        st.session_state[custom_key].pop(idx)
                        st.rerun()
            st.markdown("---")
            if st.button("Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []
                st.rerun()

        st.markdown("---")
        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(
            f"<div style=\"background:#161b22;border:1px solid #30363d;"
            f"border-radius:8px;padding:10px 14px;\">"
            f"<span style=\"color:#8b949e;font-size:12px;\">"
            f"Mandatory: <b style=\"color:#58a6ff;\">{len(schema['required_fields'])}</b>"
            f" &nbsp;|&nbsp; "
            f"Custom: <b style=\"color:#3fb950;\">{len(custom_fields)}</b>"
            f" &nbsp;|&nbsp; "
            f"Total export fields: <b style=\"color:white;\">{total}</b>"
            f"</span></div>",
            unsafe_allow_html=True
        )


# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="Loss Run Parser")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
    /* ── Typography ── */
    h1,h2,h3,h4,h5,h6,
    .main-title, .mid-header-title, .sheet-title-value, .sheet-title-label,
    .incurred-label, .incurred-amount,
    [class*="heading"], [class*="-title"], [class*="-label"] {
        font-family: "Segoe UI", "Segoe UI Variable", Arial, sans-serif !important;
    }

    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title {
        font-size: 26px; font-weight: 600; padding: 10px 0;
        border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white;
        text-shadow: 0 0 10px rgba(88,166,255,0.7);
        font-family: "Segoe UI", Arial, sans-serif !important;
    }
    .sheet-title-banner {
        background: #161b22;
        border: 1px solid #30363d;
        border-left: 4px solid #58a6ff;
        border-radius: 6px;
        padding: 10px 16px;
        margin-bottom: 14px;
    }
    .sheet-title-label  { font-size:10px;color:#8b949e;text-transform:uppercase;font-weight:bold;letter-spacing:1px;margin-bottom:3px; }
    .sheet-title-value  { font-size:15px;color:#e6edf3;font-weight:600; }
    .sheet-subtitle-value { font-size:12px;color:#8b949e;margin-top:3px; }
    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; cursor: pointer;
        transition: all .25s ease;
    }
    .claim-card:hover { border-color:#58a6ff; box-shadow:0 0 12px rgba(88,166,255,0.6); transform:translateY(-2px); }
    .selected-card { border-left:4px solid #58a6ff;background:#1c2128;box-shadow:0 0 16px rgba(88,166,255,0.8); }
    .status-text     { font-size:12px;color:#3fb950;margin-top:5px; }
    .status-progress { font-size:12px;color:#d29922;margin-top:5px; }
    .mid-header-title  { font-size:26px;font-weight:bold;color:white;margin-bottom:0px; }
    .mid-header-sub    { font-size:15px;color:#8b949e;margin-top:5px;margin-bottom:5px; }
    .mid-header-status { font-size:13px;color:#3fb950;margin-bottom:15px; }
    .incurred-label    { font-size:14px;color:#8b949e;margin-bottom:0px; }
    .incurred-amount   { font-size:26px;font-weight:bold;color:#3fb950;margin-top:0px;margin-bottom:20px; }
    div[data-baseweb="input"],div[data-baseweb="base-input"],div[data-baseweb="select"] {
        background-color:#161b22!important;border:1px solid #30363d!important;border-radius:6px!important;
    }
    div[data-baseweb="input"] input {
        color:#ffffff!important;-webkit-text-fill-color:#ffffff!important;
        background-color:transparent!important;font-size:15px!important;padding:8px 12px!important;
    }
    div[data-baseweb="input"]:has(input:disabled),div[data-baseweb="base-input"]:has(input:disabled) {
        background-color:transparent!important;border:none!important;
    }
    div[data-baseweb="input"] input:disabled {
        color:#e6edf3!important;-webkit-text-fill-color:#e6edf3!important;
        cursor:default!important;padding-left:0px!important;
    }
    div[data-testid="stButton"] button {
        background-color:transparent!important;color:#8b949e!important;
        border:1px solid #30363d!important;border-radius:6px!important;
        padding:2px 8px!important;transition:0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color:#58a6ff!important;color:#58a6ff!important;background-color:#1c2128!important;
    }
    div[data-testid="stButton"] button:disabled { opacity:0.3!important; }
    div[role="dialog"] { background-color:#0d1117!important;border:1px solid #30363d!important;border-radius:10px!important; }
    div[role="dialog"] * { color:#c9d1d9!important; }
    div[role="dialog"] button { background-color:transparent!important;border:1px solid #30363d!important;color:#8b949e!important; }
    div[role="dialog"] button:hover { border-color:#58a6ff!important;color:#58a6ff!important;background-color:#1c2128!important; }
    .conf-bar-wrap { background:#21262d;border-radius:6px;height:8px;width:100%;margin-top:4px;overflow:hidden; }
    .conf-bar-fill { height:100%;border-radius:6px;background:linear-gradient(90deg,#3fb950,#58a6ff);transition:width 0.3s ease; }
    .field-pill { display:inline-block;background:#161b22;border:1px solid #30363d;border-radius:12px;padding:3px 10px;font-size:11px;color:#c9d1d9;margin:2px 3px; }
    .field-pill-required { border-color:#58a6ff!important;color:#58a6ff!important;background:#1c2128!important; }
    .field-pill-custom   { border-color:#3fb950!important;color:#3fb950!important;background:#1c2128!important; }
    .schema-badge {
        display:inline-flex;align-items:center;gap:6px;
        background:#1c2128;border:1px solid #58a6ff;border-radius:20px;
        padding:3px 10px;font-size:11px;color:#58a6ff;font-weight:600;
        margin-left:10px;vertical-align:middle;
    }
    .schema-badge-duck  { border-color:#f0883e!important;color:#f0883e!important; }
    .schema-badge-guide { border-color:#58a6ff!important;color:#58a6ff!important; }
    .settings-btn div[data-testid="stButton"] button {
        background:transparent!important;border:1px solid #30363d!important;
        border-radius:8px!important;color:#8b949e!important;font-size:18px!important;
        padding:4px 10px!important;transition:all 0.2s!important;
    }
    .settings-btn div[data-testid="stButton"] button:hover {
        border-color:#58a6ff!important;color:#58a6ff!important;
        background:#1c2128!important;box-shadow:0 0 8px rgba(88,166,255,0.4)!important;
    }
    .mandatory-asterisk {
        display: inline-block;
        font-size: 13px;
        color: #58a6ff;
        font-weight: bold;
        margin-left: 4px;
        vertical-align: middle;
        line-height: 1;
    }
    .optional-badge {
        display:inline-block;background:#1c2128;border:1px solid #30363d;
        border-radius:3px;font-size:9px;color:#8b949e;
        padding:0 4px;margin-left:5px;vertical-align:middle;
    }
    .export-sel-btn div[data-testid="stButton"] > button {
        width:100%!important;padding:0!important;font-size:11px!important;
        display:flex!important;align-items:center!important;justify-content:center!important;
    }
    div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] { display:none!important; }
    div[data-testid="stForm"] { border:none!important;padding:0!important; }
</style>
""", unsafe_allow_html=True)


# ==============================
# SHEET HELPERS
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others


def classify_sheet(rows):
    text = " ".join(
        str(cell).lower()
        for row in rows[:20]
        for cell in row
        if cell
    )
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no"
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury",
        "date of incident", "injury date"
    ])
    has_financial = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense"
    ])
    if has_claim and (has_loss or has_financial):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return {}
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    merged_info = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        cell = ws.cell(mn_r, mn_c)
        val  = str(cell.value).strip() if cell.value else ""
        span_cols = mx_c - mn_c + 1
        span_rows = mx_r - mn_r + 1
        region_type = "TITLE" if mn_r <= 3 and span_cols >= 3 else \
                      "HEADER" if span_cols >= 2 and span_rows == 1 else "DATA"
        merged_info[f"R{mn_r}C{mn_c}"] = {
            "value": val, "type": region_type,
            "row_start": mn_r, "col_start": mn_c,
            "row_end": mx_r, "col_end": mx_c,
            "span_cols": span_cols, "span_rows": span_rows,
            "excel_row": mn_r, "excel_col": mn_c,
        }
    wb.close()
    return merged_info


def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    ext    = os.path.splitext(file_path)[1].lower()
    totals = {}
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows = raw_rows
        wb.close()

    if not rows:
        return totals

    header_row_index = None
    headers = []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}"
                       for j, h in enumerate(row)]
            break

    if header_row_index is None or not headers:
        return totals

    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index + 1:]):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total", "subtotal", "grand total", "sum", "totals"]):
            row_data = {}
            cell_row = cell_rows[header_row_index + 1 + r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers):
                    continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column \
                        if hasattr(cell_row[c_idx_0], 'column') else c_idx_0 + 1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0 + 1
                if clean_val:
                    row_data[headers[c_idx_0]] = {
                        "value": clean_val, "excel_row": r_idx, "excel_col": real_col,
                    }
            if row_data:
                totals_rows.append(row_data)

    if totals_rows:
        totals["rows"]     = totals_rows
        totals["excel_row"] = totals_rows[0].get(
            list(totals_rows[0].keys())[0], {}).get("excel_row", 9999)
        agg = {}
        for row_data in totals_rows:
            for field, info in row_data.items():
                try:
                    num = float(str(info["value"]).replace(",", "").replace("$", ""))
                    if field not in agg:
                        agg[field] = 0.0
                    agg[field] += num
                except:
                    pass
        totals["aggregated"] = {k: round(v, 2) for k, v in agg.items()}

    return totals


# ==============================
# CELL FORMATTING
# ==============================
_THEME_COLORS = {
    0: "FFFFFF", 1: "000000", 2: "EEECE1", 3: "1F497D",
    4: "4F81BD", 5: "C0504D", 6: "9BBB59", 7: "8064A2",
    8: "4BACC6", 9: "F79646",
}


def _resolve_color(color_obj, default="FFFFFF") -> str:
    if color_obj is None:
        return default
    t = color_obj.type
    if t == "rgb":
        rgb = color_obj.rgb or ""
        if len(rgb) == 8 and rgb not in ("00000000", "FF000000"):
            return rgb[2:]
        if len(rgb) == 6:
            return rgb
        return default
    if t == "theme":
        base = _THEME_COLORS.get(color_obj.theme, default)
        tint = color_obj.tint or 0.0
        if tint != 0.0:
            r, g, b = int(base[0:2], 16), int(base[2:4], 16), int(base[4:6], 16)
            if tint > 0:
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            else:
                r, g, b = int(r*(1+tint)), int(g*(1+tint)), int(b*(1+tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        return base
    if t == "indexed":
        indexed_map = {0:"000000",1:"FFFFFF",2:"FF0000",3:"00FF00",
                       4:"0000FF",5:"FFFF00",6:"FF00FF",7:"00FFFF",
                       64:"000000",65:"FFFFFF"}
        return indexed_map.get(color_obj.indexed, default)
    return default


def format_cell_value(value) -> str:
    if value is None: return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if (value.hour or value.minute) \
               else value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date): return value.strftime("%Y-%m-%d")
    if isinstance(value, bool): return str(value)
    if isinstance(value, int): return str(value)
    if isinstance(value, float):
        if value == int(value): return f"{int(value)}.0"
        formatted = f"{value:.10f}".rstrip('0')
        if '.' not in formatted: formatted += '.0'
        return formatted
    return normalize_str(str(value).strip())


def _apply_date_number_format(dt, nf: str) -> str:
    if not nf or nf.lower() in ("general", "@", ""):
        return dt.strftime("%m-%d-%Y")
    fmt = re.sub(r'\[.*?\]', '', nf)
    fmt = re.sub(r'["_*\\]', '', fmt)
    result = fmt
    result = re.sub(r'(?i)(?<=h)mm', '__MIN__', result)
    result = re.sub(r'(?i)mm(?=ss)', '__MIN__', result)

    def _tok(m):
        tok = m.group(0).lower()
        return {'yyyy':'%Y','yy':'%y','mmmm':'%B','mmm':'%b','mm':'%m',
                '__min__':'%M','m':'%m','dd':'%d','d':'%d',
                'hh':'%H','h':'%H','ss':'%S','s':'%S',
                'am/pm':'%p','a/p':'%p'}.get(tok, m.group(0))

    result = re.sub(
        r'(?i)yyyy|yy|mmmm|mmm|__min__|mm|dd|hh|ss|am/pm|a/p|d|h|s|m',
        _tok, result)
    try:
        return dt.strftime(result)
    except Exception:
        return dt.strftime("%m-%d-%Y")


def format_cell_value_with_fmt(cell) -> str:
    value = cell.value
    if value is None: return ""
    nf = (cell.number_format or "").strip()

    if isinstance(value, (datetime.datetime, datetime.date)):
        return _apply_date_number_format(value, nf)
    if isinstance(value, bool): return str(value)
    if isinstance(value, (int, float)):
        decimal_places = None
        if nf and nf.lower() not in ("general", "@", ""):
            clean_nf = re.sub(r'[$€£¥"_*\\]', '', nf)
            is_date_fmt = (
                any(x in clean_nf.lower() for x in ['yy','mm','dd','hh','ss'])
                and not any(ch in clean_nf for ch in ['0','#'])
            )
            if not is_date_fmt:
                if '.' in clean_nf:
                    after_dot = clean_nf.split('.')[1]
                    after_dot = re.sub(r'\[.*?\]', '', after_dot)
                    decimal_places = sum(1 for ch in after_dot if ch in '0#')
                else:
                    decimal_places = 0
        if decimal_places is not None:
            fval = float(value)
            return str(int(round(fval))) if decimal_places == 0 \
                   else f"{fval:.{decimal_places}f}"
        if isinstance(value, int): return str(value)
        fval      = float(value)
        remainder = fval - int(fval)
        if remainder == 0.0: return f"{fval:.2f}"
        formatted = f"{fval:.10f}".rstrip('0')
        if '.' not in formatted: formatted += '.00'
        elif len(formatted.split('.')[1]) < 2: formatted = f"{fval:.2f}"
        return formatted
    return normalize_str(str(value).strip())


# ==============================
# EXCEL RENDERER
# ==============================
def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd = ws.column_dimensions.get(letter)
    w  = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(20, int(w * 8 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str, scale: float = 1.0) -> tuple:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))

    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    img  = Image.new("RGB", (col_starts[-1], row_starts[-1]), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    merged_master: dict = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    drawn_merges: set = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            merge_info = merged_master.get((r, c))
            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue
                drawn_merges.add((mn_r, mn_c))
                x1, y1 = col_starts[mn_c-1], row_starts[mn_r-1]
                x2, y2 = col_starts[mx_c], row_starts[mx_r]
                cell = ws.cell(mn_r, mn_c)
            else:
                x1, y1 = col_starts[c-1], row_starts[r-1]
                x2, y2 = col_starts[c], row_starts[r]
                cell = ws.cell(r, c)

            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")
            draw.rectangle([x1, y1, x2-1, y2-1],
                           fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"
                bold     = bool(cell.font and cell.font.bold)
                text     = format_cell_value_with_fmt(cell) if cell.value is not None else ""
                cell_w   = x2 - x1
                ch_w     = 8 if bold else 7
                max_chars = max(1, (cell_w - 8) // ch_w)
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"
                draw.text((x1+4, y1+4), text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts, merged_master


def get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col, merged_master=None):
    c = max(1, min(target_col, len(col_starts)-1))
    r = max(1, min(target_row, len(row_starts)-1))
    if merged_master:
        info = merged_master.get((r, c))
        if info:
            mn_r, mn_c, mx_r, mx_c = info
            return (col_starts[mn_c-1], row_starts[mn_r-1],
                    col_starts[min(mx_c, len(col_starts)-1)],
                    row_starts[min(mx_r, len(row_starts)-1)])
    return (col_starts[c-1], row_starts[r-1],
            col_starts[min(c, len(col_starts)-1)],
            row_starts[min(r, len(row_starts)-1)])


def crop_context(img, x1, y1, x2, y2, pad_x=220, pad_y=160):
    iw, ih = img.size
    cx1 = max(0, x1-pad_x); cy1 = max(0, y1-pad_y)
    cx2 = min(iw, x2+pad_x); cy2 = min(ih, y2+pad_y)
    cropped = img.crop((cx1, cy1, cx2, cy2))
    return cropped, x1-cx1, y1-cy1, x2-cx1, y2-cy1


# ==============================
# EYE POPUP
# ==============================
@st.dialog("Cell View", width="large")
def show_eye_popup(field, info, excel_path, sheet_name):
    st.markdown(f"### 📍 {field}")
    value      = info.get("modified", info["value"])
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        st.markdown("**Extracted Value**")
        st.code(value if value else "(empty)")
    with col_b:
        col_letter = get_column_letter(target_col) if target_col else "?"
        st.markdown(f"""
            <div style="padding:10px 0;color:#8b949e;font-size:14px;">
                Cell: <span style="color:#58a6ff;font-weight:bold;">{col_letter}{target_row or '?'}</span>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                Row <span style="color:#c9d1d9;">{target_row or '?'}</span>
                · Col <span style="color:#c9d1d9;">{target_col or '?'}</span>
            </div>
        """, unsafe_allow_html=True)

    if not target_row or not target_col:
        st.warning("No cell location recorded for this field.")
        return

    ext = os.path.splitext(excel_path)[1].lower()
    if ext == ".csv":
        st.markdown("---")
        try:
            with open(excel_path, "r", encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            if not all_rows:
                st.warning("CSV file is empty.")
                return
            n_rows = len(all_rows)
            n_cols = max(len(r) for r in all_rows)
            r0 = max(0, target_row-6); r1 = min(n_rows, target_row+5)
            col_headers = "".join(
                f"<th style='background:#1c2128;color:#8b949e;font-size:10px;"
                f"padding:3px 8px;border:1px solid #30363d;'>"
                f"{get_column_letter(c+1)}</th>" for c in range(n_cols)
            )
            thead = (f"<thead><tr>"
                     f"<th style='background:#1c2128;color:#8b949e;font-size:10px;"
                     f"padding:3px 8px;border:1px solid #30363d;'>#</th>"
                     f"{col_headers}</tr></thead>")
            tbody = ""
            for r_idx in range(r0, r1):
                row_data      = all_rows[r_idx] if r_idx < len(all_rows) else []
                is_target_row = (r_idx + 1 == target_row)
                row_num_style = (
                    "background:#1c3a5e;color:#58a6ff;font-weight:bold;"
                    if is_target_row else "background:#161b22;color:#8b949e;"
                )
                cells = (f"<td style='{row_num_style}font-size:11px;padding:3px 8px;"
                         f"border:1px solid #30363d;text-align:center;'>{r_idx+1}</td>")
                for c_idx in range(n_cols):
                    cell_val = row_data[c_idx] if c_idx < len(row_data) else ""
                    is_tc    = is_target_row and (c_idx+1 == target_col)
                    if is_tc:
                        style = "background:#3a2800;border:2px solid #f0883e;color:#fff;font-weight:bold;"
                    elif is_target_row:
                        style = "background:#1a2030;border:1px solid #3a4a5e;color:#c9d1d9;"
                    else:
                        style = "background:#0d1117;border:1px solid #21262d;color:#8b949e;"
                    cells += (f"<td style='{style}font-size:12px;padding:4px 8px;"
                              f"max-width:160px;overflow:hidden;text-overflow:ellipsis;"
                              f"white-space:nowrap;'>{cell_val}</td>")
                tbody += f"<tr>{cells}</tr>"
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:6px;border:1px solid #30363d;'>"
                f"<table style='border-collapse:collapse;width:100%;font-family:monospace;'>"
                f"{thead}<tbody>{tbody}</tbody></table></div>",
                unsafe_allow_html=True
            )
        except Exception as e:
            st.error(f"CSV preview error: {e}")
        return

    st.markdown("---")
    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0)
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]

    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")
        x1, y1, x2, y2 = get_cell_pixel_bbox(
            col_starts, row_starts, target_row, target_col, merged_master)
        draw.rectangle([x1+1, y1+1, x2-1, y2-1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(255, 180, 0, 255), width=3)
        draw.rectangle([x1+3, y1+3, x2-3, y2-3], outline=(255, 255, 255, 160), width=1)
        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)
        col_letter = get_column_letter(target_col)
        st.image(cropped, use_container_width=True,
                 caption=f"Cell {col_letter}{target_row}  ·  Value: {value or '(empty)'}")
    except Exception as e:
        st.error(f"Rendering error: {e}")


# ==============================
# FORMAT CONVERTERS
# ==============================
def to_duck_creek_xml(mapped_records: list, sheet_meta: dict) -> str:
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    root = ET.Element("ClaimTransactionBatch")
    root.set("xmlns",      "http://www.duckcreek.com/claims/transaction/v6")
    root.set("xmlns:xsi",  "http://www.w3.org/2001/XMLSchema-instance")
    root.set("batchDate",  datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))
    root.set("source",     "TPA_Claims_Review_Portal")
    root.set("recordCount", str(len(mapped_records)))

    _DC_XML_MAP = {
        "Claim Id":"ClaimId","Transaction Id":"TransactionId",
        "Claimant Name":"ClaimantName","Loss Date":"LossDate",
        "Date Reported":"DateReported","Total Incurred":"TotalIncurred",
        "Total Paid":"TotalPaid","Reserve":"Reserve",
        "Indemnity Paid":"IndemnityPaid","Medical Paid":"MedicalPaid",
        "Expense Paid":"ExpensePaid","Policy Number":"PolicyNumber",
        "Policy Effective Date":"PolicyEffectiveDate",
        "Policy Expiry Date":"PolicyExpiryDate",
        "Claim Status":"ClaimStatus","Cause of Loss":"CauseOfLoss",
        "Description of Loss":"LossDescription","Insured Name":"InsuredName",
        "Carrier Name":"CarrierName","Line of Business":"LineOfBusiness",
        "Adjuster Id":"AdjusterId","Adjuster Name":"AdjusterName",
        "Office Code":"OfficeCode","Jurisdiction":"Jurisdiction",
        "State Code":"StateCode","Deductible Amount":"DeductibleAmount",
        "Subrogation Flag":"SubrogationFlag","Recovery Amount":"RecoveryAmount",
        "Litigation Flag":"LitigationFlag","Date Closed":"DateClosed",
        "Date Reopened":"DateReopened","Last Updated Date":"LastUpdatedDate",
        "Notes":"Notes",
    }

    for rec in mapped_records:
        txn = ET.SubElement(root, "ClaimTransaction")
        txn.set("transactionType", "UPDATE")
        txn.set("confidence", str(rec.get("_avg_confidence", "")))
        claim_el = ET.SubElement(txn, "Claim")
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            xml_tag = _DC_XML_MAP.get(schema_field, schema_field.replace(" ", ""))
            el      = ET.SubElement(claim_el, xml_tag)
            el.text = str(field_data.get("value", ""))
            if field_data.get("edited"):
                el.set("edited", "true")
                el.set("originalValue", str(field_data.get("original", "")))
            el.set("confidence", str(field_data.get("confidence", "")))

    xml_str = ET.tostring(root, encoding="unicode")
    pretty  = __import__('xml.dom.minidom', fromlist=['minidom']).minidom \
              .parseString(xml_str).toprettyxml(indent="  ")
    lines = pretty.split("\n")
    return "\n".join(lines[1:]) if lines[0].startswith("<?xml") else pretty


def to_duck_creek_json(mapped_records: list, sheet_meta: dict) -> dict:
    transactions = []
    for rec in mapped_records:
        claim_obj = {}
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            claim_obj[schema_field] = {
                "value":      field_data.get("value", ""),
                "confidence": field_data.get("confidence", 0),
                "edited":     field_data.get("edited", False),
            }
            if field_data.get("edited"):
                claim_obj[schema_field]["originalValue"] = field_data.get("original", "")
        transactions.append({
            "transactionType": "UPDATE",
            "avgConfidence":   rec.get("_avg_confidence", 0),
            "claim":           claim_obj,
        })
    return {
        "schema":       "DuckCreek.Claims.Transaction.v6",
        "exportDate":   datetime.datetime.now().isoformat(),
        "source":       "TPA_Claims_Review_Portal",
        "sheetName":    sheet_meta.get("sheet_name", ""),
        "recordCount":  len(transactions),
        "transactions": transactions,
    }


def to_guidewire_json(mapped_records: list, sheet_meta: dict) -> dict:
    _GW_FIELD_MAP = {
        "Claim Number":"claimNumber","Claimant Name":"claimantName",
        "Loss Date":"lossDate","Date Reported":"reportedDate",
        "Total Incurred":"totalIncurredAmount","Total Paid":"totalPaidAmount",
        "Reserve":"reserveAmount","Indemnity Paid":"indemnityPaidAmount",
        "Medical Paid":"medicalPaidAmount","Expense Paid":"expensePaidAmount",
        "Status":"status","Line of Business":"lineOfBusinessCode",
        "Policy Number":"policyNumber","Policy Period Start":"policyPeriodStart",
        "Policy Period End":"policyPeriodEnd","Carrier":"carrierName",
        "Insured Name":"insuredName","Description of Loss":"lossDescription",
        "Cause of Loss":"causeOfLoss","Litigation Flag":"litigationFlag",
        "Adjuster Name":"adjusterName","Adjuster Phone":"adjusterPhone",
        "Branch Code":"branchCode","Department Code":"departmentCode",
        "Coverage Type":"coverageType","Deductible":"deductibleAmount",
        "Subrogation Amount":"subrogationAmount","Recovery Amount":"recoveryAmount",
        "Open/Closed":"openClosedStatus","Reopen Date":"reopenDate",
        "Last Activity Date":"lastActivityDate","Notes":"notes",
    }

    claims = []
    for rec in mapped_records:
        claim_obj  = {"_type": "cc.Claim", "_confidence": rec.get("_avg_confidence", 0)}
        financials = {}
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            gw_key = _GW_FIELD_MAP.get(
                schema_field,
                schema_field[0].lower() + schema_field[1:].replace(" ", "")
            )
            val = field_data.get("value", "")
            if any(x in schema_field.lower() for x in
                   ["paid", "reserve", "incurred", "deductible", "recovery", "subrogation"]):
                financials[gw_key] = {"amount": val, "currency": "USD",
                                       "confidence": field_data.get("confidence", 0)}
                if field_data.get("edited"):
                    financials[gw_key]["originalValue"] = field_data.get("original", "")
            else:
                claim_obj[gw_key] = {"value": val, "confidence": field_data.get("confidence", 0)}
                if field_data.get("edited"):
                    claim_obj[gw_key]["originalValue"] = field_data.get("original", "")
        if financials:
            claim_obj["financials"] = financials
        claims.append(claim_obj)

    return {
        "schema":      "Guidewire.ClaimCenter.REST.v1",
        "exportDate":  datetime.datetime.now().isoformat(),
        "source":      "TPA_Claims_Review_Portal",
        "sheetName":   sheet_meta.get("sheet_name", ""),
        "recordCount": len(claims),
        "data":        {"claims": claims},
    }


def build_mapped_records_for_export(data: list, schema_name: str,
                                    selected_sheet: str) -> list:
    records     = []
    schema      = SCHEMAS[schema_name]
    custom_flds = st.session_state.get(f"custom_fields_{schema_name}", [])
    export_flds = list(schema["required_fields"]) + [
        f for f in custom_flds if f not in schema["required_fields"]
    ]
    _sheet_cache = st.session_state.get("sheet_cache", {})
    title_fields = _sheet_cache.get(selected_sheet, {}).get("title_fields", {})

    for i, row in enumerate(data):
        c_id   = detect_claim_id(row, i)
        mapped = map_claim_to_schema(row, schema_name, title_fields)
        rec    = {}
        confs  = []

        for sf in export_flds:
            if sf not in mapped:
                rec[sf] = {"value": "", "confidence": 0, "edited": False, "original": ""}
                confs.append(0)
                continue
            m       = mapped[sf]
            mk_key  = f"mod_{selected_sheet}_{c_id}_schema_{sf}"
            live_val = st.session_state.get(mk_key, None)
            orig    = m["info"].get("value", "")
            final   = live_val if live_val is not None else m["value"]
            rec[sf] = {
                "value": final, "original": orig, "edited": final != orig,
                "confidence": m["confidence"],
                "excel_row": m["info"].get("excel_row"),
                "excel_col": m["info"].get("excel_col"),
            }
            confs.append(m["confidence"])

        rec["_avg_confidence"] = round(sum(confs)/len(confs)) if confs else 0
        rec["_claim_id"]       = c_id
        records.append(rec)
    return records


def to_standard_json(export_data: dict, sheet_meta: dict,
                     totals: dict, merged_meta: dict) -> dict:
    titles_section = []
    sorted_merges = sorted(
        [(k, v) for k, v in merged_meta.items() if v.get("value")],
        key=lambda x: (x[1]["row_start"], x[1]["col_start"])
    )
    for _, m in sorted_merges:
        titles_section.append({
            "type": m["type"], "value": m["value"],
            "excel_row": m["excel_row"], "excel_col": m["excel_col"],
            "span_cols": m["span_cols"], "span_rows": m["span_rows"],
        })
    totals_section = {}
    if totals:
        totals_section = {
            "excel_row":  totals.get("excel_row"),
            "rows":       totals.get("rows", []),
            "aggregated": totals.get("aggregated", {}),
        }
    return {
        "exportDate":  datetime.datetime.now().isoformat(),
        "sheetMeta":   {"sheet_name": sheet_meta.get("sheet_name"),
                        "record_count": sheet_meta.get("record_count")},
        "titleRows":   titles_section,
        "records":     export_data,
        "totals":      totals_section,
        "recordCount": len(export_data),
    }


# ==============================
# UTILS
# ==============================
def get_val(claim: dict, keys: list, default: str = "") -> str:
    for pk in keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"] or default
    return default


def detect_claim_id(row, index=None):
    keys = ["claim id", "claim_id", "claimid", "claim number", "claim no",
            "claim #", "claim ref", "claim reference", "file number", "record id"]
    for k, v in row.items():
        name = str(k).lower().replace("_", " ").strip()
        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip():
                return str(val)
    if index is not None:
        return str(index + 1)
    return ""


def clean_duplicate_fields(record: dict) -> dict:
    seen, out = set(), {}
    for k, v in record.items():
        if k.strip() not in seen:
            seen.add(k.strip())
            out[k.strip()] = v
    return out


# ── Audit / Feature Store helpers ────────────────────────────────────────
AUDIT_LOG_PATH = os.path.join(FEATURE_STORE_PATH, "_audit_log.json")


def _load_audit_log() -> list:
    if os.path.exists(AUDIT_LOG_PATH):
        try:
            with open(AUDIT_LOG_PATH) as f:
                return json.load(f)
        except Exception:
            pass
    return []


def _save_audit_log(log: list):
    with open(AUDIT_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2, ensure_ascii=False)


def compute_file_sha256(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def compute_sheet_sha256(data: list) -> str:
    """Hash the extracted row data for a single sheet."""
    payload = json.dumps(data, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode()).hexdigest()


def check_file_duplicate(file_hash: str) -> dict | None:
    """Return the existing audit entry if this file hash was seen before."""
    for entry in _load_audit_log():
        if entry.get("file_hash") == file_hash:
            return entry
    return None


def record_audit_entry(filename: str, file_hash: str, sheet_hashes: dict,
                        num_sheets: int) -> dict:
    log  = _load_audit_log()
    ts   = datetime.datetime.now().isoformat(timespec="seconds")
    entry = {
        "filename":     filename,
        "file_hash":    file_hash,
        "sheet_hashes": sheet_hashes,
        "num_sheets":   num_sheets,
        "uploaded_at":  ts,
    }
    log.append(entry)
    _save_audit_log(log)
    return entry


def save_feature_store(sheet_name: str, data: dict,
                       filename: str = "", file_hash: str = "") -> str:
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = sheet_name.replace(" ", "_").replace("/", "-")
    path = os.path.join(FEATURE_STORE_PATH, f"{safe}_{ts}.json")

    def _sanitize(obj):
        if isinstance(obj, dict): return {k: _sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list): return [_sanitize(i) for i in obj]
        if isinstance(obj, str):  return normalize_str(obj)
        return obj

    envelope = {
        "_meta": {
            "filename":    filename,
            "sheet_name":  sheet_name,
            "file_hash":   file_hash,
            "exported_at": datetime.datetime.now().isoformat(timespec="seconds"),
        },
        "data": _sanitize(data),
    }
    with open(path, "w") as f:
        json.dump(envelope, f, indent=2, ensure_ascii=False)
    return path


def list_feature_store_entries() -> list:
    """Return all saved JSON entries sorted newest-first as list of dicts."""
    entries = []
    for fname in os.listdir(FEATURE_STORE_PATH):
        if not fname.endswith(".json") or fname.startswith("_"):
            continue
        fpath = os.path.join(FEATURE_STORE_PATH, fname)
        try:
            with open(fpath) as f:
                obj = json.load(f)
            meta = obj.get("_meta", {})
            entries.append({
                "file":       fname,
                "path":       fpath,
                "filename":   meta.get("filename", "—"),
                "sheet_name": meta.get("sheet_name", fname.rsplit("_", 2)[0]),
                "file_hash":  meta.get("file_hash", "—"),
                "exported_at":meta.get("exported_at", "—"),
                "data":       obj.get("data", obj),
            })
        except Exception:
            pass
    entries.sort(key=lambda x: x["exported_at"], reverse=True)
    return entries


def extract_from_excel(file_path, sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN"
        return parse_rows(classify_sheet(rows), rows)
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN"
        return parse_rows_with_cells(classify_sheet(raw_rows), raw_rows, cell_rows)


def parse_rows_with_cells(sheet_type, rows, cell_rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type
        headers = [str(h).strip() if h is not None else f"Column_{i}"
                   for i, h in enumerate(rows[header_row_index])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(
            zip(rows[header_row_index+1:], cell_rows[header_row_index+1:])
        ):
            r_idx = header_row_index + 2 + r_idx_rel
            if not any(raw_row): continue
            row_data = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers): continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0+1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if ("claim" in row_text or "employee name" in row_text or "driver name" in row_text) and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
            or "injury" in row_text or "incident" in row_text
        ):
            header_row_index = i
            break
    if header_row_index is None:
        return [], sheet_type

    headers = [str(h).strip() if h is not None else f"Column_{i}"
               for i, h in enumerate(rows[header_row_index])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(
        zip(rows[header_row_index+1:], cell_rows[header_row_index+1:])
    ):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row): continue
        if any(str(c).lower().strip() in ["totals","total","grand total"]
               for c in raw_row if c):
            break
        row_data = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers): continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0+1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def parse_rows(sheet_type, rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type
        headers = [str(h).strip() if h is not None else f"Column_{i}"
                   for i, h in enumerate(rows[header_row_index])]
        extracted = []
        for r_idx, row in enumerate(rows[header_row_index+1:], start=header_row_index+2):
            if not any(row): continue
            row_data = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx-1 >= len(headers): continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx-1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if ("claim" in row_text or "employee name" in row_text or "driver name" in row_text) and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
            or "injury" in row_text or "incident" in row_text
        ):
            header_row_index = i
            break
    if header_row_index is None:
        return [], sheet_type

    headers = [str(h).strip() if h is not None else f"Column_{i}"
               for i, h in enumerate(rows[header_row_index])]
    extracted = []
    for r_idx, row in enumerate(rows[header_row_index+1:], start=header_row_index+2):
        if not any(row): continue
        if any(str(cell).lower().strip() in ["totals","total","grand total"]
               for cell in row if cell):
            break
        row_data = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx-1 >= len(headers): continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx-1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


# ==============================
# SESSION STATE DEFAULTS  ← FIX: all keys initialised before any widget reads them
# ==============================
for _k, _v in [
    ("conf_threshold",      80),
    ("use_conf_threshold",  True),
    ("active_schema",       None),
    ("schema_popup_target", None),
    ("schema_popup_tab",    "required"),
    ("settings_saved",      False),
    # ── keys that were previously only set inside `if uploaded:` ──────────
    ("sheet_cache",         {}),
    ("sheet_names",         []),
    ("selected_idx",        0),
    ("sheet_hashes",        {}),
    ("last_uploaded",       None),
    ("is_duplicate",        False),
    ("file_hash",           ""),
    ("file_bytes",          0),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ==============================
# TOP BAR  —  Logo always shown via embedded base64
# ==============================
col_title, col_gear, col_sheet_dropdown = st.columns([3.5, 0.5, 1])

with col_title:
    active_schema = st.session_state.get("active_schema", None)
    badge_html = ""
    if active_schema and active_schema in SCHEMAS:
        sc = SCHEMAS[active_schema]
        badge_html = (
            f'<span class="schema-badge schema-badge-{sc["css_cls"]}" '
            f'style="border-color:{sc["color"]};color:{sc["color"]};background:#1c2128;">'
            f'{sc["icon"]} {active_schema} · {sc["version"]}'
            f'</span>'
        )

    logo_html = _logo_img_tag(height=38)

    st.markdown(
        f'<div class="main-title" style="display:flex;align-items:center;">'
        f'{logo_html}📄 Loss Run Parser{badge_html}</div>',
        unsafe_allow_html=True
    )

with col_gear:
    st.markdown("<div style='margin-top:14px;' class='settings-btn'>", unsafe_allow_html=True)
    if st.button("⚙", key="open_settings", help="Open Settings", use_container_width=True):
        show_settings_dialog()
    st.markdown("</div>", unsafe_allow_html=True)

if st.session_state.get("schema_popup_target"):
    _target = st.session_state["schema_popup_target"]
    st.session_state["schema_popup_target"] = None
    show_schema_fields_dialog(_target)

# ==============================
# FILE UPLOAD
# ==============================
uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        raw_bytes  = uploaded.read()
        file_hash  = compute_file_sha256(raw_bytes)
        with open(excel_path, "wb") as f:
            f.write(raw_bytes)
        st.session_state.last_uploaded = uploaded.name
        st.session_state.file_hash     = file_hash
        st.session_state.file_bytes    = len(raw_bytes)
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None
        st.session_state.sheet_hashes  = {}
        for key in list(st.session_state.keys()):
            if key.startswith("_rendered_"):
                del st.session_state[key]
        dup = check_file_duplicate(file_hash)
        if dup:
            st.warning(
                f"⚠️ **Duplicate detected** — this file was previously uploaded on "
                f"`{dup['uploaded_at']}` as `{dup['filename']}`. "
                f"Processing will continue but no new audit entry will be created."
            )
            st.session_state.is_duplicate = True
        else:
            st.session_state.is_duplicate = False

    # ── File-level high-level summary ─────────────────────────────────────
    _file_hash    = st.session_state.get("file_hash", "")
    _file_bytes   = st.session_state.get("file_bytes", 0)
    _sheet_names  = st.session_state.get("sheet_names", [])
    _is_dup       = st.session_state.get("is_duplicate", False)
    _dup_badge    = ("<span style='background:#d29922;color:#0d1117;font-size:9px;"
                     "border-radius:4px;padding:1px 6px;margin-left:8px;font-weight:bold;'>"
                     "DUPLICATE</span>" if _is_dup else
                     "<span style='background:#3fb950;color:#0d1117;font-size:9px;"
                     "border-radius:4px;padding:1px 6px;margin-left:8px;font-weight:bold;'>"
                     "UNIQUE</span>")
    st.markdown(f"""
    <div style="background:#161b22;border:1px solid #30363d;border-left:4px solid #58a6ff;
                border-radius:6px;padding:12px 16px;margin-bottom:12px;">
        <div style="font-family:'Segoe UI',Arial,sans-serif;font-size:10px;color:#8b949e;
                    text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;">
            📁 File Summary
        </div>
        <div style="display:flex;flex-wrap:wrap;gap:20px;align-items:center;">
            <div>
                <span style="font-family:'Segoe UI',Arial,sans-serif;font-size:14px;
                             font-weight:600;color:#e6edf3;">{uploaded.name}</span>
                {_dup_badge}
            </div>
            <div style="font-size:12px;color:#8b949e;">
                📄 <b style="color:#c9d1d9;">{len(_sheet_names)}</b> sheet(s)
                &nbsp;|&nbsp;
                💾 <b style="color:#c9d1d9;">{_file_bytes/1024:.1f} KB</b>
                &nbsp;|&nbsp;
                🔑 SHA-256: <code style="font-size:10px;color:#58a6ff;">
                    {_file_hash[:16]}…{_file_hash[-8:]}</code>
            </div>
        </div>
        <div style="margin-top:8px;display:flex;flex-wrap:wrap;gap:6px;">
            {"".join(
                f'<span style="background:#21262d;border:1px solid #30363d;border-radius:4px;'
                f'padding:2px 8px;font-size:11px;color:#c9d1d9;">{s}</span>'
                for s in _sheet_names
            )}
        </div>
    </div>
    """, unsafe_allow_html=True)

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox(
            "Sheet", st.session_state.sheet_names,
            index=0, label_visibility="collapsed"
        )

    st.markdown("<hr style='border-color:#30363d;margin-top:-10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Reading '{selected_sheet}'..."):
            data, sheet_type = extract_from_excel(excel_path, selected_sheet)
            merged_meta      = extract_merged_cell_metadata(excel_path, selected_sheet)
            totals_data      = extract_totals_row(excel_path, selected_sheet)
            if not data:
                st.warning(f"No data found in sheet '{selected_sheet}'.")
                st.stop()
            for row in data:
                for fld, inf in row.items():
                    for key in ("value", "modified"):
                        if key in inf and isinstance(inf[key], str):
                            inf[key] = normalize_str(inf[key])
            _title_flds = extract_title_fields(merged_meta)

            _sheet_hash = compute_sheet_sha256(
                [{k: v.get("value","") for k,v in r.items()} for r in data]
            )
            st.session_state.sheet_hashes = {
                **st.session_state.get("sheet_hashes", {}),
                selected_sheet: _sheet_hash,
            }
            if not st.session_state.get("is_duplicate", False):
                record_audit_entry(
                    filename     = uploaded.name,
                    file_hash    = st.session_state.get("file_hash", ""),
                    sheet_hashes = st.session_state.sheet_hashes,
                    num_sheets   = len(st.session_state.sheet_names),
                )

            st.session_state.sheet_cache[selected_sheet] = {
                "data":         data,
                "merged_meta":  merged_meta,
                "totals":       totals_data,
                "title_fields": _title_flds,
                "sheet_type":   sheet_type,
                "sheet_hash":   _sheet_hash,
            }
            st.session_state.selected_idx = 0
            st.session_state.focus_field  = None

    active       = st.session_state.sheet_cache[selected_sheet]
    data         = active["data"]
    merged_meta  = active.get("merged_meta", {})
    totals_data  = active.get("totals", {})
    title_fields = active.get("title_fields", {})
    _s_type      = active.get("sheet_type", "UNKNOWN")
    _s_hash      = active.get("sheet_hash", "")

    # ── Sheet-level summary banner ─────────────────────────────────────────
    _num_rows = len(data)
    _num_cols = max((len(r) for r in data), default=0) if data else 0
    _totals_flag = "Yes" if totals_data else "No"
    _merged_count = len([v for v in merged_meta.values() if v.get("value")])
    _lob_val  = title_fields.get("Line of Business", {}).get("value", "—")
    _pol_val  = title_fields.get("Policy Number",    {}).get("value", "—")
    st.markdown(f"""
    <div style="background:#161b22;border:1px solid #30363d;border-left:4px solid #3fb950;
                border-radius:6px;padding:10px 16px;margin-bottom:10px;">
        <div style="font-family:'Segoe UI',Arial,sans-serif;font-size:10px;color:#8b949e;
                    text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;">
            📊 Sheet Summary — <b style="color:#c9d1d9;">{selected_sheet}</b>
        </div>
        <div style="display:flex;flex-wrap:wrap;gap:18px;font-size:12px;color:#8b949e;">
            <span>🗂 Type: <b style="color:#c9d1d9;">{_s_type}</b></span>
            <span>📋 Rows: <b style="color:#c9d1d9;">{_num_rows}</b></span>
            <span>📐 Columns: <b style="color:#c9d1d9;">{_num_cols}</b></span>
            <span>∑ Totals row: <b style="color:#c9d1d9;">{_totals_flag}</b></span>
            <span>🔗 Merged regions: <b style="color:#c9d1d9;">{_merged_count}</b></span>
            <span>🏷 LOB: <b style="color:#c9d1d9;">{_lob_val}</b></span>
            <span>📄 Policy: <b style="color:#c9d1d9;">{_pol_val}</b></span>
        </div>
        {"" if not _s_hash else
         f'<div style="margin-top:6px;font-size:10px;color:#8b949e;">'
         f'🔑 Sheet SHA-256: <code style="color:#58a6ff;">'
         f'{_s_hash[:16]}…{_s_hash[-8:]}</code></div>'}
    </div>
    """, unsafe_allow_html=True)

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim = data[st.session_state.selected_idx]

    col_nav, col_main, col_fmt = st.columns([1.2, 3.2, 1.4], gap="large")

    # ── LEFT NAV ────────────────────────────────────────────────────────
    with col_nav:
        with st.container(height=700, border=False):
            st.markdown(
                "<p style='color:#8b949e;font-weight:bold;font-size:12px;"
                "text-transform:uppercase;'>CLAIM RECORDS</p>",
                unsafe_allow_html=True
            )
            for i, row_data in enumerate(data):
                is_sel   = "selected-card" if st.session_state.selected_idx == i else ""
                c_id     = detect_claim_id(row_data, i)
                c_name   = get_val(row_data, ["Insured Name","Name","Company",
                                              "Claimant","TPA_NAME"], "Unknown Entity")
                raw_st   = get_val(row_data, ["Status","CLAIM_STATUS"], "")
                c_status = raw_st or ("Yet to Review" if i == 0 else
                                      "In Progress"   if i == 1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or \
                           c_status.lower() == "open" else "status-text"
                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold;color:white;font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e;font-size:13px;margin-top:2px;">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Select", key=f"sel_{selected_sheet}_{i}",
                             use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field  = None
                    st.rerun()

    # ── MIDDLE PANEL ────────────────────────────────────────────────────
    with col_main:
        sorted_titles = sorted(
            [(k, v) for k, v in merged_meta.items() if v.get("value")],
            key=lambda x: (x[1]["row_start"], x[1]["col_start"])
        )
        if sorted_titles:
            main_title_val = ""
            sub_title_val  = ""
            for _, m in sorted_titles:
                if m["type"] == "TITLE":
                    if not main_title_val: main_title_val = m["value"]
                    elif not sub_title_val: sub_title_val = m["value"]
            if main_title_val or sub_title_val:
                st.markdown(f"""
                <div class="sheet-title-banner">
                    <div class="sheet-title-label">Sheet Title</div>
                    <div class="sheet-title-value">{main_title_val}</div>
                    {"" if not sub_title_val else
                     f'<div class="sheet-subtitle-value">{sub_title_val}</div>'}
                </div>
                """, unsafe_allow_html=True)

        head_left, head_right = st.columns([3, 1])
        curr_claim_id = detect_claim_id(curr_claim)

        with head_left:
            st.markdown(
                "<p style='color:#8b949e;font-weight:bold;font-size:12px;"
                "text-transform:uppercase;'>Review Details</p>",
                unsafe_allow_html=True
            )
            h_name   = get_val(curr_claim, ["Insured Name","Name","Claimant","TPA_NAME"],
                               "Unknown Entity")
            h_date   = get_val(curr_claim, ["Loss Date","Date","LOSS_DATE"], "N/A")
            h_status = get_val(curr_claim, ["Status","CLAIM_STATUS"], "Submitted")
            h_total  = get_val(curr_claim,
                               ["Total Incurred","Incurred","Total","Amount","TOTAL_INCURRED"],
                               "$0")
            st.markdown(f"""
                <div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} — {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown(
                "<p style='color:#8b949e;font-weight:bold;font-size:12px;"
                "text-transform:uppercase;text-align:right;'>Export Selection</p>",
                unsafe_allow_html=True
            )
            b1, b2 = st.columns([1, 1])
            with b1:
                if st.button("✔ All", key=f"all_{selected_sheet}_{curr_claim_id}",
                             use_container_width=True):
                    for fld in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{fld}"] = True
                    st.rerun()
            with b2:
                if st.button("✘ None", key=f"none_{selected_sheet}_{curr_claim_id}",
                             use_container_width=True):
                    for fld in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{fld}"] = False
                    st.rerun()

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        _active_schema = st.session_state.get("active_schema", None)
        _use_conf      = st.session_state.get("use_conf_threshold", True)
        _conf_thresh   = st.session_state.get("conf_threshold", 80) if _use_conf else 0
        _show_conf     = _use_conf

        if _active_schema and _active_schema in SCHEMAS:
            # ── SCHEMA MODE ──────────────────────────────────────────
            _schema_def  = SCHEMAS[_active_schema]
            _mapped      = map_claim_to_schema(curr_claim, _active_schema, title_fields)
            _custom_flds = st.session_state.get(f"custom_fields_{_active_schema}", [])

            _display_fields = list(_schema_def["required_fields"]) + [
                f for f in _custom_flds if f not in _schema_def["required_fields"]
            ]

            _low_conf = [
                sf for sf in _display_fields
                if sf in _mapped and _mapped[sf]["confidence"] < _conf_thresh and _use_conf
            ]
            _missing  = [sf for sf in _schema_def["required_fields"] if sf not in _mapped]

            if _missing:
                st.markdown(
                    f"<div style=\"background:#2d1515;border:1px solid #f85149;"
                    f"border-radius:6px;padding:8px 12px;margin-bottom:8px;"
                    f"font-size:12px;color:#f85149;\">"
                    f"⚠ {len(_missing)} mandatory field(s) could not be mapped: "
                    f"{', '.join(_missing)}</div>",
                    unsafe_allow_html=True
                )
            if _low_conf:
                st.markdown(
                    f"<div style=\"background:#2d2208;border:1px solid #d29922;"
                    f"border-radius:6px;padding:8px 12px;margin-bottom:8px;"
                    f"font-size:12px;color:#d29922;\">"
                    f"⚡ {len(_low_conf)} field(s) below confidence threshold "
                    f"({_conf_thresh}%): {', '.join(_low_conf)}</div>",
                    unsafe_allow_html=True
                )

            if _show_conf:
                hc = st.columns([1.8, 1.7, 1.8, 1.8, 0.55, 0.55, 0.45])
                for col_i, lbl in enumerate(["**SCHEMA FIELD**", "**CONFIDENCE**",
                                              "**EXTRACTED VALUE**", "**MODIFIED VALUE**"]):
                    with hc[col_i]: st.markdown(lbl)
            else:
                hc = st.columns([1.8, 1.8, 1.8, 0.55, 0.55, 0.45])
                for col_i, lbl in enumerate(["**SCHEMA FIELD**",
                                              "**EXTRACTED VALUE**", "**MODIFIED VALUE**"]):
                    with hc[col_i]: st.markdown(lbl)

            for schema_field in _display_fields:
                if schema_field not in _mapped:
                    is_req = schema_field in _schema_def["required_fields"]
                    if is_req:
                        _nf_bg, _nf_border, _nf_color = "#1a0e0e", "#f85149", "#f85149"
                        _nf_badge_bg, _nf_badge_color  = "#f85149", "white"
                        _nf_label = "MANDATORY · NOT FOUND"
                    else:
                        _nf_bg, _nf_border, _nf_color = "#161b22", "#30363d", "#8b949e"
                        _nf_badge_bg, _nf_badge_color  = "#21262d", "#8b949e"
                        _nf_label = "OPTIONAL · NOT IN SHEET"
                    st.markdown(
                        f"<div style=\"display:flex;align-items:center;gap:8px;"
                        f"background:{_nf_bg};border:1px solid {_nf_border};"
                        f"border-radius:6px;padding:6px 10px;margin:2px 0;\">"
                        f"<span style=\"color:{_nf_color};font-size:12px;font-weight:bold;"
                        f"text-transform:uppercase;\">{schema_field}</span>"
                        f"<span style=\"background:{_nf_badge_bg};color:{_nf_badge_color};"
                        f"font-size:9px;border-radius:4px;padding:1px 5px;"
                        f"border:1px solid {_nf_border};\">"
                        f"{_nf_label}</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                    continue

                m        = _mapped[schema_field]
                conf     = m["confidence"]
                excel_f  = m["excel_field"]
                info     = m["info"]
                is_req   = m["is_required"]
                is_title_sourced = m.get("from_title", False)

                if not _use_conf:
                    conf_col, row_border, row_bg = "#8b949e", "#30363d", "#161b22"
                elif conf < _conf_thresh:
                    conf_col, row_border, row_bg = "#f85149", "#f85149", "#1f0d0d"
                elif conf < 75:
                    conf_col, row_border, row_bg = "#f0883e", "#f0883e", "#1f1508"
                elif conf < 88:
                    conf_col, row_border, row_bg = "#d29922", "#30363d", "#161b22"
                else:
                    conf_col, row_border, row_bg = "#3fb950", "#30363d", "#161b22"

                ek = f"edit_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"

                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state:
                    st.session_state[mk] = info.get("modified", info["value"])

                st.markdown(
                    f"<div style=\"border-left:3px solid {row_border};background:{row_bg};"
                    f"border-radius:0 4px 4px 0;padding:2px 0 2px 4px;margin:1px 0;\"></div>",
                    unsafe_allow_html=True
                )

                _cur_val = st.session_state.get(mk, info.get("modified", info["value"]))
                _edited  = _cur_val != info["value"]
                _dot     = "<span style=\"color:#d29922;font-size:8px;\">●</span> " \
                           if _edited else ""

                if is_req:
                    _badge_html = (
                        "<span class='mandatory-asterisk' "
                        "title='Mandatory field — required by schema'>*</span>"
                    )
                else:
                    _badge_html = "<span class='optional-badge'>OPT</span>"

                _field_label_html = (
                    f"<div style=\"min-height:40px;display:flex;flex-direction:column;"
                    f"justify-content:center;color:#c9d1d9;font-size:11px;font-weight:bold;"
                    f"text-transform:uppercase;\">"
                    f"<div style='display:flex;align-items:center;gap:4px;'>"
                    f"{_dot}{schema_field}{_badge_html}</div>"
                    f"</div>"
                )

                _conf_html = (
                    f"<div style=\"min-height:40px;display:flex;flex-direction:column;"
                    f"justify-content:center;gap:3px;\">"
                    f"<span style=\"background:{conf_col}22;border:1px solid {conf_col};"
                    f"border-radius:10px;padding:2px 8px;font-size:12px;"
                    f"color:{conf_col};font-weight:bold;\">{conf}%</span>"
                    f"<div style=\"background:#21262d;border-radius:3px;height:4px;width:80%;\">"
                    f"<div style=\"background:{conf_col};height:4px;border-radius:3px;"
                    f"width:{conf}%;\"></div></div>"
                    f"<span style=\"font-size:9px;color:#8b949e;\">"
                    f"H:{m['header_score']}% V:{m['value_score']}%</span>"
                    f"</div>"
                )

                def _render_edit_col(ek, mk, info, is_title_sourced, selected_sheet,
                                     curr_claim_id, schema_field, active):
                    if st.session_state[ek]:
                        with st.form(
                            key=f"form_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                            border=False
                        ):
                            nv = st.text_input(
                                "m",
                                value=st.session_state.get(
                                    mk, info.get("modified", info["value"])),
                                label_visibility="collapsed"
                            )
                            submitted = st.form_submit_button("", use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv
                                if (not is_title_sourced and
                                    excel_f in active["data"][st.session_state.selected_idx]):
                                    active["data"][
                                        st.session_state.selected_idx][excel_f]["modified"] = nv
                                st.session_state[ek] = False
                                st.rerun()
                    else:
                        st.text_input("m", key=mk, label_visibility="collapsed", disabled=True)
                    if (not is_title_sourced and
                        excel_f in active["data"][st.session_state.selected_idx]):
                        active["data"][
                            st.session_state.selected_idx][excel_f]["modified"] = \
                            st.session_state.get(mk, info.get("modified", info["value"]))

                def _render_edit_btn(ek, selected_sheet, curr_claim_id, schema_field):
                    if not st.session_state[ek]:
                        if st.button("✏",
                                     key=f"ed_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                                     use_container_width=True, help="Edit field"):
                            st.session_state[ek] = True
                            st.rerun()
                    else:
                        st.markdown(
                            "<div style=\"height:38px;display:flex;align-items:center;"
                            "justify-content:center;color:#3fb950;font-size:11px;"
                            "border:1px solid #30363d;border-radius:6px;\">↵</div>",
                            unsafe_allow_html=True
                        )

                if _show_conf:
                    cl, cc, co, cm, ce, cb, cx = st.columns(
                        [1.8, 1.5, 1.8, 1.8, 0.55, 0.55, 0.45], gap="small")
                    with cl: st.markdown(_field_label_html, unsafe_allow_html=True)
                    with cc: st.markdown(_conf_html, unsafe_allow_html=True)
                    with co:
                        st.text_input(
                            "o", value=info["value"],
                            key=f"orig_{selected_sheet}_{curr_claim_id}_schema_{schema_field}",
                            label_visibility="collapsed", disabled=True)
                    with cm:
                        _render_edit_col(ek, mk, info, is_title_sourced, selected_sheet,
                                         curr_claim_id, schema_field, active)
                    with ce:
                        if st.button("👁",
                                     key=f"eye_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                                     use_container_width=True):
                            show_eye_popup(schema_field, info, excel_path, selected_sheet)
                    with cb:
                        _render_edit_btn(ek, selected_sheet, curr_claim_id, schema_field)
                    with cx:
                        st.markdown("<div style=\"height:8px;\"></div>", unsafe_allow_html=True)
                        st.checkbox("", key=xk, label_visibility="collapsed")

                else:
                    cl, co, cm, ce, cb, cx = st.columns(
                        [1.8, 1.8, 1.8, 0.55, 0.55, 0.45], gap="small")
                    with cl: st.markdown(_field_label_html, unsafe_allow_html=True)
                    with co:
                        st.text_input(
                            "o", value=info["value"],
                            key=f"orig_{selected_sheet}_{curr_claim_id}_schema_{schema_field}",
                            label_visibility="collapsed", disabled=True)
                    with cm:
                        _render_edit_col(ek, mk, info, is_title_sourced, selected_sheet,
                                         curr_claim_id, schema_field, active)
                    with ce:
                        if st.button("👁",
                                     key=f"eye_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                                     use_container_width=True):
                            show_eye_popup(schema_field, info, excel_path, selected_sheet)
                    with cb:
                        _render_edit_btn(ek, selected_sheet, curr_claim_id, schema_field)
                    with cx:
                        st.markdown("<div style=\"height:8px;\"></div>", unsafe_allow_html=True)
                        st.checkbox("", key=xk, label_visibility="collapsed")

        else:
            # ── PLAIN MODE ──────────────────────────────────────────
            hc = st.columns([2, 2.6, 2.6, 0.6, 0.6, 0.5])
            with hc[0]: st.markdown("**FIELD**")
            with hc[1]: st.markdown("**EXTRACTED VALUE**")
            with hc[2]: st.markdown("**MODIFIED VALUE**")

            for field, info in curr_claim.items():
                ek = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_{field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_{field}"

                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state:
                    st.session_state[mk] = info.get("modified", info["value"])

                cl, co, cm, ce, cb, cx = st.columns([2, 2.6, 2.6, 0.9, 0.9, 0.5], gap="small")

                with cl:
                    _current_val = st.session_state.get(mk, info.get("modified", info["value"]))
                    _dot         = "<span style=\"color:#d29922;margin-left:4px;font-size:8px;\">" \
                                   "●</span>" if _current_val != info["value"] else ""
                    st.markdown(
                        f"<div style=\"height:40px;display:flex;align-items:center;"
                        f"color:#c9d1d9;font-size:12px;font-weight:bold;text-transform:uppercase;\">"
                        f"{field}{_dot}</div>",
                        unsafe_allow_html=True
                    )

                with co:
                    st.text_input("o", value=info["value"],
                                  key=f"orig_{selected_sheet}_{curr_claim_id}_{field}",
                                  label_visibility="collapsed", disabled=True)

                with cm:
                    if st.session_state[ek]:
                        with st.form(
                            key=f"form_{selected_sheet}_{curr_claim_id}_{field}", border=False
                        ):
                            nv = st.text_input(
                                "m",
                                value=st.session_state.get(mk, info.get("modified", info["value"])),
                                label_visibility="collapsed"
                            )
                            submitted = st.form_submit_button("", use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv
                                active["data"][st.session_state.selected_idx][field]["modified"] = nv
                                st.session_state[ek] = False
                                st.rerun()
                    else:
                        st.text_input("m", key=mk, label_visibility="collapsed", disabled=True)
                    active["data"][st.session_state.selected_idx][field]["modified"] = \
                        st.session_state.get(mk, info.get("modified", info["value"]))

                with ce:
                    if st.button("👁", key=f"eye_{selected_sheet}_{curr_claim_id}_{field}",
                                 use_container_width=True):
                        show_eye_popup(field, info, excel_path, selected_sheet)

                with cb:
                    if not st.session_state[ek]:
                        if st.button("✏", key=f"ed_{selected_sheet}_{curr_claim_id}_{field}",
                                     use_container_width=True, help="Edit field"):
                            st.session_state[ek] = True
                            st.rerun()
                    else:
                        st.markdown(
                            "<div style=\"height:38px;display:flex;align-items:center;"
                            "justify-content:center;color:#3fb950;font-size:11px;"
                            "border:1px solid #30363d;border-radius:6px;\">↵</div>",
                            unsafe_allow_html=True
                        )

                with cx:
                    st.markdown("<div style=\"height:8px;\"></div>", unsafe_allow_html=True)
                    st.checkbox("", key=xk, label_visibility="collapsed")

        # Totals
        if totals_data:
            st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>",
                        unsafe_allow_html=True)
            st.markdown("**📊 Sheet Totals**")
            agg = totals_data.get("aggregated", {})
            if agg:
                t_cols = st.columns(min(4, len(agg)))
                for idx, (k, v) in enumerate(agg.items()):
                    with t_cols[idx % len(t_cols)]:
                        st.markdown(f"""
                        <div style="background:#161b22;border:1px solid #30363d;
                                    border-radius:6px;padding:8px 12px;margin-bottom:6px;">
                            <div style="font-size:11px;color:#8b949e;text-transform:uppercase;">{k}</div>
                            <div style="font-size:16px;font-weight:bold;color:#3fb950;">{v:,.2f}</div>
                        </div>""", unsafe_allow_html=True)

    # ── RIGHT PANEL — Export ─────────────────────────────────────────
    with col_fmt:
        st.markdown(
            "<p style='color:#8b949e;font-weight:bold;font-size:12px;"
            "text-transform:uppercase;'>Export Format</p>",
            unsafe_allow_html=True
        )

        _active = st.session_state.get("active_schema", None)
        if _active and _active in SCHEMAS:
            _sc       = SCHEMAS[_active]
            _cf_count = len(st.session_state.get(f"custom_fields_{_active}", []))
            st.markdown(
                f"<div style='background:#1c2128;border:1px solid {_sc['color']};"
                f"border-radius:8px;padding:10px 12px;margin-bottom:8px;'>"
                f"<div style='font-size:13px;font-weight:bold;color:{_sc['color']};'>"
                f"{_sc['icon']} {_active} Active</div>"
                f"<div style='font-size:11px;color:#8b949e;margin-top:3px;'>{_sc['version']}</div>"
                f"<div style='font-size:11px;color:#8b949e;margin-top:2px;'>"
                f"Mandatory: {len(_sc['required_fields'])} fields &nbsp;|&nbsp; Custom: {_cf_count}"
                f"</div></div>",
                unsafe_allow_html=True
            )

        _use_conf_disp = st.session_state.get("use_conf_threshold", True)
        _conf          = st.session_state.get("conf_threshold", 80)
        if _use_conf_disp:
            _bar_col = "#3fb950" if _conf >= 70 else "#d29922" if _conf >= 40 else "#f85149"
            st.markdown(
                f"<div style='margin-bottom:10px;'>"
                f"<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
                f"font-weight:bold;margin-bottom:3px;'>Confidence Threshold</div>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div class='conf-bar-wrap' style='flex:1;'>"
                f"<div class='conf-bar-fill' style='width:{_conf}%;background:{_bar_col};'>"
                f"</div></div>"
                f"<span style='color:{_bar_col};font-size:12px;font-weight:bold;'>"
                f"{_conf}%</span></div></div>",
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='margin-bottom:10px;background:#161b22;border:1px solid #30363d;"
                "border-radius:6px;padding:6px 10px;font-size:11px;color:#8b949e;'>"
                "⚠ Confidence threshold disabled</div>",
                unsafe_allow_html=True
            )

        def _sanitize_for_json(obj):
            if isinstance(obj, dict): return {k: _sanitize_for_json(v) for k, v in obj.items()}
            if isinstance(obj, list): return [_sanitize_for_json(i) for i in obj]
            if isinstance(obj, str):  return normalize_str(obj)
            return obj

        _sheet_meta = {"sheet_name": selected_sheet, "record_count": len(data)}

        st.markdown("<hr style='border-color:#30363d;margin:8px 0;'>", unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
            "font-weight:bold;margin-bottom:6px;'>📄 Standard Export</div>",
            unsafe_allow_html=True
        )

        if st.button("⬇ Download Standard JSON", use_container_width=True, type="primary",
                     key=f"export_std_json_{selected_sheet}"):
            _std_export_data = {}
            for i, row in enumerate(data):
                c_id = detect_claim_id(row, i)
                rec  = {}
                for fld, inf in row.items():
                    if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}", True):
                        mk_key   = f"mod_{selected_sheet}_{c_id}_{fld}"
                        live_val = st.session_state.get(mk_key, None)
                        orig     = inf.get("value", "")
                        final    = live_val if live_val is not None else inf.get("modified", orig)
                        rec[fld] = {
                            "value": final, "original": orig, "edited": final != orig,
                            "excel_row": inf.get("excel_row"),
                            "excel_col": inf.get("excel_col"),
                            "record_index": i,
                        }
                _std_export_data[c_id] = clean_duplicate_fields(rec)
            output   = _sanitize_for_json(
                to_standard_json(_std_export_data, _sheet_meta, totals_data, merged_meta))
            json_str = json.dumps(output, indent=2, ensure_ascii=False)
            save_feature_store(selected_sheet, output)
            st.session_state[f"_std_json_ready_{selected_sheet}"] = json_str

        if st.session_state.get(f"_std_json_ready_{selected_sheet}"):
            st.download_button(
                "📥 Save Standard JSON",
                data=st.session_state[f"_std_json_ready_{selected_sheet}"],
                file_name=f"{selected_sheet}_standard.json",
                mime="application/json", use_container_width=True,
                key=f"dl_std_json_{selected_sheet}"
            )

        st.markdown("<hr style='border-color:#30363d;margin:10px 0;'>", unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
            "font-weight:bold;margin-bottom:6px;'>🔌 Schema Export</div>",
            unsafe_allow_html=True
        )

        _schema_sel = st.selectbox(
            "Schema export format",
            options=["— Select schema format —", "🔵 Guidewire (JSON)", "🟠 Duck Creek (JSON)"],
            key=f"schema_export_sel_{selected_sheet}",
            label_visibility="collapsed"
        )

        if _schema_sel and _schema_sel != "— Select schema format —":
            if st.button("⬇ Generate Export", use_container_width=True,
                         key=f"schema_export_go_{selected_sheet}"):
                if "Guidewire" in _schema_sel:
                    recs     = build_mapped_records_for_export(data, "Guidewire", selected_sheet)
                    gw_json  = _sanitize_for_json(to_guidewire_json(recs, _sheet_meta))
                    json_str = json.dumps(gw_json, indent=2, ensure_ascii=False)
                    save_feature_store(selected_sheet, gw_json)
                    st.session_state[f"_schema_export_data_{selected_sheet}"] = {
                        "data": json_str,
                        "filename": f"{selected_sheet}_Guidewire_ClaimCenter.json",
                        "mime": "application/json",
                        "label": "📥 Save Guidewire JSON"
                    }
                elif "Duck Creek" in _schema_sel:
                    recs     = build_mapped_records_for_export(data, "Duck Creek", selected_sheet)
                    dc_json  = _sanitize_for_json(to_duck_creek_json(recs, _sheet_meta))
                    json_str = json.dumps(dc_json, indent=2, ensure_ascii=False)
                    save_feature_store(selected_sheet, dc_json)
                    st.session_state[f"_schema_export_data_{selected_sheet}"] = {
                        "data": json_str,
                        "filename": f"{selected_sheet}_DuckCreek.json",
                        "mime": "application/json",
                        "label": "📥 Save Duck Creek JSON"
                    }
                st.success("✅ Ready!")

        _exp_ready = st.session_state.get(f"_schema_export_data_{selected_sheet}")
        if _exp_ready:
            st.download_button(
                _exp_ready["label"], data=_exp_ready["data"],
                file_name=_exp_ready["filename"], mime=_exp_ready["mime"],
                use_container_width=True,
                key=f"dl_schema_export_{selected_sheet}"
            )
            hint_file = ("config/guidewire.yaml" if "Guidewire" in _exp_ready["filename"]
                         else "config/duck_creek.yaml")
            st.markdown(
                f"<div style=\"background:#161b22;border:1px solid #30363d;border-radius:6px;"
                f"padding:6px 10px;margin-top:6px;font-size:10px;color:#8b949e;\">"
                f"📁 Schema config: <code>{hint_file}</code></div>",
                unsafe_allow_html=True
            )

        st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)

        if merged_meta:
            st.markdown(
                "<p style='color:#8b949e;font-weight:bold;font-size:11px;"
                "text-transform:uppercase;margin-top:12px;'>Merged Regions</p>",
                unsafe_allow_html=True
            )
            sorted_merges = sorted(
                [(k, v) for k, v in merged_meta.items() if v["value"]],
                key=lambda x: (x[1]["row_start"], x[1]["col_start"])
            )
            for key, m in sorted_merges[:8]:
                type_color = ("#58a6ff" if m["type"] == "TITLE"
                              else "#d29922" if m["type"] == "HEADER"
                              else "#8b949e")
                st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;
                                padding:6px 10px;margin-bottom:4px;">
                        <div style="font-size:10px;color:{type_color};">
                            {m['type']} · R{m['row_start']}C{m['col_start']}
                            →R{m['row_end']}C{m['col_end']}
                        </div>
                        <div style="font-size:12px;color:#c9d1d9;margin-top:2px;">
                            {m['value'][:35]}
                        </div>
                    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# LLM-ASSISTED INTERPRETATION  (OpenAI / ChatGPT)
# Tasks covered:
#   1. Identify unstructured fields like claim descriptions
#   2. Design prompts for claim summarization
#   3. Extract structured insights from text fields
#   4. Integrate LLM processing with the pipeline
#   5. Validate outputs before final publishing
# ══════════════════════════════════════════════════════════════════════════════

_UNSTRUCTURED_KEYWORDS = [
    "description", "narrative", "notes", "comment", "remark",
    "cause", "nature", "detail", "summary", "loss desc",
    "injury", "accident", "incident", "report",
]

def identify_unstructured_fields(claim_row: dict) -> list:
    """Task 1 — Find free-text / unstructured fields in a claim row."""
    found = []
    for field, info in claim_row.items():
        fn  = field.lower().replace("_", " ").strip()
        val = info.get("value", "") or ""
        if any(kw in fn for kw in _UNSTRUCTURED_KEYWORDS) and len(val.strip()) > 3:
            found.append((field, val.strip()))
    return found


def build_claim_summary_prompt(claim_row: dict, claim_id: str,
                                unstructured: list) -> str:
    """Task 2 — Build the insurance-domain prompt for ChatGPT."""
    field_lines   = [f"  - {f}: {i.get('value','').strip()}"
                     for f, i in claim_row.items() if i.get("value","").strip()]
    unstruct_text = "\n".join(f"  [{fn}]: {fv}" for fn, fv in unstructured) \
                    if unstructured else "  (none detected)"

    return f"""You are an expert insurance claims analyst. Analyze the following loss run claim record and provide a structured response.

CLAIM ID: {claim_id}

CLAIM DATA:
{chr(10).join(field_lines)}

UNSTRUCTURED / FREE-TEXT FIELDS:
{unstruct_text}

Please provide:
1. CLAIM SUMMARY: A concise 2-3 sentence plain-English summary of this claim.
2. KEY INSIGHTS: Up to 5 bullet points of notable findings from the text fields.
3. RISK FLAGS: Any red flags, anomalies, or items needing manual review.
4. RECOMMENDED ACTION: One-line recommended next step for the claims reviewer.
5. FIELD VALIDATION: For each unstructured field, confirm if the value is plausible or flag issues.

Respond in this exact JSON format:
{{
  "claim_summary": "...",
  "key_insights": ["...", "..."],
  "risk_flags": ["...", "..."],
  "recommended_action": "...",
  "field_validation": {{
    "<field_name>": {{"status": "ok" | "flag", "note": "..."}}
  }}
}}"""


def call_openai_llm(prompt: str, api_key: str,
                    model: str = "gpt-4o-mini") -> dict:
    """Task 4 — Call OpenAI ChatCompletions and return parsed JSON result."""
    import urllib.request
    import urllib.error

    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system",
             "content": "You are an insurance claims analyst. Always respond with valid JSON only."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.2,
        "max_tokens": 1000,
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        content = raw["choices"][0]["message"]["content"].strip()
        if content.startswith("```"):
            content = re.sub(r"^```(?:json)?\s*", "", content)
            content = re.sub(r"\s*```$", "", content.strip())
        return {"ok": True, "result": json.loads(content)}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        return {"ok": False, "error": f"HTTP {e.code}: {err_body[:300]}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def validate_llm_output(llm_result: dict) -> tuple:
    """Task 5 — Validate LLM response before publishing. Returns (is_valid, issues)."""
    required_keys = [
        "claim_summary", "key_insights",
        "risk_flags", "recommended_action", "field_validation"
    ]
    issues = []
    if not isinstance(llm_result, dict):
        return False, ["LLM response is not a JSON object."]
    for k in required_keys:
        if k not in llm_result:
            issues.append(f"Missing key: '{k}'")
    if "claim_summary" in llm_result and (
            not isinstance(llm_result["claim_summary"], str)
            or len(llm_result["claim_summary"].strip()) < 10):
        issues.append("'claim_summary' is too short or not a string.")
    if "key_insights" in llm_result:
        if not isinstance(llm_result["key_insights"], list):
            issues.append("'key_insights' must be a list.")
        elif len(llm_result["key_insights"]) == 0:
            issues.append("'key_insights' is empty.")
    if "risk_flags" in llm_result and not isinstance(llm_result["risk_flags"], list):
        issues.append("'risk_flags' must be a list.")
    if "field_validation" in llm_result:
        fv = llm_result["field_validation"]
        if not isinstance(fv, dict):
            issues.append("'field_validation' must be an object.")
        else:
            for fname, finfo in fv.items():
                if not isinstance(finfo, dict):
                    issues.append(f"field_validation['{fname}'] must be an object.")
                elif finfo.get("status") not in ("ok", "flag"):
                    issues.append(
                        f"field_validation['{fname}'].status must be 'ok' or 'flag'.")
    return len(issues) == 0, issues


def render_llm_panel(claim_row: dict, claim_id: str, selected_sheet: str):
    """Full Streamlit UI for LLM-Assisted Interpretation."""

    st.markdown(
        "<div style='font-family:\"Segoe UI\",Arial,sans-serif;font-size:18px;"
        "font-weight:600;color:#e6edf3;margin-bottom:4px;'>"
        "🤖 LLM-Assisted Interpretation</div>"
        "<div style='font-size:12px;color:#8b949e;margin-bottom:10px;'>"
        "Powered by OpenAI ChatGPT · Summarization, insight extraction "
        "&amp; pre-publish validation.</div>",
        unsafe_allow_html=True,
    )

    # API Key
    llm_key_skey = "llm_openai_api_key"
    if llm_key_skey not in st.session_state:
        st.session_state[llm_key_skey] = ""

    with st.expander("🔑 OpenAI API Key",
                     expanded=st.session_state[llm_key_skey] == ""):
        col_key, col_save = st.columns([5, 1])
        with col_key:
            entered_key = st.text_input(
                "API key", value=st.session_state[llm_key_skey],
                type="password", key="llm_key_input",
                label_visibility="collapsed", placeholder="sk-…",
            )
        with col_save:
            if st.button("Save", key="llm_key_save_btn", use_container_width=True):
                st.session_state[llm_key_skey] = entered_key.strip()
                st.success("Saved.")
                st.rerun()
        st.markdown(
            "<div style='font-size:10px;color:#8b949e;margin-top:4px;'>"
            "⚠ Key stored in session only — never written to disk.</div>",
            unsafe_allow_html=True,
        )

    api_key = st.session_state.get(llm_key_skey, "").strip()
    key_ok  = api_key.startswith("sk-") and len(api_key) > 20

    # ── Task 1: Unstructured fields ──────────────────────────────────────
    unstructured = identify_unstructured_fields(claim_row)
    st.markdown(
        "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
        "text-transform:uppercase;margin:10px 0 4px;'>"
        "① Unstructured Fields Detected</div>",
        unsafe_allow_html=True,
    )
    if unstructured:
        for fname, fval in unstructured:
            preview = fval[:120] + ("…" if len(fval) > 120 else "")
            st.markdown(
                f"<div style='background:#161b22;border:1px solid #30363d;"
                f"border-left:3px solid #58a6ff;border-radius:4px;"
                f"padding:6px 10px;margin-bottom:4px;font-size:12px;'>"
                f"<span style='color:#58a6ff;font-weight:bold;'>{fname}</span>"
                f"<span style='color:#8b949e;margin-left:8px;'>{preview}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            "<div style='color:#8b949e;font-size:12px;padding:4px 0;'>"
            "No unstructured text fields detected in this claim.</div>",
            unsafe_allow_html=True,
        )

    # ── Task 2: Prompt preview ──────────────────────────────────────────
    st.markdown(
        "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
        "text-transform:uppercase;margin:10px 0 4px;'>"
        "② Prompt (Claim Summarization)</div>",
        unsafe_allow_html=True,
    )
    with st.expander("View prompt that will be sent to ChatGPT", expanded=False):
        st.code(
            build_claim_summary_prompt(claim_row, claim_id, unstructured),
            language="text"
        )

    llm_model = st.selectbox(
        "ChatGPT model",
        options=["gpt-4o-mini", "gpt-4o", "gpt-3.5-turbo"],
        index=0,
        key=f"llm_model_sel_{selected_sheet}_{claim_id}",
        help="gpt-4o-mini is fast and cost-effective; gpt-4o is more thorough.",
    )

    # ── Task 3 + 4: Run LLM ─────────────────────────────────────────────
    run_key = f"llm_result_{selected_sheet}_{claim_id}"
    if st.button(
        "🤖 Run LLM Analysis" if key_ok else "🔑 Add API Key above to Run",
        key=f"llm_run_{selected_sheet}_{claim_id}",
        use_container_width=True,
        type="primary",
        disabled=not key_ok,
    ):
        with st.spinner("Calling ChatGPT…"):
            prompt   = build_claim_summary_prompt(claim_row, claim_id, unstructured)
            response = call_openai_llm(prompt, api_key, model=llm_model)
        if response["ok"]:
            is_valid, val_issues = validate_llm_output(response["result"])
            st.session_state[run_key] = {
                "result":     response["result"],
                "is_valid":   is_valid,
                "val_issues": val_issues,
                "model":      llm_model,
                "ran_at":     datetime.datetime.now().strftime("%H:%M:%S"),
            }
            st.rerun()
        else:
            st.error(f"OpenAI API error: {response['error']}")

    # ── Display cached results ───────────────────────────────────────────
    cached = st.session_state.get(run_key)
    if cached:
        res        = cached["result"]
        is_valid   = cached["is_valid"]
        val_issues = cached["val_issues"]

        # Task 5 — validation badge
        if is_valid:
            st.markdown(
                f"<div style='background:#1c2d1c;border:1px solid #3fb950;"
                f"border-radius:6px;padding:6px 12px;margin:8px 0;"
                f"font-size:12px;color:#3fb950;'>"
                f"✅ Validated — ready to publish &nbsp;"
                f"<span style='color:#8b949e;font-size:10px;'>"
                f"model: {cached['model']} · {cached['ran_at']}</span></div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div style='background:#2d1a08;border:1px solid #f0883e;"
                "border-radius:6px;padding:6px 12px;margin:8px 0;"
                "font-size:12px;color:#f0883e;'>"
                "⚠ Validation issues — review before publishing:<br>"
                + "".join(f"&nbsp;• {i}<br>" for i in val_issues) + "</div>",
                unsafe_allow_html=True,
            )

        # Summary
        st.markdown(
            "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
            "text-transform:uppercase;margin:12px 0 4px;'>③ Claim Summary</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"<div style='background:#161b22;border:1px solid #30363d;"
            f"border-left:4px solid #58a6ff;border-radius:6px;"
            f"padding:10px 14px;font-size:13px;color:#e6edf3;line-height:1.6;'>"
            f"{res.get('claim_summary', '—')}</div>",
            unsafe_allow_html=True,
        )

        # Key insights (Task 3 — structured extraction)
        insights = res.get("key_insights", [])
        if insights:
            st.markdown(
                "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
                "text-transform:uppercase;margin:12px 0 4px;'>"
                "③ Key Insights</div>",
                unsafe_allow_html=True,
            )
            for ins in insights:
                st.markdown(
                    f"<div style='background:#161b22;border:1px solid #30363d;"
                    f"border-radius:4px;padding:5px 10px;margin-bottom:3px;"
                    f"font-size:12px;color:#c9d1d9;'>💡 {ins}</div>",
                    unsafe_allow_html=True,
                )

        # Risk flags
        flags = res.get("risk_flags", [])
        if flags:
            st.markdown(
                "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
                "text-transform:uppercase;margin:12px 0 4px;'>⚠ Risk Flags</div>",
                unsafe_allow_html=True,
            )
            for flag in flags:
                st.markdown(
                    f"<div style='background:#1f0d0d;border:1px solid #f85149;"
                    f"border-radius:4px;padding:5px 10px;margin-bottom:3px;"
                    f"font-size:12px;color:#f85149;'>🚩 {flag}</div>",
                    unsafe_allow_html=True,
                )

        # Recommended action
        rec = res.get("recommended_action", "")
        if rec:
            st.markdown(
                "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
                "text-transform:uppercase;margin:12px 0 4px;'>"
                "✅ Recommended Action</div>",
                unsafe_allow_html=True,
            )
            st.markdown(
                f"<div style='background:#1c2d1c;border:1px solid #3fb950;"
                f"border-radius:4px;padding:8px 12px;"
                f"font-size:13px;color:#3fb950;font-weight:600;'>"
                f"→ {rec}</div>",
                unsafe_allow_html=True,
            )

        # Field validation (Task 5 detail)
        fv = res.get("field_validation", {})
        if fv:
            st.markdown(
                "<div style='font-size:11px;font-weight:bold;color:#8b949e;"
                "text-transform:uppercase;margin:12px 0 4px;'>⑤ Field Validation</div>",
                unsafe_allow_html=True,
            )
            for fname, finfo in fv.items():
                status = finfo.get("status", "ok")
                note   = finfo.get("note", "")
                bg     = "#1c2d1c" if status == "ok" else "#1f0d0d"
                border = "#3fb950" if status == "ok" else "#f85149"
                icon   = "✅" if status == "ok" else "🚩"
                color  = "#3fb950" if status == "ok" else "#f85149"
                st.markdown(
                    f"<div style='background:{bg};border:1px solid {border};"
                    f"border-radius:4px;padding:5px 10px;margin-bottom:3px;"
                    f"font-size:11px;'>"
                    f"<span style='color:{color};font-weight:bold;'>{icon} {fname}</span>"
                    f"<span style='color:#c9d1d9;margin-left:8px;'>{note}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        # Download
        st.download_button(
            "⬇ Download LLM Analysis (JSON)",
            data=json.dumps(
                {"claim_id": claim_id, "model": cached["model"],
                 "ran_at": cached["ran_at"], **res},
                indent=2, ensure_ascii=False,
            ),
            file_name=f"llm_{claim_id}_{selected_sheet}.json",
            mime="application/json",
            key=f"dl_llm_{selected_sheet}_{claim_id}",
            use_container_width=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# LLM PANEL  — rendered below the three-column portal for the active claim
# ══════════════════════════════════════════════════════════════════════════════
if uploaded and "selected_sheet" in dir() and "curr_claim" in dir():
    st.markdown("<hr style='border-color:#30363d;margin:20px 0 10px;'>",
                unsafe_allow_html=True)
    render_llm_panel(curr_claim, curr_claim_id, selected_sheet)

# ══════════════════════════════════════════════════════════════════════════════
# JSON STORE VIEWER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<hr style='border-color:#30363d;margin:30px 0 16px;'>",
            unsafe_allow_html=True)

st.markdown(
    "<div style='font-family:\"Segoe UI\",Arial,sans-serif;font-size:18px;"
    "font-weight:600;color:#e6edf3;margin-bottom:4px;'>🗄 JSON Store Viewer</div>"
    "<div style='font-size:12px;color:#8b949e;margin-bottom:14px;'>"
    "All exported JSON records saved to the feature store, grouped by file → sheet.</div>",
    unsafe_allow_html=True,
)

_store_entries = list_feature_store_entries()

if not _store_entries:
    st.info("No JSON records in the feature store yet. Export a sheet to populate it.")
else:
    from collections import defaultdict
    _by_file = defaultdict(lambda: defaultdict(list))
    for e in _store_entries:
        _by_file[e["filename"] or "—"][e["sheet_name"]].append(e)

    st.button("🔄 Refresh Store", key="store_refresh_btn")

    for _fname, _sheets in _by_file.items():
        with st.expander(f"📁 {_fname}  ({len(_sheets)} sheet(s))", expanded=False):
            for _sname, _entries in _sheets.items():
                st.markdown(
                    f"<div style='font-family:\"Segoe UI\",Arial,sans-serif;"
                    f"font-size:13px;font-weight:600;color:#58a6ff;margin:10px 0 4px;'>"
                    f"📋 {_sname}</div>",
                    unsafe_allow_html=True,
                )
                for _ent in _entries:
                    st.markdown(
                        f"<span style='font-size:10px;color:#8b949e;'>"
                        f"Exported: {_ent['exported_at']} &nbsp;|&nbsp; "
                        f"File: <code style='font-size:9px;color:#58a6ff;'>"
                        f"{_ent['file_hash'][:12] if _ent['file_hash'] != '—' else '—'}…"
                        f"</code></span>",
                        unsafe_allow_html=True,
                    )
                    _jdata = _ent["data"]
                    if isinstance(_jdata, dict):
                        _rows_for_table = []
                        for _claim_id_j, _fields in _jdata.items():
                            if isinstance(_fields, dict):
                                _flat = {"Claim ID": _claim_id_j}
                                for _f, _fv in _fields.items():
                                    if isinstance(_fv, dict):
                                        _flat[_f] = _fv.get("value", str(_fv))
                                    else:
                                        _flat[_f] = str(_fv)
                                _rows_for_table.append(_flat)
                        if _rows_for_table:
                            import pandas as pd
                            st.dataframe(
                                pd.DataFrame(_rows_for_table),
                                use_container_width=True,
                                height=min(200 + len(_rows_for_table) * 35, 500),
                            )
                        else:
                            st.json(_jdata, expanded=False)
                    else:
                        st.json(_jdata, expanded=False)

                    st.download_button(
                        "⬇ Download this export",
                        data=json.dumps(_ent["data"], indent=2, ensure_ascii=False),
                        file_name=_ent["file"],
                        mime="application/json",
                        key=f"dl_store_{_ent['file']}_{_sname}",
                        use_container_width=False,
                    )
                    st.markdown("<hr style='border-color:#21262d;margin:6px 0;'>",
                                unsafe_allow_html=True)

# ── Audit Log ─────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
with st.expander("🔍 Upload Audit Log", expanded=False):
    _audit = _load_audit_log()
    if not _audit:
        st.info("No uploads recorded yet.")
    else:
        import pandas as pd
        _audit_rows = []
        for _a in reversed(_audit):
            _audit_rows.append({
                "Uploaded At":  _a.get("uploaded_at", "—"),
                "Filename":     _a.get("filename", "—"),
                "Sheets":       _a.get("num_sheets", "—"),
                "File SHA-256": _a.get("file_hash", "—")[:24] + "…",
                "Status":       "Duplicate" if _a.get("is_duplicate") else "Unique",
            })
        st.dataframe(pd.DataFrame(_audit_rows), use_container_width=True)
        st.download_button(
            "⬇ Download Full Audit Log (JSON)",
            data=json.dumps(_audit, indent=2, ensure_ascii=False),
            file_name="_audit_log.json",
            mime="application/json",
            key="dl_audit_log",
        )
