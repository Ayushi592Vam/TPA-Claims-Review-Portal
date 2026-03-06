import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import re
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)

# ==============================
# YAML CONFIG LOADER
# ==============================
CONFIG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")

def _parse_yaml_simple(text: str) -> dict:
    """
    Minimal pure-Python YAML parser that handles the subset used in config files:
    - key: value  (string, int, float, blank/null)
    - key:        (null value)
    - - list item
    - Nested indentation
    - # comments
    Does NOT handle anchors, multi-line strings, or complex YAML.
    """
    def _cast(v: str):
        v = v.strip()
        if not v or v.lower() in ("null", "~", ""):
            return None
        if v.lower() == "true":  return True
        if v.lower() == "false": return False
        try:    return int(v)
        except: pass
        try:    return float(v)
        except: pass
        return v.strip('"').strip("'")

    lines   = text.splitlines()
    root    = {}
    stack   = [(0, root)]   # (indent_level, current_dict)
    cur_key = None

    for raw in lines:
        if not raw.strip() or raw.strip().startswith("#"):
            continue
        indent = len(raw) - len(raw.lstrip())
        line   = raw.strip()

        # Pop stack to match current indent
        while len(stack) > 1 and stack[-1][0] >= indent:
            stack.pop()
        parent = stack[-1][1]

        if line.startswith("- "):          # list item
            val = line[2:].strip()
            if cur_key and isinstance(parent, dict):
                if not isinstance(parent.get(cur_key), list):
                    parent[cur_key] = []
                parent[cur_key].append(_cast(val))
        elif ":" in line:
            parts = line.split(":", 1)
            key   = parts[0].strip().strip('"').strip("'")
            val   = parts[1].strip() if len(parts) > 1 else ""
            # Strip inline comment
            if " #" in val:
                val = val[:val.index(" #")].strip()
            cur_key = key
            if val:
                parent[key] = _cast(val)
            else:
                parent[key] = {}
                stack.append((indent + 2, parent[key]))

    return root


def load_schema_config(schema_filename: str) -> dict | None:
    """
    Load a schema config from /config/<schema_filename>.
    Returns None if file not found (app falls back to hardcoded defaults).
    """
    path = os.path.join(CONFIG_DIR, schema_filename)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = _parse_yaml_simple(f.read())
        return raw
    except Exception as e:
        return None


def _merge_schema_from_config(hardcoded: dict, cfg: dict | None) -> dict:
    """
    Merge a loaded YAML config on top of hardcoded defaults.
    YAML wins for required_fields, accepted_fields, field_aliases.
    Hardcoded UI keys (color, icon, css_cls) are preserved.
    """
    if not cfg:
        return hardcoded

    merged = dict(hardcoded)  # start from hardcoded

    schema_block = cfg.get("schema", {})
    for k in ("version", "description"):
        if schema_block.get(k):
            merged[k] = schema_block[k]

    # required_fields from YAML
    if cfg.get("required_fields"):
        rf = cfg["required_fields"]
        if isinstance(rf, dict):  # parsed as dict if items are indented as keys
            rf = list(rf.keys())
        if isinstance(rf, list):
            merged["required_fields"] = [str(f) for f in rf if f]

    # accepted_fields from YAML
    if cfg.get("accepted_fields"):
        af = cfg["accepted_fields"]
        if isinstance(af, dict):
            af = list(af.keys())
        if isinstance(af, list):
            merged["accepted_fields"] = [str(f) for f in af if f]

    # field_aliases from YAML
    if cfg.get("field_aliases") and isinstance(cfg["field_aliases"], dict):
        aliases = {}
        for field, vals in cfg["field_aliases"].items():
            if isinstance(vals, list):
                aliases[field] = [str(v) for v in vals if v]
            elif isinstance(vals, str):
                aliases[field] = [vals]
        if aliases:
            merged["field_aliases"] = aliases

    # confidence field_thresholds
    conf_block = cfg.get("confidence", {})
    if isinstance(conf_block, dict):
        if conf_block.get("field_thresholds") and isinstance(conf_block["field_thresholds"], dict):
            merged["field_thresholds"] = {
                k: int(v) for k, v in conf_block["field_thresholds"].items()
                if v is not None
            }
        if conf_block.get("global_threshold") is not None:
            merged["config_threshold"] = int(conf_block["global_threshold"])

    # export settings
    if cfg.get("export") and isinstance(cfg["export"], dict):
        merged["export_config"] = cfg["export"]

    return merged


# Track which config files loaded successfully (shown in UI)
_CONFIG_LOAD_STATUS = {}

def _load_all_configs(hardcoded_schemas: dict) -> dict:
    """Load all YAML configs and merge into hardcoded schemas."""
    filemap = {"Guidewire": "guidewire.yaml", "Duck Creek": "duck_creek.yaml"}
    result  = {}
    for name, schema in hardcoded_schemas.items():
        fname = filemap.get(name)
        cfg   = load_schema_config(fname) if fname else None
        result[name] = _merge_schema_from_config(schema, cfg)
        _CONFIG_LOAD_STATUS[name] = {
            "file":   fname,
            "loaded": cfg is not None,
            "path":   os.path.join(CONFIG_DIR, fname) if fname else "",
        }
    return result


# ==============================
# UNICODE NORMALIZER
# ==============================
_DASH_TABLE = str.maketrans({
    '\u2013': '-',   # en-dash  –
    '\u2014': '-',   # em-dash  —
    '\u2012': '-',   # figure dash
    '\u2015': '-',   # horizontal bar
    '\u2212': '-',   # minus sign
    '\ufe58': '-',   # small em-dash
    '\ufe63': '-',   # small hyphen-minus
    '\uff0d': '-',   # fullwidth hyphen-minus
    '\u2018': "'",   # left single quote  '
    '\u2019': "'",   # right single quote '
    '\u201c': '"',   # left double quote  "
    '\u201d': '"',   # right double quote "
    '\u00a0': ' ',   # non-breaking space
    '\u202f': ' ',   # narrow no-break space
})

def normalize_str(s: str) -> str:
    """Replace fancy unicode punctuation with plain ASCII equivalents."""
    if not s:
        return s
    return s.translate(_DASH_TABLE)

# ==============================
# SCHEMA DEFINITIONS
# ==============================
_HARDCODED_SCHEMAS = {
    "Guidewire": {
        "color":   "#58a6ff",
        "icon":    "🔵",
        "css_cls": "guide",
        "version": "ClaimCenter 10.x",
        "description": "Guidewire ClaimCenter 10.x compatible format",
        "required_fields": [
            "Claim Number", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Status", "Line of Business", "Policy Number",
        ],
        "accepted_fields": [
            "Claim Number", "Claimant Name", "Loss Date", "Date Reported",
            "Total Incurred", "Total Paid", "Reserve", "Indemnity Paid",
            "Medical Paid", "Expense Paid", "Status", "Line of Business",
            "Policy Number", "Policy Period Start", "Policy Period End",
            "Carrier", "Insured Name", "Description of Loss",
            "Cause of Loss", "Litigation Flag", "Adjuster Name",
            "Adjuster Phone", "Branch Code", "Department Code",
            "Coverage Type", "Deductible", "Subrogation Amount",
            "Recovery Amount", "Open/Closed", "Reopen Date",
            "Last Activity Date", "Notes",
        ],
        "field_aliases": {
            "Claim Number":      ["claim_id", "claim number", "claim no", "claim#", "claimid", "claim ref"],
            "Claimant Name":     ["claimant name", "claimant", "insured name", "name", "injured party"],
            "Loss Date":         ["date of loss", "loss date", "loss dt", "date of accident", "incident date"],
            "Date Reported":     ["date reported", "reported date", "report date"],
            "Total Incurred":    ["total incurred", "incurred", "total incurred amount"],
            "Total Paid":        ["total paid", "amount paid", "paid amount"],
            "Reserve":           ["reserve", "outstanding reserve", "case reserve"],
            "Indemnity Paid":    ["indemnity paid", "indemnity", "wage loss paid"],
            "Medical Paid":      ["medical paid", "medical", "med paid"],
            "Expense Paid":      ["expense paid", "expense", "legal expense"],
            "Status":            ["status", "claim status", "open/closed"],
            "Line of Business":  ["line of business", "lob", "coverage line"],
            "Policy Number":     ["policy number", "policy no", "policy#", "policy id"],
            "Insured Name":      ["insured name", "insured", "employer name"],
            "Description of Loss": ["description of loss", "loss description", "description", "narrative"],
            "Cause of Loss":     ["cause of loss", "cause", "type of loss", "peril"],
            "Adjuster Name":     ["adjuster name", "adjuster", "examiner"],
        },
    },
    "Duck Creek": {
        "color":   "#f0883e",
        "icon":    "🟠",
        "css_cls": "duck",
        "version": "Claims 6.x",
        "description": "Duck Creek Claims 6.x transaction format",
        "required_fields": [
            "Claim Id", "Claimant Name", "Loss Date",
            "Total Incurred", "Total Paid", "Reserve",
            "Policy Number", "Claim Status",
        ],
        "accepted_fields": [
            "Claim Id", "Transaction Id", "Claimant Name", "Loss Date",
            "Date Reported", "Total Incurred", "Total Paid", "Reserve",
            "Indemnity Paid", "Medical Paid", "Expense Paid",
            "Policy Number", "Policy Effective Date", "Policy Expiry Date",
            "Claim Status", "Cause of Loss", "Description of Loss",
            "Insured Name", "Carrier Name", "Line of Business",
            "Adjuster Id", "Adjuster Name", "Office Code",
            "Jurisdiction", "State Code", "Deductible Amount",
            "Subrogation Flag", "Recovery Amount", "Litigation Flag",
            "Date Closed", "Date Reopened", "Last Updated Date", "Notes",
        ],
        "field_aliases": {
            "Claim Id":          ["claim_id", "claim number", "claim no", "claim#", "claimid", "claim ref"],
            "Claimant Name":     ["claimant name", "claimant", "insured name", "name", "injured party"],
            "Loss Date":         ["date of loss", "loss date", "loss dt", "date of accident", "incident date"],
            "Date Reported":     ["date reported", "reported date", "report date"],
            "Total Incurred":    ["total incurred", "incurred", "total incurred amount"],
            "Total Paid":        ["total paid", "amount paid", "paid amount"],
            "Reserve":           ["reserve", "outstanding reserve", "case reserve"],
            "Indemnity Paid":    ["indemnity paid", "indemnity", "wage loss paid"],
            "Medical Paid":      ["medical paid", "medical", "med paid"],
            "Expense Paid":      ["expense paid", "expense", "legal expense"],
            "Claim Status":      ["status", "claim status", "open/closed"],
            "Line of Business":  ["line of business", "lob", "coverage line"],
            "Policy Number":     ["policy number", "policy no", "policy#", "policy id"],
            "Insured Name":      ["insured name", "insured", "employer name"],
            "Description of Loss": ["description of loss", "loss description", "description", "narrative"],
            "Cause of Loss":     ["cause of loss", "cause", "type of loss", "peril"],
            "Carrier Name":      ["carrier", "carrier name", "insurance company"],
        },
    },
}


# Load YAML configs and merge over hardcoded defaults
SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)

# ==============================
# SCHEMA MAPPING + CONFIDENCE ENGINE
# ==============================
def _word_tokens(s: str) -> set:
    """Split a field name into meaningful word tokens, ignoring stopwords."""
    stopwords = {"of", "the", "a", "an", "and", "or", "to", "in", "for"}
    words = re.sub(r"[_/#+]", " ", s.lower()).split()
    return {w for w in words if len(w) > 1 and w not in stopwords}


def _str_similarity(a: str, b: str) -> float:
    """
    Word-token Jaccard similarity — only whole words count, not characters.
    Returns 0.0–1.0: 1.0 = identical token sets, 0.0 = no shared words.
    """
    a_tok = _word_tokens(a)
    b_tok = _word_tokens(b)
    if not a_tok or not b_tok:
        return 0.0
    if a_tok == b_tok:
        return 1.0
    intersection = a_tok & b_tok
    union        = a_tok | b_tok
    return len(intersection) / len(union)


def _header_match_score(excel_col: str, schema_field: str, aliases: list) -> float:
    """Score 0-1: how well an Excel column name matches a schema field."""
    ec_norm = excel_col.lower().replace("_", " ").strip()
    for alias in aliases:
        if ec_norm == alias.lower():
            return 1.0
    best = max((_str_similarity(ec_norm, a.lower()) for a in aliases), default=0.0)
    return max(best, _str_similarity(ec_norm, schema_field.lower()))


def _value_quality_score(value: str, schema_field: str) -> float:
    """Score 0-1: how good the extracted value looks for the given field type."""
    if not value or not value.strip():
        return 0.0
    v = value.strip()
    sf = schema_field.lower()

    if any(x in sf for x in ["date", "loss dt"]):
        import re as _re
        date_patterns = [
            r"\d{2}-\d{2}-\d{4}", r"\d{4}-\d{2}-\d{2}",
            r"\d{2}/\d{2}/\d{4}", r"\d{1,2}/\d{1,2}/\d{2,4}",
        ]
        for p in date_patterns:
            if _re.fullmatch(p, v):
                return 1.0
        return 0.4

    if any(x in sf for x in ["incurred", "paid", "reserve", "amount", "deductible", "recovery"]):
        import re as _re
        clean = v.replace(",", "").replace("$", "").replace("(", "-").replace(")", "")
        try:
            float(clean)
            return 1.0
        except ValueError:
            return 0.3

    if any(x in sf for x in ["id", "number", "no", "code"]):
        if len(v) >= 2:
            return 0.9
        return 0.5

    if "status" in sf:
        known = {"open", "closed", "pending", "reopened", "denied", "settled"}
        if v.lower() in known:
            return 1.0
        return 0.7

    return 0.85 if len(v) > 0 else 0.0


_MIN_HEADER_MATCH = 0.70

def map_claim_to_schema(claim: dict, schema_name: str,
                        title_fields: dict = None) -> dict:
    """
    Map extracted Excel fields to schema fields.
    Falls back to title_fields for policy-level data not found in columns.
    Fields with no match are left unmapped (not returned).
    """
    if schema_name not in SCHEMAS:
        return {}

    schema       = SCHEMAS[schema_name]
    aliases      = schema.get("field_aliases", {})
    accepted     = schema["accepted_fields"]
    title_fields = title_fields or {}
    result       = {}

    for schema_field in accepted:
        field_aliases  = aliases.get(schema_field, [schema_field.lower()])
        best_excel_col = None
        best_header_sc = 0.0
        best_info      = None

        for excel_col, info in claim.items():
            h_sc = _header_match_score(excel_col, schema_field, field_aliases)
            if h_sc > best_header_sc:
                best_header_sc = h_sc
                best_excel_col = excel_col
                best_info      = info

        if best_header_sc >= _MIN_HEADER_MATCH and best_info is not None:
            val  = best_info.get("modified", best_info.get("value", ""))
            v_sc = _value_quality_score(val, schema_field)
            conf = round(best_header_sc * 0.40 * 100 + v_sc * 0.60 * 100)
            result[schema_field] = {
                "excel_field":  best_excel_col,
                "value":        val,
                "header_score": round(best_header_sc * 100),
                "value_score":  round(v_sc * 100),
                "confidence":   conf,
                "is_required":  schema_field in schema["required_fields"],
                "info":         best_info,
                "from_title":   False,
            }

        elif schema_field in title_fields:
            tf   = title_fields[schema_field]
            val  = tf.get("value", "")
            v_sc = _value_quality_score(val, schema_field)
            conf = min(95, round(1.0 * 0.40 * 100 + v_sc * 0.60 * 100))
            result[schema_field] = {
                "excel_field":  f"[title row {tf['excel_row']}]",
                "value":        val,
                "header_score": 100,
                "value_score":  round(v_sc * 100),
                "confidence":   conf,
                "is_required":  schema_field in schema["required_fields"],
                "info":         tf,
                "from_title":   True,
            }

    return result


def extract_title_fields(merged_meta: dict) -> dict:
    """
    Parse policy-level fields from merged title/header cells.
    Returns dict: schema_field_name -> info dict with value + source='title_row'.
    """
    found = {}

    title_rows = sorted(
        [v for v in merged_meta.values() if v.get("value") and v["type"] in ("TITLE", "HEADER")],
        key=lambda x: (x["row_start"], x["col_start"])
    )

    for m in title_rows:
        text = str(m["value"]).strip()
        r, c = m["excel_row"], m["excel_col"]

        def _info(val):
            return {"value": val, "original": val, "modified": val,
                    "source": "title_row", "excel_row": r, "excel_col": c,
                    "title_text": text}

        pol = re.search(r'Policy\s*(?:#|No\.?|Number)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/\.]+)',
                        text, re.IGNORECASE)
        if pol and "Policy Number" not in found:
            found["Policy Number"] = _info(pol.group(1).strip())

        ins = re.search(r'Insured\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if ins and "Insured Name" not in found:
            found["Insured Name"] = _info(ins.group(1).strip())

        carr = re.search(r'Carrier\s*[:\-]\s*([^\|;]+)', text, re.IGNORECASE)
        if carr:
            val = carr.group(1).strip()
            if "Carrier" not in found:
                found["Carrier"] = _info(val)
            if "Carrier Name" not in found:
                found["Carrier Name"] = _info(val)

        period = re.search(
            r'Period\s*[:\-]?\s*'
            r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
            r'[\s\u2013\u2014\-to]+'
            r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})',
            text, re.IGNORECASE
        )
        if period:
            s, e = period.group(1).strip(), period.group(2).strip()
            for k, v in [("Policy Period Start", s), ("Policy Period End", e),
                         ("Policy Effective Date", s), ("Policy Expiry Date", e)]:
                if k not in found:
                    found[k] = _info(v)

        lob_map = [
            (r'commercial\s+general\s+liability', "Commercial General Liability"),
            (r'\bCGL\b',                          "Commercial General Liability"),
            (r'workers[\s\-]+comp',               "Workers Compensation"),
            (r'commercial\s+auto',                "Commercial Auto"),
            (r'commercial\s+property',            "Commercial Property"),
            (r'professional\s+liability',         "Professional Liability"),
            (r'\bE\s*&\s*O\b',                    "Professional Liability"),
            (r'general\s+liability',              "General Liability"),
            (r'\bumbrella\b',                     "Umbrella"),
            (r'excess\s+liability',               "Excess Liability"),
        ]
        for pattern, lob_val in lob_map:
            if re.search(pattern, text, re.IGNORECASE) and "Line of Business" not in found:
                found["Line of Business"] = _info(lob_val)
                break

    return found


# ==============================
# SETTINGS DIALOG
# ==============================
@st.dialog("Settings", width="large")
def show_settings_dialog():
    global SCHEMAS
    st.markdown("### Configuration")

    st.markdown("---")
    st.markdown("#### Extraction Confidence Threshold")
    st.markdown(
        "<div style='color:#8b949e;font-size:13px;margin-bottom:8px;'>"
        "Fields with confidence below this threshold will be flagged for manual review."
        "</div>",
        unsafe_allow_html=True
    )

    conf = st.slider(
        "Confidence threshold", min_value=0, max_value=100,
        value=st.session_state.get("conf_threshold", 80),
        step=5, format="%d%%", label_visibility="visible"
    )
    st.session_state["conf_threshold"] = conf

    bar_color = "#3fb950" if conf >= 70 else "#d29922" if conf >= 40 else "#f85149"
    level_txt = (
        "High confidence — minimal manual review needed" if conf >= 70
        else "Medium — review flagged fields carefully" if conf >= 40
        else "Low — most fields will require manual review"
    )
    st.markdown(
        f"<div class=\"conf-bar-wrap\">"
        f"<div class=\"conf-bar-fill\" style=\"width:{conf}%;background:{bar_color};\"></div>"
        f"</div>"
        f"<div style=\"color:{bar_color};font-size:12px;margin-top:5px;\">{level_txt}</div>",
        unsafe_allow_html=True
    )

    st.markdown("---")
    st.markdown("#### Export Schema")
    st.markdown(
        "<div style='color:#8b949e;font-size:13px;margin-bottom:12px;'>"
        "Activate a schema to map extracted fields to a standard format. "
        "Custom fields can be added per schema."
        "</div>",
        unsafe_allow_html=True
    )

    active_schema = st.session_state.get("active_schema", None)

    for schema_name, schema_def in SCHEMAS.items():
        is_active  = active_schema == schema_name
        border_col = schema_def["color"] if is_active else "#30363d"
        bg_col     = "#1c2128" if is_active else "#161b22"
        active_tag = (
            f"<span style=\"font-size:10px;color:{schema_def['color']};margin-left:8px;font-weight:bold;\">● ACTIVE</span>"
            if is_active else ""
        )
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}", []))

        st.markdown(
            f"<div style=\"background:{bg_col};border:1px solid {border_col};border-radius:8px;"
            f"padding:12px 14px;margin-bottom:4px;\">"
            f"<div style=\"display:flex;align-items:center;\">"
            f"<span style=\"font-size:15px;font-weight:bold;color:white;\">{schema_def['icon']} {schema_name}</span>"
            f"<span style=\"font-size:11px;color:#8b949e;margin-left:8px;\">{schema_def['version']}</span>"
            f"{active_tag}</div>"
            f"<div style=\"font-size:12px;color:#8b949e;margin-top:4px;\">{schema_def['description']}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            lbl = "✓ Deactivate" if is_active else "Activate"
            if st.button(lbl, key=f"activate_{schema_name}", use_container_width=True):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(f"Custom Fields ({custom_count})", key=f"custom_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()

        st.markdown("<div style=\"height:6px;\"></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(
        "<div style='color:#8b949e;font-size:13px;margin-bottom:8px;'>"
        f"Config directory: <code>{CONFIG_DIR}</code>"
        "</div>",
        unsafe_allow_html=True
    )
    for schema_name, status in _CONFIG_LOAD_STATUS.items():
        sc      = SCHEMAS.get(schema_name, {})
        col_st  = sc.get("color", "#8b949e")
        if status["loaded"]:
            badge = f"<span style=\"background:#1c2d1c;border:1px solid #3fb950;border-radius:4px;padding:1px 7px;font-size:10px;color:#3fb950;\">✓ Loaded</span>"
        else:
            badge = f"<span style=\"background:#2d1515;border:1px solid #f85149;border-radius:4px;padding:1px 7px;font-size:10px;color:#f85149;\">✗ Not found — using defaults</span>"
        st.markdown(
            f"<div style=\"background:#161b22;border:1px solid #30363d;border-radius:6px;"
            f"padding:8px 12px;margin-bottom:6px;\">"
            f"<div style=\"display:flex;align-items:center;gap:8px;\">"
            f"<span style=\"color:{col_st};font-weight:bold;font-size:13px;\">{sc.get('icon','')} {schema_name}</span>"
            f"{badge}</div>"
            f"<div style=\"font-size:10px;color:#8b949e;margin-top:3px;\">📄 {status['file']}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.success("✅ Configs reloaded from disk")
        st.rerun()

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("Reset Defaults", use_container_width=True):
            st.session_state["conf_threshold"] = 80
            st.session_state["active_schema"]  = None
            for s in SCHEMAS:
                st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name):
    schema     = SCHEMAS[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state:
        st.session_state[custom_key] = []

    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(
        f"<div style='color:#8b949e;font-size:13px;margin-bottom:12px;'>{schema['description']}</div>",
        unsafe_allow_html=True
    )

    tab_req, tab_accepted, tab_custom = st.tabs([
        "Required Fields", "All Accepted Fields", "My Custom Fields"
    ])

    with tab_req:
        st.markdown(
            "These fields are **required** by the schema and will always be "
            "included when this schema is active."
        )
        pills = "".join(
            f"<span class=\"field-pill field-pill-required\">✓ {f}</span>"
            for f in schema["required_fields"]
        )
        st.markdown(f"<div style=\"margin:12px 0;\">{pills}</div>", unsafe_allow_html=True)

    with tab_accepted:
        st.markdown(
            "All fields **accepted** by this schema. Only these fields can be "
            "added as custom fields."
        )
        optional = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(
            f"<span class=\"field-pill field-pill-required\">✓ {f}</span>"
            for f in schema["required_fields"]
        )
        opt_pills = "".join(
            f"<span class=\"field-pill\">{f}</span>"
            for f in optional
        )
        st.markdown(
            f"<div style=\"margin:10px 0;\"><b style=\"color:#8b949e;font-size:11px;\">REQUIRED</b><br>{req_pills}</div>"
            f"<div style=\"margin:10px 0;\"><b style=\"color:#8b949e;font-size:11px;\">OPTIONAL</b><br>{opt_pills}</div>",
            unsafe_allow_html=True
        )

    with tab_custom:
        st.markdown(
            "Select **optional fields** from the accepted list to include "
            "alongside required fields in the export."
        )
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]

        if available:
            st.markdown("#### Add Optional Field")
            sel_col, add_col = st.columns([4, 1])
            with sel_col:
                chosen = st.selectbox(
                    "Pick field",
                    options=["— select a field —"] + available,
                    key=f"new_field_sel_{schema_name}",
                    label_visibility="collapsed"
                )
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}",
                             use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()
        else:
            st.info("All accepted optional fields have already been added.")

        st.markdown("#### Active Custom Fields")
        if not custom_fields:
            st.markdown(
                "<div style=\"color:#8b949e;font-size:13px;padding:8px 0;\">"
                "No optional fields added yet."
                "</div>",
                unsafe_allow_html=True
            )
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5, 1])
                with cf1:
                    is_req = cf in schema["required_fields"]
                    cls    = "field-pill-required" if is_req else "field-pill-custom"
                    st.markdown(
                        f"<span class=\"field-pill {cls}\">{'✓' if is_req else '+'} {cf}</span>",
                        unsafe_allow_html=True
                    )
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}",
                                 use_container_width=True):
                        st.session_state[custom_key].pop(idx)
                        st.rerun()
            st.markdown("---")
            if st.button("Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []
                st.rerun()

        st.markdown("---")
        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(
            f"<div style=\"background:#161b22;border:1px solid #30363d;"
            f"border-radius:8px;padding:10px 14px;\">"
            f"<span style=\"color:#8b949e;font-size:12px;\">"
            f"Required: <b style=\"color:#58a6ff;\">{len(schema['required_fields'])}</b> &nbsp;|&nbsp; "
            f"Custom: <b style=\"color:#3fb950;\">{len(custom_fields)}</b> &nbsp;|&nbsp; "
            f"Total export fields: <b style=\"color:white;\">{total}</b>"
            f"</span></div>",
            unsafe_allow_html=True
        )


# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title {
        font-size: 26px; font-weight: 600; padding: 10px 0;
        border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white;
        text-shadow: 0 0 10px rgba(88,166,255,0.7);
    }
    .sheet-title-banner {
        background: #161b22;
        border: 1px solid #30363d;
        border-left: 4px solid #58a6ff;
        border-radius: 6px;
        padding: 10px 16px;
        margin-bottom: 14px;
    }
    .sheet-title-label {
        font-size: 10px;
        color: #8b949e;
        text-transform: uppercase;
        font-weight: bold;
        letter-spacing: 1px;
        margin-bottom: 3px;
    }
    .sheet-title-value {
        font-size: 15px;
        color: #e6edf3;
        font-weight: 600;
    }
    .sheet-subtitle-value {
        font-size: 12px;
        color: #8b949e;
        margin-top: 3px;
    }
    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; cursor: pointer;
        box-shadow: 0 0 0 transparent; transition: all .25s ease;
    }
    .claim-card:hover {
        border-color: #58a6ff;
        box-shadow: 0 0 12px rgba(88,166,255,0.6);
        transform: translateY(-2px);
    }
    .selected-card { border-left: 4px solid #58a6ff; background: #1c2128; box-shadow: 0 0 16px rgba(88,166,255,0.8); }
    .status-text     { font-size: 12px; color: #3fb950; margin-top: 5px; }
    .status-progress { font-size: 12px; color: #d29922; margin-top: 5px; }
    .mid-header-title  { font-size: 26px; font-weight: bold; color: white; margin-bottom: 0px; }
    .mid-header-sub    { font-size: 15px; color: #8b949e; margin-top: 5px; margin-bottom: 5px; }
    .mid-header-status { font-size: 13px; color: #3fb950; margin-bottom: 15px; }
    .incurred-label    { font-size: 14px; color: #8b949e; margin-bottom: 0px; }
    .incurred-amount   { font-size: 26px; font-weight: bold; color: #3fb950; margin-top: 0px; margin-bottom: 20px; }
    div[data-baseweb="input"],
    div[data-baseweb="base-input"],
    div[data-baseweb="select"] {
        background-color: #161b22 !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
    }
    div[data-baseweb="select"] input {
        caret-color: transparent !important;
        pointer-events: none !important;
        user-select: none !important;
        cursor: pointer !important;
    }
    div[data-baseweb="select"]:focus-within {
        border: 1px solid #30363d !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-baseweb="select"] [data-baseweb="input"] {
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus {
        border-color: #30363d !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-baseweb="input"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background-color: transparent !important;
        font-size: 15px !important;
        padding: 8px 12px !important;
    }
    div[data-baseweb="input"]:has(input:disabled),
    div[data-baseweb="base-input"]:has(input:disabled) {
        background-color: transparent !important;
        border: none !important;
    }
    div[data-baseweb="input"] input:disabled {
        color: #e6edf3 !important;
        -webkit-text-fill-color: #e6edf3 !important;
        cursor: default !important;
        padding-left: 0px !important;
    }
    div[data-testid="stButton"] button {
        background-color: transparent !important;
        color: #8b949e !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        transition: 0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    div[data-testid="stButton"] button:disabled { opacity: 0.3 !important; }
    /* Keep action icon buttons (👁 ✏) fixed-size and inline on all viewports */
    div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] button {
        min-width: 32px !important;
        max-width: 44px !important;
        width: 100% !important;
        padding: 4px 2px !important;
        font-size: 14px !important;
        overflow: hidden !important;
        white-space: nowrap !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    div[role="dialog"] {
        background-color: #0d1117 !important;
        border: 1px solid #30363d !important;
        border-radius: 10px !important;
    }
    div[role="dialog"] * { color: #c9d1d9 !important; }
    div[role="dialog"] button {
        background-color: transparent !important;
        border: 1px solid #30363d !important;
        color: #8b949e !important;
    }
    div[role="dialog"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    .format-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 12px; margin-bottom: 8px;
    }
    .format-card.selected {
        border-color: #58a6ff;
        box-shadow: 0 0 10px rgba(88,166,255,0.4);
    }
    .merged-badge {
        display:inline-block; background:#1c2128; border:1px solid #58a6ff;
        border-radius:4px; padding:1px 6px; font-size:10px; color:#58a6ff;
        margin-left:6px; vertical-align:middle;
    }
    .totals-badge {
        display:inline-block; background:#1c2128; border:1px solid #3fb950;
        border-radius:4px; padding:1px 6px; font-size:10px; color:#3fb950;
        margin-left:6px; vertical-align:middle;
    }
    div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] {
        display: none !important;
    }
    div[data-testid="stForm"] {
        border: none !important;
        padding: 0 !important;
    }
    section[data-testid="stVerticalBlock"] div[data-testid="stVerticalBlockBorderWrapper"] > div {
        max-height: none !important;
    }
    div[data-testid="stVerticalBlock"] > div[style*="overflow"] {
        height: calc(100vh - 180px) !important;
        max-height: calc(100vh - 180px) !important;
    }
    .export-sel-btn div[data-testid="stButton"],
    .export-sel-btn div[data-testid="stButton"] > button {
        height: 38px !important;
        min-height: 38px !important;
        max-height: 38px !important;
    }
    .export-sel-btn div[data-testid="stButton"] > button {
        width: 100% !important;
        padding: 0 !important;
        font-size: 11px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        line-height: 38px !important;
    }
    .export-sel-btn div[data-testid="stButton"] p,
    .export-sel-btn div[data-testid="stButton"] span {
        margin: 0 !important;
        padding: 0 !important;
        line-height: 38px !important;
        min-height: unset !important;
        height: 38px !important;
        display: inline !important;
    }
    .settings-btn div[data-testid="stButton"] button {
        background: transparent !important;
        border: 1px solid #30363d !important;
        border-radius: 8px !important;
        color: #8b949e !important;
        font-size: 18px !important;
        padding: 4px 10px !important;
        transition: all 0.2s !important;
    }
    .settings-btn div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background: #1c2128 !important;
        box-shadow: 0 0 8px rgba(88,166,255,0.4) !important;
    }
    .schema-badge {
        display: inline-flex; align-items: center; gap: 6px;
        background: #1c2128; border: 1px solid #58a6ff;
        border-radius: 20px; padding: 3px 10px;
        font-size: 11px; color: #58a6ff; font-weight: 600;
        margin-left: 10px; vertical-align: middle;
    }
    .schema-badge-duck  { border-color: #f0883e !important; color: #f0883e !important; }
    .schema-badge-guide { border-color: #58a6ff !important; color: #58a6ff !important; }
    .conf-bar-wrap {
        background: #21262d; border-radius: 6px; height: 8px;
        width: 100%; margin-top: 4px; overflow: hidden;
    }
    .conf-bar-fill {
        height: 100%; border-radius: 6px;
        background: linear-gradient(90deg, #3fb950, #58a6ff);
        transition: width 0.3s ease;
    }
    .field-pill {
        display: inline-block; background: #161b22;
        border: 1px solid #30363d; border-radius: 12px;
        padding: 3px 10px; font-size: 11px; color: #c9d1d9;
        margin: 2px 3px;
    }
    .field-pill-required { border-color: #58a6ff !important; color: #58a6ff !important; background: #1c2128 !important; }
    .field-pill-custom   { border-color: #3fb950 !important; color: #3fb950 !important; background: #1c2128 !important; }
    div[role="dialog"] .stSlider { padding: 0 !important; }
</style>
""", unsafe_allow_html=True)


# ==============================
# SHEET NAMES
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others


# ==============================
# CLASSIFICATION
# ==============================
def classify_sheet(rows):
    text = " ".join(
        str(cell).lower()
        for row in rows[:20]
        for cell in row
        if cell
    )
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id",
        "claim ref", "claimant", "file number", "file no"
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date"
    ])
    has_financial = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense"
    ])
    if has_claim and (has_loss or has_financial):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    """Extract merged cell ranges and classify them as title/header/data."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return {}

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    merged_info = {}

    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        cell = ws.cell(mn_r, mn_c)
        val  = str(cell.value).strip() if cell.value else ""

        span_cols = mx_c - mn_c + 1
        span_rows = mx_r - mn_r + 1

        if mn_r <= 3 and span_cols >= 3:
            region_type = "TITLE"
        elif span_cols >= 2 and span_rows == 1:
            region_type = "HEADER"
        else:
            region_type = "DATA"

        key = f"R{mn_r}C{mn_c}"
        merged_info[key] = {
            "value":       val,
            "type":        region_type,
            "row_start":   mn_r,
            "col_start":   mn_c,
            "row_end":     mx_r,
            "col_end":     mx_c,
            "span_cols":   span_cols,
            "span_rows":   span_rows,
            "excel_row":   mn_r,
            "excel_col":   mn_c,
        }

    wb.close()
    return merged_info


def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    """Find and extract totals/summary rows from the sheet."""
    ext    = os.path.splitext(file_path)[1].lower()
    totals = {}

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows = raw_rows
        wb.close()

    if not rows:
        return totals

    header_row_index = None
    headers = []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and ("date" in row_text or "incurred" in row_text or "paid" in row_text):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}" for j, h in enumerate(row)]
            break

    if header_row_index is None or not headers:
        return totals

    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index + 1:]):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total", "subtotal", "grand total", "sum", "totals"]):
            row_data = {}
            cell_row = cell_rows[header_row_index + 1 + r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers):
                    continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column if hasattr(cell_row[c_idx_0], 'column') else c_idx_0 + 1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0 + 1
                if clean_val:
                    row_data[headers[c_idx_0]] = {
                        "value":     clean_val,
                        "excel_row": r_idx,
                        "excel_col": real_col,
                    }
            if row_data:
                totals_rows.append(row_data)

    if totals_rows:
        totals["rows"] = totals_rows
        totals["excel_row"] = totals_rows[0].get(list(totals_rows[0].keys())[0], {}).get("excel_row", 9999)
        agg = {}
        for row_data in totals_rows:
            for field, info in row_data.items():
                try:
                    num = float(str(info["value"]).replace(",", "").replace("$", ""))
                    if field not in agg:
                        agg[field] = 0.0
                    agg[field] += num
                except:
                    pass
        totals["aggregated"] = {k: round(v, 2) for k, v in agg.items()}

    return totals


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute else value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return f"{int(value)}.0"
        formatted = f"{value:.10f}".rstrip('0')
        if '.' not in formatted:
            formatted += '.0'
        return formatted
    return normalize_str(str(value).strip())


def _apply_date_number_format(dt, nf: str) -> str:
    if not nf or nf.lower() in ("general", "@", ""):
        return dt.strftime("%m-%d-%Y")

    fmt = re.sub(r'\[.*?\]', '', nf)
    fmt = re.sub(r'["_*\\]', '', fmt)

    result = fmt
    result = re.sub(r'(?i)(?<=h)mm', '__MIN__', result)
    result = re.sub(r'(?i)mm(?=ss)', '__MIN__', result)

    def _tok(m):
        tok = m.group(0).lower()
        return {
            'yyyy': '%Y', 'yy': '%y',
            'mmmm': '%B', 'mmm': '%b', 'mm': '%m', '__min__': '%M', 'm': '%m',
            'dd': '%d', 'd': '%d',
            'hh': '%H', 'h': '%H',
            'ss': '%S', 's': '%S',
            'am/pm': '%p', 'a/p': '%p',
        }.get(tok, m.group(0))

    result = re.sub(
        r'(?i)yyyy|yy|mmmm|mmm|__min__|mm|dd|hh|ss|am/pm|a/p|d|h|s|m',
        _tok, result
    )
    try:
        return dt.strftime(result)
    except Exception:
        return dt.strftime("%m-%d-%Y")


def format_cell_value_with_fmt(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    nf = (cell.number_format or "").strip()

    if isinstance(value, (datetime.datetime, datetime.date)):
        return _apply_date_number_format(value, nf)

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, (int, float)):
        decimal_places = None

        if nf and nf.lower() not in ("general", "@", ""):
            clean_nf = re.sub(r'[$€£¥"_*\\]', '', nf)
            is_date_fmt = (
                any(x in clean_nf.lower() for x in ['yy', 'mm', 'dd', 'hh', 'ss'])
                and not any(ch in clean_nf for ch in ['0', '#'])
            )
            if not is_date_fmt:
                if '.' in clean_nf:
                    after_dot = clean_nf.split('.')[1]
                    after_dot = re.sub(r'\[.*?\]', '', after_dot)
                    dp = sum(1 for ch in after_dot if ch in '0#')
                    decimal_places = dp
                else:
                    decimal_places = 0

        if decimal_places is not None:
            fval = float(value)
            if decimal_places == 0:
                return str(int(round(fval)))
            return f"{fval:.{decimal_places}f}"

        if isinstance(value, int):
            return str(value)

        fval = float(value)
        remainder = fval - int(fval)
        if remainder == 0.0:
            return f"{fval:.2f}"
        else:
            formatted = f"{fval:.10f}".rstrip('0')
            if '.' not in formatted:
                formatted += '.00'
            elif len(formatted.split('.')[1]) < 2:
                formatted = f"{fval:.2f}"
            return formatted

    return normalize_str(str(value).strip())


def extract_from_excel(file_path, sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN"
        sheet_type = classify_sheet(rows)
        return parse_rows(sheet_type, rows)
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows   = []
        cell_rows  = []
        for row in ws.iter_rows():
            raw_rows.append([cell.value for cell in row])
            cell_rows.append(list(row))
        wb.close()

        if not raw_rows:
            return [], "UNKNOWN"

        sheet_type = classify_sheet(raw_rows)
        return parse_rows_with_cells(sheet_type, raw_rows, cell_rows)


def parse_rows_with_cells(sheet_type, rows, cell_rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type

        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(rows[header_row_index])
        ]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(
            zip(rows[header_row_index + 1:], cell_rows[header_row_index + 1:])
        ):
            r_idx = header_row_index + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                header    = headers[c_idx_0]
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0 + 1
                row_data[header] = {
                    "value":    clean_val,
                    "modified": clean_val,
                    "excel_row": r_idx,
                    "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            break

    if header_row_index is None:
        return [], sheet_type

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(rows[header_row_index])
    ]
    extracted = []
    data_rows      = rows[header_row_index + 1:]
    data_cell_rows = cell_rows[header_row_index + 1:]

    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(data_rows, data_cell_rows)):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total"] for c in raw_row if c):
            break
        row_data = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header    = headers[c_idx_0]
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value":     clean_val,
                "modified":  clean_val,
                "excel_row": r_idx,
                "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def parse_rows(sheet_type, rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type
        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(rows[header_row_index])
        ]
        extracted = []
        for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if not any(row):
                continue
            row_data = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                header    = headers[c_idx - 1]
                clean_val = str(value).strip() if value is not None else ""
                row_data[header] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            break

    if header_row_index is None:
        return [], sheet_type

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(rows[header_row_index])
    ]
    extracted = []
    for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        row_data = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


# ==============================
# EXCEL CELL RENDERER
# ==============================
_THEME_COLORS = {
    0: "FFFFFF", 1: "000000", 2: "EEECE1", 3: "1F497D",
    4: "4F81BD", 5: "C0504D", 6: "9BBB59", 7: "8064A2",
    8: "4BACC6", 9: "F79646",
}


def _resolve_color(color_obj, default="FFFFFF") -> str:
    if color_obj is None:
        return default
    t = color_obj.type
    if t == "rgb":
        rgb = color_obj.rgb or ""
        if len(rgb) == 8 and rgb not in ("00000000", "FF000000"):
            return rgb[2:]
        if len(rgb) == 6:
            return rgb
        return default
    if t == "theme":
        base = _THEME_COLORS.get(color_obj.theme, default)
        tint = color_obj.tint or 0.0
        if tint != 0.0:
            r, g, b = int(base[0:2], 16), int(base[2:4], 16), int(base[4:6], 16)
            if tint > 0:
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            else:
                r = int(r * (1 + tint))
                g = int(g * (1 + tint))
                b = int(b * (1 + tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        return base
    if t == "indexed":
        indexed_map = {
            0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00",
            4: "0000FF", 5: "FFFF00", 6: "FF00FF", 7: "00FFFF",
            64: "000000", 65: "FFFFFF",
        }
        return indexed_map.get(color_obj.indexed, default)
    return default


def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd = ws.column_dimensions.get(letter)
    w  = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(20, int(w * 8 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str,
                        scale: float = 1.0) -> tuple:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))

    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    img_w = col_starts[-1]
    img_h = row_starts[-1]

    img  = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    merged_master: dict = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    drawn_merges: set = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            merge_info = merged_master.get((r, c))

            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue
                drawn_merges.add((mn_r, mn_c))
                x1 = col_starts[mn_c - 1]; y1 = row_starts[mn_r - 1]
                x2 = col_starts[mx_c];      y2 = row_starts[mx_r]
                cell = ws.cell(mn_r, mn_c)
            else:
                x1 = col_starts[c - 1]; y1 = row_starts[r - 1]
                x2 = col_starts[c];      y2 = row_starts[r]
                cell = ws.cell(r, c)

            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")

            draw.rectangle([x1, y1, x2 - 1, y2 - 1],
                           fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"

                bold    = bool(cell.font and cell.font.bold)
                text    = format_cell_value_with_fmt(cell) if cell.value is not None else ""
                cell_w  = x2 - x1
                ch_w    = 7 if not bold else 8
                max_chars = max(1, (cell_w - 8) // ch_w)
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"

                draw.text((x1 + 4, y1 + 4), text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts, merged_master


def get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col,
                        merged_master=None):
    c = max(1, min(target_col, len(col_starts) - 1))
    r = max(1, min(target_row, len(row_starts) - 1))

    if merged_master:
        info = merged_master.get((r, c))
        if info:
            mn_r, mn_c, mx_r, mx_c = info
            x1 = col_starts[mn_c - 1]
            y1 = row_starts[mn_r - 1]
            x2 = col_starts[min(mx_c, len(col_starts) - 1)]
            y2 = row_starts[min(mx_r, len(row_starts) - 1)]
            return x1, y1, x2, y2

    x1 = col_starts[c - 1]
    y1 = row_starts[r - 1]
    x2 = col_starts[min(c, len(col_starts) - 1)]
    y2 = row_starts[min(r, len(row_starts) - 1)]
    return x1, y1, x2, y2


def crop_context(img, x1, y1, x2, y2, pad_x=220, pad_y=160):
    iw, ih = img.size
    cx1 = max(0, x1 - pad_x);  cy1 = max(0, y1 - pad_y)
    cx2 = min(iw, x2 + pad_x); cy2 = min(ih, y2 + pad_y)
    cropped = img.crop((cx1, cy1, cx2, cy2))
    return cropped, x1 - cx1, y1 - cy1, x2 - cx1, y2 - cy1


# ==============================
# EYE POPUP
# ==============================
@st.dialog("Cell View", width="large")
def show_eye_popup(field, info, excel_path, sheet_name):
    st.markdown(f"### 📍 {field}")

    value      = info.get("modified", info["value"])
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        st.markdown("**Extracted Value**")
        st.code(value if value else "(empty)")
    with col_b:
        r_lbl = target_row or "?"
        c_lbl = target_col or "?"
        col_letter = get_column_letter(target_col) if target_col else "?"
        st.markdown(f"""
            <div style="padding:10px 0; color:#8b949e; font-size:14px;">
                📌 Cell: <span style="color:#58a6ff; font-weight:bold;">{col_letter}{r_lbl}</span>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                Row <span style="color:#c9d1d9;">{r_lbl}</span> · Col <span style="color:#c9d1d9;">{c_lbl}</span>
            </div>
        """, unsafe_allow_html=True)

    if not target_row or not target_col:
        st.warning("No cell location recorded for this field.")
        return

    ext = os.path.splitext(excel_path)[1].lower()
    if ext == ".csv":
        st.info("Cell preview is not available for CSV files.")
        return

    st.markdown("---")
    st.markdown("**📊 Excel Cell Location**")

    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0
            )
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]

    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")

        x1, y1, x2, y2 = get_cell_pixel_bbox(
            col_starts, row_starts, target_row, target_col, merged_master
        )

        draw.rectangle([x1 + 1, y1 + 1, x2 - 1, y2 - 1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(255, 180, 0, 255), width=3)
        draw.rectangle([x1 + 3, y1 + 3, x2 - 3, y2 - 3], outline=(255, 255, 255, 160), width=1)

        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)

        col_letter = get_column_letter(target_col)
        st.image(
            cropped,
            use_container_width=True,
            caption=f"Cell {col_letter}{target_row}  ·  Value: {value or '(empty)'}"
        )

    except Exception as e:
        st.error(f"Rendering error: {e}")
        import traceback
        st.code(traceback.format_exc())


# ==============================
# FORMAT CONVERTERS
# ==============================
def to_duck_creek_xml(mapped_records: list, sheet_meta: dict) -> str:
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    root = ET.Element("ClaimTransactionBatch")
    root.set("xmlns", "http://www.duckcreek.com/claims/transaction/v6")
    root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    root.set("batchDate", datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))
    root.set("source", "TPA_Claims_Review_Portal")
    root.set("recordCount", str(len(mapped_records)))

    _DC_XML_MAP = {
        "Claim Id":          "ClaimId",
        "Transaction Id":    "TransactionId",
        "Claimant Name":     "ClaimantName",
        "Loss Date":         "LossDate",
        "Date Reported":     "DateReported",
        "Total Incurred":    "TotalIncurred",
        "Total Paid":        "TotalPaid",
        "Reserve":           "Reserve",
        "Indemnity Paid":    "IndemnityPaid",
        "Medical Paid":      "MedicalPaid",
        "Expense Paid":      "ExpensePaid",
        "Policy Number":     "PolicyNumber",
        "Policy Effective Date": "PolicyEffectiveDate",
        "Policy Expiry Date":    "PolicyExpiryDate",
        "Claim Status":      "ClaimStatus",
        "Cause of Loss":     "CauseOfLoss",
        "Description of Loss": "LossDescription",
        "Insured Name":      "InsuredName",
        "Carrier Name":      "CarrierName",
        "Line of Business":  "LineOfBusiness",
        "Adjuster Id":       "AdjusterId",
        "Adjuster Name":     "AdjusterName",
        "Office Code":       "OfficeCode",
        "Jurisdiction":      "Jurisdiction",
        "State Code":        "StateCode",
        "Deductible Amount": "DeductibleAmount",
        "Subrogation Flag":  "SubrogationFlag",
        "Recovery Amount":   "RecoveryAmount",
        "Litigation Flag":   "LitigationFlag",
        "Date Closed":       "DateClosed",
        "Date Reopened":     "DateReopened",
        "Last Updated Date": "LastUpdatedDate",
        "Notes":             "Notes",
    }

    for rec in mapped_records:
        txn = ET.SubElement(root, "ClaimTransaction")
        txn.set("transactionType", "UPDATE")
        txn.set("confidence", str(rec.get("_avg_confidence", "")))

        claim_el = ET.SubElement(txn, "Claim")
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            xml_tag = _DC_XML_MAP.get(schema_field, schema_field.replace(" ", ""))
            el = ET.SubElement(claim_el, xml_tag)
            el.text = str(field_data.get("value", ""))
            if field_data.get("edited"):
                el.set("edited", "true")
                el.set("originalValue", str(field_data.get("original", "")))
            el.set("confidence", str(field_data.get("confidence", "")))

    xml_str = ET.tostring(root, encoding="unicode")
    pretty  = minidom.parseString(xml_str).toprettyxml(indent="  ")
    lines = pretty.split("\n")
    return "\n".join(lines[1:]) if lines[0].startswith("<?xml") else pretty


def to_duck_creek_json(mapped_records: list, sheet_meta: dict) -> dict:
    transactions = []
    for rec in mapped_records:
        claim_obj = {}
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            claim_obj[schema_field] = {
                "value":      field_data.get("value", ""),
                "confidence": field_data.get("confidence", 0),
                "edited":     field_data.get("edited", False),
            }
            if field_data.get("edited"):
                claim_obj[schema_field]["originalValue"] = field_data.get("original", "")

        transactions.append({
            "transactionType":  "UPDATE",
            "avgConfidence":    rec.get("_avg_confidence", 0),
            "claim":            claim_obj,
        })

    return {
        "schema":        "DuckCreek.Claims.Transaction.v6",
        "exportDate":    datetime.datetime.now().isoformat(),
        "source":        "TPA_Claims_Review_Portal",
        "sheetName":     sheet_meta.get("sheet_name", ""),
        "recordCount":   len(transactions),
        "transactions":  transactions,
    }


def to_guidewire_json(mapped_records: list, sheet_meta: dict) -> dict:
    _GW_FIELD_MAP = {
        "Claim Number":        "claimNumber",
        "Claimant Name":       "claimantName",
        "Loss Date":           "lossDate",
        "Date Reported":       "reportedDate",
        "Total Incurred":      "totalIncurredAmount",
        "Total Paid":          "totalPaidAmount",
        "Reserve":             "reserveAmount",
        "Indemnity Paid":      "indemnityPaidAmount",
        "Medical Paid":        "medicalPaidAmount",
        "Expense Paid":        "expensePaidAmount",
        "Status":              "status",
        "Line of Business":    "lineOfBusinessCode",
        "Policy Number":       "policyNumber",
        "Policy Period Start": "policyPeriodStart",
        "Policy Period End":   "policyPeriodEnd",
        "Carrier":             "carrierName",
        "Insured Name":        "insuredName",
        "Description of Loss": "lossDescription",
        "Cause of Loss":       "causeOfLoss",
        "Litigation Flag":     "litigationFlag",
        "Adjuster Name":       "adjusterName",
        "Adjuster Phone":      "adjusterPhone",
        "Branch Code":         "branchCode",
        "Department Code":     "departmentCode",
        "Coverage Type":       "coverageType",
        "Deductible":          "deductibleAmount",
        "Subrogation Amount":  "subrogationAmount",
        "Recovery Amount":     "recoveryAmount",
        "Open/Closed":         "openClosedStatus",
        "Reopen Date":         "reopenDate",
        "Last Activity Date":  "lastActivityDate",
        "Notes":               "notes",
    }

    claims = []
    for rec in mapped_records:
        claim_obj = {
            "_type":          "cc.Claim",
            "_confidence":    rec.get("_avg_confidence", 0),
        }
        financials  = {}
        has_finance = False
        for schema_field, field_data in rec.items():
            if schema_field.startswith("_"):
                continue
            gw_key = _GW_FIELD_MAP.get(schema_field, schema_field[0].lower() + schema_field[1:].replace(" ", ""))
            val    = field_data.get("value", "")

            if any(x in schema_field.lower() for x in ["paid", "reserve", "incurred", "deductible", "recovery", "subrogation"]):
                financials[gw_key] = {
                    "amount":     val,
                    "currency":   "USD",
                    "confidence": field_data.get("confidence", 0),
                }
                if field_data.get("edited"):
                    financials[gw_key]["originalValue"] = field_data.get("original", "")
                has_finance = True
            else:
                claim_obj[gw_key] = {
                    "value":      val,
                    "confidence": field_data.get("confidence", 0),
                }
                if field_data.get("edited"):
                    claim_obj[gw_key]["originalValue"] = field_data.get("original", "")

        if has_finance:
            claim_obj["financials"] = financials
        claims.append(claim_obj)

    return {
        "schema":         "Guidewire.ClaimCenter.REST.v1",
        "exportDate":     datetime.datetime.now().isoformat(),
        "source":         "TPA_Claims_Review_Portal",
        "sheetName":      sheet_meta.get("sheet_name", ""),
        "recordCount":    len(claims),
        "data": {
            "claims": claims,
        },
    }


def build_mapped_records_for_export(data: list, schema_name: str,
                                     selected_sheet: str) -> list:
    records = []
    schema  = SCHEMAS[schema_name]
    custom_flds = st.session_state.get(f"custom_fields_{schema_name}", [])
    export_flds = list(schema["required_fields"]) + [
        f for f in custom_flds if f not in schema["required_fields"]
    ]

    for i, row in enumerate(data):
        c_id   = detect_claim_id(row, i)
        mapped = map_claim_to_schema(row, schema_name)
        rec    = {}
        confs  = []

        for sf in export_flds:
            if sf not in mapped:
                rec[sf] = {"value": "", "confidence": 0, "edited": False, "original": ""}
                confs.append(0)
                continue
            m        = mapped[sf]
            excel_f  = m["excel_field"]
            mk_key   = f"mod_{selected_sheet}_{c_id}_schema_{sf}"
            live_val = st.session_state.get(mk_key, None)
            orig     = m["info"].get("value", "")
            final    = live_val if live_val is not None else m["value"]
            rec[sf]  = {
                "value":      final,
                "original":   orig,
                "edited":     final != orig,
                "confidence": m["confidence"],
                "excel_row":  m["info"].get("excel_row"),
                "excel_col":  m["info"].get("excel_col"),
            }
            confs.append(m["confidence"])

        rec["_avg_confidence"] = round(sum(confs) / len(confs)) if confs else 0
        rec["_claim_id"]       = c_id
        records.append(rec)
    return records


def to_standard_json(export_data: dict, sheet_meta: dict, totals: dict, merged_meta: dict) -> dict:
    titles_section = []
    sorted_merges = sorted(
        [(k, v) for k, v in merged_meta.items() if v.get("value")],
        key=lambda x: (x[1]["row_start"], x[1]["col_start"])
    )
    for key, m in sorted_merges:
        titles_section.append({
            "type":      m["type"],
            "value":     m["value"],
            "excel_row": m["excel_row"],
            "excel_col": m["excel_col"],
            "span_cols": m["span_cols"],
            "span_rows": m["span_rows"],
        })

    records_section = export_data

    totals_section = {}
    if totals:
        totals_section = {
            "excel_row":  totals.get("excel_row"),
            "rows":       totals.get("rows", []),
            "aggregated": totals.get("aggregated", {}),
        }

    return {
        "exportDate":   datetime.datetime.now().isoformat(),
        "sheetMeta": {
            "sheet_name":    sheet_meta.get("sheet_name"),
            "record_count":  sheet_meta.get("record_count"),
        },
        "titleRows":    titles_section,
        "records":      records_section,
        "totals":       totals_section,
        "recordCount":  len(export_data),
    }


# ==============================
# UTILS
# ==============================
def get_val(claim: dict, keys: list, default: str = "") -> str:
    for pk in keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"] or default
    return default


def detect_claim_id(row, index=None):
    keys = [
        "claim id", "claim_id", "claimid",
        "claim number", "claim no", "claim #",
        "claim ref", "claim reference",
        "file number", "record id"
    ]
    for k, v in row.items():
        name = str(k).lower().replace("_", " ").strip()
        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip():
                return str(val)
    if index is not None:
        return str(index + 1)
    return ""


def clean_duplicate_fields(record: dict) -> dict:
    seen, out = set(), {}
    for k, v in record.items():
        if k.strip() not in seen:
            seen.add(k.strip())
            out[k.strip()] = v
    return out


def save_feature_store(sheet_name: str, data: dict) -> str:
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(FEATURE_STORE_PATH, f"{sheet_name}_{ts}.json")
    def _sanitize(obj):
        if isinstance(obj, dict): return {k: _sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list): return [_sanitize(i) for i in obj]
        if isinstance(obj, str): return normalize_str(obj)
        return obj
    with open(path, "w") as f:
        json.dump(_sanitize(data), f, indent=2, ensure_ascii=False)
    return path


# ==============================
# MAIN APP
# ==============================

for _k, _v in [
    ("conf_threshold", 80),
    ("active_schema",  None),
    ("schema_popup_target", None),
    ("schema_popup_tab",    "required"),
    ("settings_saved", False),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

col_title, col_gear, col_sheet_dropdown = st.columns([3.5, 0.5, 1])
with col_title:
    active_schema = st.session_state.get("active_schema", None)
    badge_html = ""
    if active_schema and active_schema in SCHEMAS:
        sc = SCHEMAS[active_schema]
        badge_html = (
            f'<span class="schema-badge schema-badge-{sc["css_cls"]}" '
            f'style="border-color:{sc["color"]};color:{sc["color"]};background:#1c2128;">'
            f'{sc["icon"]} {active_schema} · {sc["version"]}'
            f'</span>'
        )
    st.markdown(
        f'<div class="main-title">🛡️ TPA Claims Review Portal{badge_html}</div>',
        unsafe_allow_html=True
    )

with col_gear:
    st.markdown("<div style='margin-top:14px;' class='settings-btn'>", unsafe_allow_html=True)
    if st.button("⚙", key="open_settings", help="Open Settings", use_container_width=True):
        show_settings_dialog()
    st.markdown("</div>", unsafe_allow_html=True)

if st.session_state.get("schema_popup_target"):
    _target = st.session_state["schema_popup_target"]
    st.session_state["schema_popup_target"] = None
    show_schema_fields_dialog(_target)


uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path, "wb") as f:
            f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None
        for key in list(st.session_state.keys()):
            if key.startswith("_rendered_"):
                del st.session_state[key]

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox(
            "Sheet", st.session_state.sheet_names,
            index=0,
            label_visibility="collapsed"
        )

    st.markdown("<hr style='border-color:#30363d; margin-top:-10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Reading '{selected_sheet}'..."):
            data, sheet_type = extract_from_excel(excel_path, selected_sheet)
            merged_meta      = extract_merged_cell_metadata(excel_path, selected_sheet)
            totals_data      = extract_totals_row(excel_path, selected_sheet)
            st.info(f"Detected Sheet Type: **{sheet_type}** | Merged Regions: **{len(merged_meta)}** | Totals Found: **{'Yes' if totals_data else 'No'}**")
            if not data:
                st.warning(f"No data found in sheet '{selected_sheet}'.")
                st.stop()
            for row in data:
                for fld, inf in row.items():
                    for key in ("value", "modified"):
                        if key in inf and isinstance(inf[key], str):
                            inf[key] = normalize_str(inf[key])
            _title_flds = extract_title_fields(merged_meta)
            st.session_state.sheet_cache[selected_sheet] = {
                "data":        data,
                "merged_meta": merged_meta,
                "totals":      totals_data,
                "title_fields": _title_flds,
            }
            st.session_state.selected_idx = 0
            st.session_state.focus_field  = None

    active       = st.session_state.sheet_cache[selected_sheet]
    data         = active["data"]
    merged_meta  = active.get("merged_meta", {})
    totals_data  = active.get("totals", {})
    title_fields = active.get("title_fields", {})

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim = data[st.session_state.selected_idx]

    col_nav, col_main, col_fmt = st.columns([1.2, 3.2, 1.4], gap="large")

    # ── LEFT PANEL ──────────────────────────────────────────────────────
    with col_nav:
        with st.container(height=700, border=False):
            st.markdown("<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>TPA Records</p>", unsafe_allow_html=True)
            for i, row_data in enumerate(data):
                is_sel   = "selected-card" if st.session_state.selected_idx == i else ""
                c_id     = detect_claim_id(row_data, i)
                c_name   = get_val(row_data, ["Insured Name", "Name", "Company", "Claimant", "TPA_NAME"], "Unknown Entity")
                raw_st   = get_val(row_data, ["Status", "CLAIM_STATUS"], "")
                c_status = raw_st or ("Yet to Review" if i == 0 else "In Progress" if i == 1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or c_status.lower() == "open" else "status-text"
                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold;color:white;font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e;font-size:13px;margin-top:2px;">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Select", key=f"sel_{selected_sheet}_{i}", use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field  = None
                    st.rerun()

    # ── MIDDLE PANEL ────────────────────────────────────────────────────
    with col_main:

        sorted_titles = sorted(
            [(k, v) for k, v in merged_meta.items() if v.get("value")],
            key=lambda x: (x[1]["row_start"], x[1]["col_start"])
        )
        if sorted_titles:
            main_title_val = ""
            sub_title_val  = ""
            for _, m in sorted_titles:
                if m["type"] == "TITLE":
                    if not main_title_val:
                        main_title_val = m["value"]
                    elif not sub_title_val:
                        sub_title_val = m["value"]
            if main_title_val or sub_title_val:
                st.markdown(f"""
                <div class="sheet-title-banner">
                    <div class="sheet-title-label">📄 Sheet Title</div>
                    <div class="sheet-title-value">{main_title_val}</div>
                    {"" if not sub_title_val else f'<div class="sheet-subtitle-value">{sub_title_val}</div>'}
                </div>
                """, unsafe_allow_html=True)

        head_left, head_right = st.columns([3, 1])
        curr_claim_id = detect_claim_id(curr_claim)

        with head_left:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Review Details</p>", unsafe_allow_html=True)
            h_name   = get_val(curr_claim, ["Insured Name", "Name", "Claimant", "TPA_NAME"], "Unknown Entity")
            h_date   = get_val(curr_claim, ["Loss Date", "Date", "LOSS_DATE"], "N/A")
            h_status = get_val(curr_claim, ["Status", "CLAIM_STATUS"], "Submitted")
            h_total  = get_val(curr_claim, ["Total Incurred", "Incurred", "Total", "Amount", "TOTAL_INCURRED"], "$0")
            st.markdown(f"""
                <div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} — {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;text-align:right;'>Export Selection</p>", unsafe_allow_html=True)
            st.markdown("""
                <div class="export-sel-btn" style="display:flex;justify-content:flex-end;gap:6px;margin-top:2px;">
            """, unsafe_allow_html=True)
            b1, b2 = st.columns([1, 1])
            with b1:
                if st.button("✔ All", key=f"all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = True
                    st.rerun()
            with b2:
                if st.button("✘ None", key=f"none_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = False
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        if merged_meta:
            titles  = [v for v in merged_meta.values() if v["type"] == "TITLE" and v["value"]]
            headers = [v for v in merged_meta.values() if v["type"] == "HEADER" and v["value"]]
            if titles or headers:
                merge_html = "<div style='margin-bottom:10px;'>"
                for t in titles[:3]:
                    merge_html += f'<span class="merged-badge">📌 TITLE: {t["value"][:30]}</span> '
                for h in headers[:4]:
                    merge_html += f'<span class="merged-badge">⊞ HEADER: {h["value"][:20]}</span> '
                merge_html += "</div>"
                st.markdown(merge_html, unsafe_allow_html=True)

        _active_schema = st.session_state.get("active_schema", None)
        _conf_thresh   = st.session_state.get("conf_threshold", 80)

        if _active_schema and _active_schema in SCHEMAS:
            # ── SCHEMA MODE ──
            _schema_def  = SCHEMAS[_active_schema]
            _mapped      = map_claim_to_schema(curr_claim, _active_schema, title_fields)
            _custom_flds = st.session_state.get(f"custom_fields_{_active_schema}", [])

            _display_fields = list(_schema_def["required_fields"]) + [
                f for f in _custom_flds if f not in _schema_def["required_fields"]
            ]

            _low_conf = [
                sf for sf in _display_fields
                if sf in _mapped and _mapped[sf]["confidence"] < _conf_thresh
            ]
            _missing  = [sf for sf in _schema_def["required_fields"] if sf not in _mapped]

            if _missing:
                st.markdown(
                    f"<div style=\"background:#2d1515;border:1px solid #f85149;border-radius:6px;"
                    f"padding:8px 12px;margin-bottom:8px;font-size:12px;color:#f85149;\">"
                    f"⚠ {len(_missing)} required field(s) could not be mapped from this sheet: "
                    f"{', '.join(_missing)}</div>",
                    unsafe_allow_html=True
                )
            if _low_conf:
                st.markdown(
                    f"<div style=\"background:#2d2208;border:1px solid #d29922;border-radius:6px;"
                    f"padding:8px 12px;margin-bottom:8px;font-size:12px;color:#d29922;\">"
                    f"⚡ {len(_low_conf)} field(s) below confidence threshold ({_conf_thresh}%): "
                    f"{', '.join(_low_conf)}</div>",
                    unsafe_allow_html=True
                )

            hc = st.columns([1.6, 1.3, 1.6, 1.6, 0.7, 0.7, 0.5])
            with hc[0]: st.markdown("**SCHEMA FIELD**")
            with hc[1]: st.markdown("**CONF**")
            with hc[2]: st.markdown("**EXTRACTED VALUE**")
            with hc[3]: st.markdown("**MODIFIED VALUE**")

            for schema_field in _display_fields:
                if schema_field not in _mapped:
                    is_req = schema_field in _schema_def["required_fields"]
                    st.markdown(
                        f"<div style=\"display:flex;align-items:center;gap:8px;"
                        f"background:#1a0e0e;border:1px solid #f85149;border-radius:6px;"
                        f"padding:6px 10px;margin:2px 0;\">"
                        f"<span style=\"color:#f85149;font-size:12px;font-weight:bold;text-transform:uppercase;\">"
                        f"{schema_field}</span>"
                        f"<span style=\"background:#f85149;color:white;font-size:9px;border-radius:4px;padding:1px 5px;\">"
                        f"{'REQUIRED' if is_req else 'OPTIONAL'} · NOT FOUND</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )
                    continue

                m        = _mapped[schema_field]
                conf     = m["confidence"]
                excel_f  = m["excel_field"]
                info     = m["info"]
                is_req   = m["is_required"]
                # Flag whether this field came from a title row (virtual key)
                is_title_sourced = m.get("from_title", False)

                if conf < _conf_thresh:
                    conf_col   = "#f85149"
                    row_border = "#f85149"
                    row_bg     = "#1f0d0d"
                elif conf < 75:
                    conf_col   = "#f0883e"
                    row_border = "#f0883e"
                    row_bg     = "#1f1508"
                elif conf < 88:
                    conf_col   = "#d29922"
                    row_border = "#30363d"
                    row_bg     = "#161b22"
                else:
                    conf_col   = "#3fb950"
                    row_border = "#30363d"
                    row_bg     = "#161b22"

                ek = f"edit_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_schema_{schema_field}"

                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state: st.session_state[mk] = info.get("modified", info["value"])

                st.markdown(
                    f"<div style=\"border-left:3px solid {row_border};background:{row_bg};"
                    f"border-radius:0 4px 4px 0;padding:2px 0 2px 4px;margin:1px 0;\"></div>",
                    unsafe_allow_html=True
                )

                cl, cc, co, cm, ce, cb, cx = st.columns([1.6, 1.3, 1.6, 1.6, 0.7, 0.7, 0.5], gap="small")

                with cl:
                    _cur_val  = st.session_state.get(mk, info.get("modified", info["value"]))
                    _edited   = _cur_val != info["value"]
                    _dot      = "<span style=\"color:#d29922;font-size:8px;\">●</span> " if _edited else ""
                    _req_tag  = (
                        f"<span style=\"background:#1c2128;border:1px solid {conf_col};"
                        f"border-radius:3px;font-size:9px;color:{conf_col};"
                        f"padding:0 4px;margin-left:4px;\">"
                        f"{'REQ' if is_req else 'OPT'}</span>"
                    )
                    _from_title  = m.get("from_title", False)
                    _src_label   = (
                        "<span style=\"color:#d29922;font-size:9px;\">📌 from title row</span>"
                        if _from_title else
                        f"<span style=\"font-size:9px;color:#8b949e;\">← {excel_f}</span>"
                    )
                    st.markdown(
                        f"<div style=\"min-height:40px;display:flex;flex-direction:column;"
                        f"justify-content:center;color:#c9d1d9;font-size:11px;font-weight:bold;"
                        f"text-transform:uppercase;\">"
                        f"{_dot}{schema_field}{_req_tag}"
                        f"<div style=\"font-weight:normal;text-transform:none;margin-top:1px;\">{_src_label}</div>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                with cc:
                    st.markdown(
                        f"<div style=\"min-height:40px;display:flex;flex-direction:column;"
                        f"justify-content:center;gap:3px;\">"
                        f"<span style=\"background:{conf_col}22;border:1px solid {conf_col};"
                        f"border-radius:10px;padding:2px 8px;font-size:12px;"
                        f"color:{conf_col};font-weight:bold;\">{conf}%</span>"
                        f"<div style=\"background:#21262d;border-radius:3px;height:4px;width:80%;\">"
                        f"<div style=\"background:{conf_col};height:4px;border-radius:3px;"
                        f"width:{conf}%;\"></div></div>"
                        f"<span style=\"font-size:9px;color:#8b949e;\">"
                        f"H:{m['header_score']}% V:{m['value_score']}%</span>"
                        f"</div>",
                        unsafe_allow_html=True
                    )

                with co:
                    st.text_input("o", value=info["value"],
                                  key=f"orig_{selected_sheet}_{curr_claim_id}_schema_{schema_field}",
                                  label_visibility="collapsed", disabled=True)

                with cm:
                    if st.session_state[ek]:
                        with st.form(key=f"form_s_{selected_sheet}_{curr_claim_id}_{schema_field}", border=False):
                            nv = st.text_input(
                                "m", value=st.session_state.get(mk, info.get("modified", info["value"])),
                                label_visibility="collapsed"
                            )
                            submitted = st.form_submit_button("", use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv
                                # ── FIX: only write back to claim dict if the
                                #    excel_field actually exists (title-sourced
                                #    fields use virtual keys like "[title row 1]"
                                #    which are NOT keys in the claim row dict) ──
                                if not is_title_sourced and excel_f in st.session_state.sheet_cache[selected_sheet]["data"][st.session_state.selected_idx]:
                                    st.session_state.sheet_cache[selected_sheet]["data"][
                                        st.session_state.selected_idx][excel_f]["modified"] = nv
                                st.session_state[ek] = False
                                st.rerun()
                    else:
                        nv = st.text_input("m", key=mk, label_visibility="collapsed", disabled=True)

                    # ── FIX: guard the writeback — skip for title-sourced fields ──
                    if not is_title_sourced and excel_f in st.session_state.sheet_cache[selected_sheet]["data"][st.session_state.selected_idx]:
                        st.session_state.sheet_cache[selected_sheet]["data"][
                            st.session_state.selected_idx][excel_f]["modified"] = st.session_state.get(
                            mk, info.get("modified", info["value"]))

                with ce:
                    if st.button("👁", key=f"eye_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                                 use_container_width=True):
                        show_eye_popup(schema_field, info, excel_path, selected_sheet)

                with cb:
                    if not st.session_state[ek]:
                        if st.button("✏", key=f"ed_s_{selected_sheet}_{curr_claim_id}_{schema_field}",
                                     use_container_width=True, help="Edit field"):
                            st.session_state[ek] = True
                            st.rerun()
                    else:
                        st.markdown(
                            "<div style=\"height:38px;display:flex;align-items:center;"
                            "justify-content:center;color:#3fb950;font-size:11px;"
                            "border:1px solid #30363d;border-radius:6px;\">↵</div>",
                            unsafe_allow_html=True
                        )

                with cx:
                    st.markdown("<div style=\"height:8px;\"></div>", unsafe_allow_html=True)
                    st.checkbox("", key=xk, label_visibility="collapsed")

        else:
            # ── PLAIN MODE ──
            hc = st.columns([1.8, 2.4, 2.4, 0.8, 0.8, 0.4])
            with hc[0]: st.markdown("**FIELD**")
            with hc[1]: st.markdown("**EXTRACTED VALUE**")
            with hc[2]: st.markdown("**MODIFIED VALUE**")

            for field, info in curr_claim.items():
                ek = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
                xk = f"chk_{selected_sheet}_{curr_claim_id}_{field}"
                mk = f"mod_{selected_sheet}_{curr_claim_id}_{field}"

                if ek not in st.session_state: st.session_state[ek] = False
                if xk not in st.session_state: st.session_state[xk] = True
                if mk not in st.session_state: st.session_state[mk] = info.get("modified", info["value"])

                cl, co, cm, ce, cb, cx = st.columns([1.8, 2.4, 2.4, 0.8, 0.8, 0.4], gap="small")

                with cl:
                    _current_val = st.session_state.get(mk, info.get("modified", info["value"]))
                    _is_edited   = _current_val != info["value"]
                    _edit_dot    = "<span style=\"color:#d29922;margin-left:4px;font-size:8px;\">●</span>" if _is_edited else ""
                    st.markdown(
                        f"<div style=\"height:40px;display:flex;align-items:center;"
                        f"color:#c9d1d9;font-size:12px;font-weight:bold;text-transform:uppercase;\">"
                        f"{field}{_edit_dot}</div>", unsafe_allow_html=True)

                with co:
                    st.text_input("o", value=info["value"],
                                  key=f"orig_{selected_sheet}_{curr_claim_id}_{field}",
                                  label_visibility="collapsed", disabled=True)

                with cm:
                    if st.session_state[ek]:
                        with st.form(key=f"form_{selected_sheet}_{curr_claim_id}_{field}", border=False):
                            nv = st.text_input(
                                "m", value=st.session_state.get(mk, info.get("modified", info["value"])),
                                label_visibility="collapsed"
                            )
                            submitted = st.form_submit_button("", use_container_width=False)
                            if submitted:
                                st.session_state[mk] = nv
                                st.session_state.sheet_cache[selected_sheet]["data"][
                                    st.session_state.selected_idx][field]["modified"] = nv
                                st.session_state[ek] = False
                                st.rerun()
                    else:
                        nv = st.text_input("m", key=mk, label_visibility="collapsed", disabled=True)
                    st.session_state.sheet_cache[selected_sheet]["data"][
                        st.session_state.selected_idx][field]["modified"] = st.session_state.get(
                        mk, info.get("modified", info["value"]))

                with ce:
                    if st.button("👁", key=f"eye_{selected_sheet}_{curr_claim_id}_{field}",
                                 use_container_width=True):
                        show_eye_popup(field, info, excel_path, selected_sheet)

                with cb:
                    if not st.session_state[ek]:
                        if st.button("✏", key=f"ed_{selected_sheet}_{curr_claim_id}_{field}",
                                     use_container_width=True, help="Edit field"):
                            st.session_state[ek] = True
                            st.rerun()
                    else:
                        st.markdown(
                            "<div style=\"height:38px;display:flex;align-items:center;"
                            "justify-content:center;color:#3fb950;font-size:11px;"
                            "border:1px solid #30363d;border-radius:6px;\">↵</div>",
                            unsafe_allow_html=True
                        )

                with cx:
                    st.markdown("<div style=\"height:8px;\"></div>", unsafe_allow_html=True)
                    st.checkbox("", key=xk, label_visibility="collapsed")

        # Totals section
        if totals_data:
            st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)
            st.markdown("**📊 Sheet Totals**")
            agg = totals_data.get("aggregated", {})
            if agg:
                t_cols = st.columns(min(4, len(agg)))
                for idx, (k, v) in enumerate(agg.items()):
                    with t_cols[idx % len(t_cols)]:
                        st.markdown(f"""
                        <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;padding:8px 12px;margin-bottom:6px;">
                            <div style="font-size:11px;color:#8b949e;text-transform:uppercase;">{k}</div>
                            <div style="font-size:16px;font-weight:bold;color:#3fb950;">{v:,.2f}</div>
                        </div>""", unsafe_allow_html=True)

    # ── RIGHT PANEL ──────────────────────────────────────────────────────
    with col_fmt:
        st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Export Format</p>", unsafe_allow_html=True)

        _active = st.session_state.get("active_schema", None)
        if _active and _active in SCHEMAS:
            _sc = SCHEMAS[_active]
            _cf_count = len(st.session_state.get(f"custom_fields_{_active}", []))
            st.markdown(
                f"""<div style='background:#1c2128;border:1px solid {_sc['color']};
border-radius:8px;padding:10px 12px;margin-bottom:8px;'>
    <div style='font-size:13px;font-weight:bold;color:{_sc['color']};'>{_sc['icon']} {_active} Active</div>
    <div style='font-size:11px;color:#8b949e;margin-top:3px;'>{_sc['version']}</div>
    <div style='font-size:11px;color:#8b949e;margin-top:2px;'>
        Required: {len(_sc['required_fields'])} fields &nbsp;|&nbsp; Custom: {_cf_count}
    </div>
</div>""",
                unsafe_allow_html=True
            )

        _conf = st.session_state.get("conf_threshold", 80)
        _bar_col = "#3fb950" if _conf >= 70 else "#d29922" if _conf >= 40 else "#f85149"
        st.markdown(
            f"<div style='margin-bottom:10px;'>"
            f"<div style='font-size:10px;color:#8b949e;text-transform:uppercase;font-weight:bold;margin-bottom:3px;'>Confidence Threshold</div>"
            f"<div style='display:flex;align-items:center;gap:8px;'>"
            f"<div class='conf-bar-wrap' style='flex:1;'>"
            f"<div class='conf-bar-fill' style='width:{_conf}%;background:{_bar_col};'></div>"
            f"</div>"
            f"<span style='color:{_bar_col};font-size:12px;font-weight:bold;'>{_conf}%</span>"
            f"</div></div>",
            unsafe_allow_html=True
        )

        st.markdown(f"""
            <div style="background:#1c2128;border:1px solid #58a6ff;border-radius:8px;
                        padding:10px 12px;margin-bottom:4px;">
                <div style="font-size:14px;color:white;font-weight:bold;">📄 Standard JSON</div>
                <div style="font-size:11px;color:#8b949e;margin-top:3px;">Raw extracted claims data with titles, records &amp; totals in Excel order</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)

        if merged_meta:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:11px;text-transform:uppercase;margin-top:12px;'>Merged Regions</p>", unsafe_allow_html=True)
            sorted_merges = sorted(
                [(k, v) for k, v in merged_meta.items() if v["value"]],
                key=lambda x: (x[1]["row_start"], x[1]["col_start"])
            )
            for key, m in sorted_merges[:8]:
                type_color = "#58a6ff" if m["type"] == "TITLE" else "#d29922" if m["type"] == "HEADER" else "#8b949e"
                st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;
                                padding:6px 10px;margin-bottom:4px;">
                        <div style="font-size:10px;color:{type_color};">{m['type']} · R{m['row_start']}C{m['col_start']}→R{m['row_end']}C{m['col_end']}</div>
                        <div style="font-size:12px;color:#c9d1d9;margin-top:2px;">{m['value'][:35]}</div>
                    </div>""", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        _exp_schema = st.session_state.get("active_schema", None)
        _sheet_meta = {"sheet_name": selected_sheet, "record_count": len(data)}

        def _sanitize_for_json(obj):
            if isinstance(obj, dict):
                return {k: _sanitize_for_json(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [_sanitize_for_json(i) for i in obj]
            if isinstance(obj, str):
                return normalize_str(obj)
            return obj

        st.markdown(
            "<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
            "font-weight:bold;margin-bottom:4px;'>Standard</div>",
            unsafe_allow_html=True
        )
        std_j, std_y = st.columns(2)

        with std_j:
            if st.button("JSON", use_container_width=True, type="primary",
                         key=f"export_std_json_{selected_sheet}"):
                _std_export_data = {}
                for i, row in enumerate(data):
                    c_id = detect_claim_id(row, i)
                    rec  = {}
                    for fld, inf in row.items():
                        if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}", True):
                            mk_key    = f"mod_{selected_sheet}_{c_id}_{fld}"
                            live_val  = st.session_state.get(mk_key, None)
                            orig      = inf.get("value", "")
                            final_val = live_val if live_val is not None else inf.get("modified", orig)
                            rec[fld]  = {
                                "value": final_val, "original": orig,
                                "edited": final_val != orig,
                                "excel_row": inf.get("excel_row"),
                                "excel_col": inf.get("excel_col"),
                                "record_index": i,
                            }
                    _std_export_data[c_id] = clean_duplicate_fields(rec)
                output   = _sanitize_for_json(
                    to_standard_json(_std_export_data, _sheet_meta, totals_data, merged_meta)
                )
                json_str = json.dumps(output, indent=2, ensure_ascii=False)
                save_feature_store(selected_sheet, output)
                st.success("✅ Ready!")
                st.download_button(
                    "📥 Download JSON", data=json_str,
                    file_name=f"{selected_sheet}_standard.json",
                    mime="application/json", use_container_width=True,
                    key=f"dl_std_json_{selected_sheet}"
                )

        with std_y:
            st.markdown(
                "<div style=\"background:#161b22;border:1px solid #30363d;border-radius:6px;"
                "padding:8px 10px;font-size:11px;color:#8b949e;text-align:center;\">"
                "YAML available<br>as config only<br><span style=\"font-size:9px;\">see /config folder</span>"
                "</div>",
                unsafe_allow_html=True
            )

        if _exp_schema == "Duck Creek":
            st.markdown(
                "<div style='background:#1a1208;border:1px solid #f0883e;border-radius:6px;"
                "padding:6px 10px;margin:8px 0 4px 0;font-size:11px;color:#f0883e;font-weight:bold;'>"
                "🟠 Duck Creek</div>",
                unsafe_allow_html=True
            )
            st.markdown(
                "<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
                "font-weight:bold;margin-bottom:4px;'>API submission (JSON / XML)</div>",
                unsafe_allow_html=True
            )
            dc_j, dc_x = st.columns(2)
            with dc_j:
                if st.button("JSON", use_container_width=True,
                             key=f"export_dc_json_{selected_sheet}"):
                    recs     = build_mapped_records_for_export(data, "Duck Creek", selected_sheet)
                    dc_json  = _sanitize_for_json(to_duck_creek_json(recs, _sheet_meta))
                    json_str = json.dumps(dc_json, indent=2, ensure_ascii=False)
                    save_feature_store(selected_sheet, dc_json)
                    st.success("✅ Ready!")
                    st.download_button(
                        "📥 Download JSON", data=json_str,
                        file_name=f"{selected_sheet}_DuckCreek.json",
                        mime="application/json", use_container_width=True,
                        key=f"dl_dc_json_{selected_sheet}"
                    )
            with dc_x:
                if st.button("XML", use_container_width=True,
                             key=f"export_dc_xml_{selected_sheet}"):
                    recs    = build_mapped_records_for_export(data, "Duck Creek", selected_sheet)
                    xml_out = to_duck_creek_xml(recs, _sheet_meta)
                    save_feature_store(selected_sheet, {"format": "DuckCreek_XML", "count": len(recs)})
                    st.success("✅ Ready!")
                    st.download_button(
                        "📥 Download XML", data=xml_out,
                        file_name=f"{selected_sheet}_DuckCreek.xml",
                        mime="application/xml", use_container_width=True,
                        key=f"dl_dc_xml_{selected_sheet}"
                    )

            st.markdown(
                "<div style=\"background:#161b22;border:1px solid #30363d;border-radius:6px;"
                "padding:6px 10px;margin-top:6px;font-size:10px;color:#8b949e;\">"
                "📁 Schema config: <code>config/duck_creek.yaml</code>"
                "</div>",
                unsafe_allow_html=True
            )

        elif _exp_schema == "Guidewire":
            st.markdown(
                "<div style='background:#0e1a2e;border:1px solid #58a6ff;border-radius:6px;"
                "padding:6px 10px;margin:8px 0 4px 0;font-size:11px;color:#58a6ff;font-weight:bold;'>"
                "🔵 Guidewire ClaimCenter</div>",
                unsafe_allow_html=True
            )
            st.markdown(
                "<div style='font-size:10px;color:#8b949e;text-transform:uppercase;"
                "font-weight:bold;margin-bottom:4px;'>API submission (JSON)</div>",
                unsafe_allow_html=True
            )
            if st.button("JSON", use_container_width=True, type="primary",
                         key=f"export_gw_json_{selected_sheet}"):
                recs     = build_mapped_records_for_export(data, "Guidewire", selected_sheet)
                gw_json  = _sanitize_for_json(to_guidewire_json(recs, _sheet_meta))
                json_str = json.dumps(gw_json, indent=2, ensure_ascii=False)
                save_feature_store(selected_sheet, gw_json)
                st.success("✅ Ready!")
                st.download_button(
                    "📥 Download JSON", data=json_str,
                    file_name=f"{selected_sheet}_Guidewire_ClaimCenter.json",
                    mime="application/json", use_container_width=True,
                    key=f"dl_gw_json_{selected_sheet}"
                )

            st.markdown(
                "<div style=\"background:#161b22;border:1px solid #30363d;border-radius:6px;"
                "padding:6px 10px;margin-top:6px;font-size:10px;color:#8b949e;\">"
                "📁 Schema config: <code>config/guidewire.yaml</code>"
                "</div>",
                unsafe_allow_html=True
            )
